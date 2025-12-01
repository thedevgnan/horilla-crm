import base64
import csv
import functools
import importlib
import inspect
import json
import logging
import re
from decimal import Decimal, InvalidOperation
from functools import cached_property, reduce
from io import BytesIO
from operator import or_
from typing import Any
from urllib.parse import parse_qs, urlencode, urlparse, urlunparse

from auditlog.models import LogEntry
from django import forms
from django.apps import apps
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import redirect_to_login
from django.contrib.contenttypes.fields import GenericRelation
from django.contrib.contenttypes.models import ContentType
from django.core.cache import cache
from django.core.exceptions import (
    FieldDoesNotExist,
    FieldError,
    ImproperlyConfigured,
    ObjectDoesNotExist,
    ValidationError,
)
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import IntegrityError, models, transaction
from django.db.models import Case, ForeignKey, Max, Q, When
from django.db.models.fields.related import ForeignKey, ManyToManyField
from django.forms import ValidationError
from django.http import Http404, HttpResponse, QueryDict
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import resolve, reverse, reverse_lazy
from django.utils import timezone, translation
from django.utils.dateparse import parse_date, parse_datetime
from django.utils.decorators import method_decorator
from django.utils.translation import gettext_lazy as _
from django.views.generic import (
    DeleteView,
    DetailView,
    FormView,
    ListView,
    TemplateView,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from horilla.exceptions import HorillaHttp404
from horilla_core.decorators import htmx_required, permission_required_or_denied
from horilla_core.mixins import OwnerQuerysetMixin
from horilla_core.models import (
    ActiveTab,
    HorillaAttachment,
    KanbanGroupBy,
    ListColumnVisibility,
    PinnedView,
    RecentlyViewed,
    RecycleBin,
)
from horilla_core.utils import get_field_permissions_for_model
from horilla_generics.forms import (
    HorillaAttachmentForm,
    HorillaHistoryForm,
    HorillaModelForm,
    HorillaMultiStepForm,
)
from horilla_utils.methods import closest_numbers, get_section_info_for_model
from horilla_utils.middlewares import _thread_local

logger = logging.getLogger(__name__)


class HorillaView(TemplateView):
    """
    A generic class-based view for rendering templates with context data.
    """

    template_name = "base.html"
    list_url: str = ""
    kanban_url: str = ""
    nav_url: str = ""

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        filter_form = self.request.headers.get("HX-Trigger")
        if filter_form == "filter-form":
            context["filter_form"] = filter_form
        context["nav_url"] = self.nav_url
        context["list_url"] = self.list_url
        context["kanban_url"] = self.kanban_url
        return context


class HorillaNavView(TemplateView):

    template_name = "navbar.html"
    nav_title: str = ""
    filterset_class = None
    search_url: str = ""
    main_url: str = ""
    kanban_url: str = ""
    actions: list = []
    new_button: dict = None
    second_button: dict = None
    model_name: str = ""
    model_app_label: str = ""
    custom_view_type: dict = {}
    nav_width = True
    recently_viewed_option = True
    all_view_types = True
    filter_option = True
    one_view_only = False
    reload_option = True
    border_enabled = True
    search_option = True
    navbar_indication = False
    gap_enabled = True
    navbar_indication_attrs: dict = {}
    exclude_kanban_fields: str = ""
    enable_actions = False

    def get_navbar_indication_attrs(self):
        if self.navbar_indication:
            return self.navbar_indication_attrs

    def get_default_view_type(self):
        """Return the pinned view_type if available, else 'all'."""
        pinned_view = PinnedView.all_objects.filter(
            user=self.request.user, model_name=self.model_name
        ).first()
        return pinned_view.view_type if pinned_view else "all"

    def show_list_only(self):
        """Check if kanban should be hidden based on current view type."""
        current_view_type = (
            self.request.GET.get("view_type") or self.get_default_view_type()
        )

        # Check if current view type has hide_kanban setting
        for view_key, view_config in self.custom_view_type.items():
            if view_key == current_view_type:
                if isinstance(view_config, dict):
                    return view_config.get("show_list_only", False)
                break
        return False

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["nav_title"] = self.nav_title
        context["search_url"] = self.search_url or self.request.path
        context["main_url"] = self.main_url or self.request.path
        context["kanban_url"] = self.kanban_url
        context["actions"] = self.actions
        context["new_button"] = self.new_button or {}
        context["second_button"] = self.second_button or {}
        context["model_name"] = self.model_name
        context["model_app_label"] = self.model_app_label
        context["nav_width"] = self.nav_width
        context["view_type"] = (
            self.request.GET.get("view_type") or self.get_default_view_type()
        )
        context["show_list_only"] = self.show_list_only()
        context["custom_view_type"] = self.custom_view_type
        context["pinned_view"] = PinnedView.all_objects.filter(
            user=self.request.user, model_name=self.model_name
        ).first()
        context["recently_viewed_option"] = self.recently_viewed_option
        context["all_view_types"] = self.all_view_types
        context["filter_option"] = self.filter_option
        context["one_view_only"] = self.one_view_only
        context["reload_option"] = self.reload_option
        context["search_option"] = self.search_option
        context["border_enabled"] = self.border_enabled
        context["navbar_indication"] = self.navbar_indication
        context["gap_enabled"] = self.gap_enabled
        context["enable_actions"] = self.enable_actions
        context["navbar_indication_attrs"] = self.get_navbar_indication_attrs()
        return context

    @cached_property
    def actions(self):
        """Actions for lead"""
        view_perm = f"{self.model_app_label}.view_{self.model_name.lower()}"
        view_own_perm = f"{self.model_app_label}.view_own_{self.model_name.lower()}"
        can_import_perm = f"{self.model_app_label}.can_import_{self.model_name.lower()}"
        resolved = resolve(str(self.search_url))
        single_import = True
        url_name = resolved.url_name

        actions = []
        if self.request.user.has_perm(view_perm) or self.request.user.has_perm(
            view_own_perm
        ):
            if self.request.user.has_perm(can_import_perm):
                actions.append(
                    {
                        "action": "Import",
                        "attrs": f"""
                        hx-get="{reverse_lazy('horilla_core:import_data')}?single_import={str(single_import).lower()}&model_name={self.model_name}&app_label={self.model_app_label}"
                        onclick="openModal()"
                        hx-target="#modalBox"
                        hx-swap="innerHTML"
                        """,
                    }
                )

            actions.extend(
                [
                    {
                        "action": "Kanban Settings",
                        "attrs": f"""
                        hx-get="{reverse_lazy('horilla_generics:create_kanban_group')}?model={self.model_name}&app_label={self.model_app_label}&exclude_fields={self.exclude_kanban_fields}"
                        onclick="openModal()"
                        hx-target="#modalBox"
                        hx-swap="innerHTML"
                        """,
                    },
                    {
                        "action": "Add column to list",
                        "attrs": f"""
                        hx-get="{reverse_lazy('horilla_generics:column_selector')}?app_label={self.model_app_label}&model_name={self.model_name}&url_name={url_name}"
                        onclick="openModal()"
                        hx-target="#modalBox"
                        hx-swap="innerHTML"
                        """,
                    },
                ]
            )
        return actions


class HorillaListView(ListView):
    """
    A customizable ListView that provides automatic column generation,
    filtering, searching, and action handling with infinite scrolling pagination.
    """

    template_name = "list_view.html"
    context_object_name = "queryset"
    columns = []
    actions = []
    view_id = ""
    action_method = ""
    exclude_columns = []
    default_sort_field = None
    default_sort_direction = "asc"
    sort_by_mapping = []
    paginate_by = 100
    page_kwarg = "page"
    main_url: str = ""
    search_url: str = ""
    filterset_class = None
    filter_url_push = True
    max_visible_actions = 4
    bulk_update_fields = []
    bulk_update_two_column = False
    raw_attrs: list = []
    number_of_recent_view = 20
    bulk_delete_enabled = True
    header_attrs = []
    col_attrs = []
    clear_session_button_enabled = True
    bulk_select_option = True
    no_record_section = True
    no_record_add_button: dict = None
    no_record_msg: str = None
    table_width = True
    table_class = True
    table_height_as_class = ""
    table_height = True
    bulk_update_option = True
    store_ordered_ids = False
    save_to_list_option = True
    enable_sorting = True
    custom_bulk_actions = []
    bulk_export_option = True
    additional_action_button = []
    list_column_visibility = True
    owner_filtration = True
    sorting_target = None
    exclude_columns_from_sorting = []

    def __init__(self, **kwargs):
        self._model_fields_cache = None
        super().__init__(**kwargs)
        if self.store_ordered_ids:
            self.ordered_ids_key = f"ordered_ids_{self.model.__name__.lower()}"
        self.kwargs = kwargs

        if self.columns:
            resolved_columns = []
            instance = self.model()
            for col in self.columns:
                if isinstance(col, (tuple, list)) and len(col) >= 2:
                    resolved_columns.append((str(col[0]), str(col[1])))
                elif isinstance(col, str):
                    try:
                        field = instance._meta.get_field(col)
                        verbose_name = str(field.verbose_name)
                        resolved_columns.append((verbose_name, col))
                    except Exception:
                        resolved_columns.append((col.replace("_", " ").title(), col))
                else:
                    resolved_columns.append((str(col), str(col)))

            self.columns = resolved_columns

    def get_default_view_type(self):
        """Return the pinned view_type if available, else 'all'."""
        pinned_view = PinnedView.all_objects.filter(
            user=self.request.user, model_name=self.model.__name__
        ).first()
        return pinned_view.view_type if pinned_view else "all"

    def get_queryset(self):
        """Get filtered queryset based on search, filter, or view type parameters."""

        queryset = super().get_queryset()
        view_type = self.request.GET.get("view_type") or self.get_default_view_type()

        is_bulk_operation = (
            (
                self.request.method == "POST"
                and self.request.POST.get("action")
                in [
                    "bulk_delete",
                    "delete_item_with_dependencies",
                    "delete_all_dependencies",
                ]
            )
            or self.request.POST.get("bulk_delete_form") == "true"
            or self.request.POST.get("soft_delete_form") == "true"
        )

        if is_bulk_operation:
            view_type = self.request.GET.get("view_type")
            if not view_type:
                view_type = "all"
        else:
            view_type = (
                self.request.GET.get("view_type") or self.get_default_view_type()
            )

        if view_type == "recently_viewed":
            recently_viewed_items = RecentlyViewed.objects.get_recently_viewed(
                user=self.request.user, model_class=self.model
            )
            pks = [item.pk for item in recently_viewed_items if item]
            queryset = queryset.filter(pk__in=pks)

        elif view_type in ("recently_created", "recently_modified"):
            sort_field = (
                "-created_at" if view_type == "recently_created" else "-updated_at"
            )
            recent_queryset = queryset.order_by(sort_field)[:20]
            recent_ids = list(recent_queryset.values_list("pk", flat=True))
            queryset = queryset.filter(pk__in=recent_ids)

        elif view_type.startswith("saved_list_"):
            saved_list_id = view_type.replace("saved_list_", "")
            try:
                saved_list = self.request.user.saved_filter_lists.filter(
                    id=saved_list_id
                ).first()
                filter_params = saved_list.get_filter_params()
                merged_params = QueryDict(mutable=True)
                for key, values in filter_params.items():
                    for value in values:
                        merged_params.appendlist(key, value)

                search_keys = [
                    "field",
                    "operator",
                    "value",
                    "start_value",
                    "end_value",
                    "search",
                ]
                for key, values in self.request.GET.lists():
                    if key in search_keys:
                        for value in values:
                            merged_params.appendlist(key, value)

                if self.filterset_class:
                    self.filterset = self.filterset_class(
                        merged_params, queryset=queryset, request=self.request
                    )
                    queryset = self.filterset.filter_queryset(queryset)
            except Exception:
                pass

        if self.filterset_class and not (
            view_type.startswith("saved_list_") and getattr(self, "filterset", None)
        ):
            self.filterset = self.filterset_class(
                self.request.GET, queryset=queryset, request=self.request
            )
            queryset = self.filterset.filter_queryset(queryset)

        sort_field = self.request.GET.get("sort")
        sort_direction = self.request.GET.get("direction", self.default_sort_direction)

        if view_type == "recently_viewed" and not sort_field and "pks" in locals():
            preserved_order = Case(
                *[When(pk=pk, then=pos) for pos, pk in enumerate(pks)]
            )
            queryset = queryset.order_by(preserved_order)
        elif sort_field:
            queryset = self._apply_sorting(queryset, sort_field, sort_direction)
        elif view_type == "recently_created":
            queryset = queryset.order_by("-created_at")
        elif view_type == "recently_modified":
            queryset = queryset.order_by("-updated_at")
        elif self.default_sort_field:
            # Use default_sort_field if specified in child class
            order_prefix = "-" if self.default_sort_direction == "desc" else ""
            queryset = queryset.order_by(f"{order_prefix}{self.default_sort_field}")
        else:
            queryset = queryset.order_by("-id")
        # elif sort_field:
        #     queryset = self._apply_sorting(queryset, sort_field, sort_direction)
        # elif view_type == "recently_created":
        #     queryset = queryset.order_by("-created_at")
        # elif view_type == "recently_modified":
        #     queryset = queryset.order_by("-updated_at")
        # else:
        #     queryset = queryset.order_by("-id")

        if self.store_ordered_ids:
            ordered_ids = list(queryset.values_list("pk", flat=True))
            self.request.session[self.ordered_ids_key] = ordered_ids

        if self.owner_filtration:
            user = self.request.user
            app_label = self.model._meta.app_label
            model_name = self.model._meta.model_name
            view_perm = f"{app_label}.view_{model_name}"
            view_own_perm = f"{app_label}.view_own_{model_name}"

            if user.has_perm(view_perm):
                return queryset

            if user.has_perm(view_own_perm):
                owner_fields = getattr(self.model, "OWNER_FIELDS", None)

                if owner_fields:
                    query = reduce(
                        or_,
                        (Q(**{field_name: user}) for field_name in owner_fields),
                        Q(),
                    )
                    return queryset.filter(query).distinct()

            return queryset.none()
        return queryset.distinct()

    def _get_columns(self):
        """Get columns configuration based on model fields and methods."""

        if not self.list_column_visibility:
            return [[col[0], col[1]] for col in self.columns] if self.columns else []
        app_label = self.model._meta.app_label
        model_name = self.model.__name__
        context = (
            urlparse(self.request.META.get("HTTP_REFERER", ""))
            .path.strip("/")
            .replace("/", "_")
        )
        context = re.sub(r"_\d+$", "", context)
        current_path = resolve(self.request.path_info).url_name
        cache_key = f"visible_columns_{self.request.user.id}_{app_label}_{model_name}_{context}_{current_path}"
        cached_columns = cache.get(cache_key)
        if cached_columns:
            return cached_columns

        visibility = ListColumnVisibility.all_objects.filter(
            user=self.request.user,
            model_name=model_name,
            app_label=app_label,
            context=context,
            url_name=current_path,
        ).first()

        if visibility:
            visible_fields = visibility.visible_fields
            model_fields = self.model._meta.get_fields()
            field_mapping = {}
            for field in model_fields:
                if hasattr(field, "verbose_name") and not field.is_relation:
                    field_mapping[str(field.verbose_name)] = field.name

            columns = []
            for visible_field in visible_fields:
                if isinstance(visible_field, list) and len(visible_field) >= 2:
                    columns.append([visible_field[0], visible_field[1]])
                else:
                    verbose_name = visible_field
                    for col_verbose_name, col_field_name in self.columns:
                        if str(col_verbose_name) == verbose_name:
                            columns.append([col_field_name, verbose_name])
                            break
                    else:
                        if verbose_name in field_mapping:
                            field_name = field_mapping[verbose_name]
                            field = next(
                                (f for f in model_fields if f.name == field_name), None
                            )
                            if field and getattr(field, "choices", None):
                                columns.append(
                                    [verbose_name, f"get_{field_name}_display"]
                                )
                            else:
                                columns.append([verbose_name, field_name])
                        else:
                            display_name = str(verbose_name.replace("_", " ").title())
                            columns.append(
                                [
                                    display_name,
                                    verbose_name.lower().replace(" ", "_"),
                                ]
                            )
            cache.set(cache_key, columns)
            return columns

        elif self.columns:
            with translation.override("en"):
                serializable_columns = []
                for col in self.columns:
                    if isinstance(col, (list, tuple)) and len(col) >= 2:
                        serializable_columns.append([str(col[0]), str(col[1])])
                    else:
                        serializable_columns.append([str(col[0]) if col else "", ""])

                visibility = ListColumnVisibility.all_objects.create(
                    user=self.request.user,
                    app_label=self.model._meta.app_label,
                    model_name=self.model.__name__,
                    visible_fields=serializable_columns,
                    context=context,
                    url_name=current_path,
                )
                columns = [[col[0], col[1]] for col in serializable_columns]
                cache.set(cache_key, columns)
                return columns

        auto_columns = []
        for field in self.model._meta.fields:
            if (
                not field.auto_created
                and field.name != "id"
                and field.name not in self.exclude_columns
            ):
                verbose = str(field.verbose_name)
                auto_columns.append(
                    [
                        verbose,
                        field.name,
                    ]
                )
        return auto_columns

    def _apply_sorting(self, queryset, field, direction):
        """Fast sorting: uses DB fields or mapped aliases only."""

        if not field:
            return queryset

        if field.startswith("get_") and field.endswith("_display"):
            field = field[4:-8]

        mapped_field = next(
            (
                item[1]
                for item in getattr(self, "sort_by_mapping", [])
                if item[0] == field
            ),
            field,
        )
        model_class = queryset.model
        if not hasattr(model_class, mapped_field):
            return queryset

        attr = getattr(model_class, mapped_field)
        if callable(attr) or isinstance(attr, property):
            return queryset

        # Check if the field is a GenericForeignKey
        try:
            field_obj = model_class._meta.get_field(mapped_field)
            from django.contrib.contenttypes.fields import GenericForeignKey

            if isinstance(field_obj, GenericForeignKey):
                # Sort by content_type_id and then object_id
                ct_field = field_obj.ct_field + "_id"  # Usually 'content_type_id'
                fk_field = field_obj.fk_field  # Usually 'object_id'

                if direction == "desc":
                    return queryset.order_by(f"-{ct_field}", f"-{fk_field}")
                else:
                    return queryset.order_by(ct_field, fk_field)
        except Exception:
            pass

        order_field = f"-{mapped_field}" if direction == "desc" else mapped_field

        try:
            return queryset.order_by(order_field)
        except Exception as e:
            import logging

            logger = logging.getLogger(__name__)
            logger.warning(f"Could not sort by field '{mapped_field}': {str(e)}")
            return queryset

    def render_to_response(self, context, **response_kwargs):
        """Override to handle different types of requests appropriately."""
        is_htmx = self.request.headers.get("HX-Request") == "true"
        context["request_params"] = self.request.GET.copy()

        if is_htmx:
            return render(self.request, "list_view.html", context)

        return super().render_to_response(context, **response_kwargs)

    def _get_model_fields(self):
        """Extract model fields with metadata for filtering UI."""
        if self._model_fields_cache is not None:
            return self._model_fields_cache

        FIELD_TYPE_MAP = {
            "CharField": "text",
            "TextField": "text",
            "BooleanField": "boolean",
            "IntegerField": "number",
            "FloatField": "float",
            "DecimalField": "decimal",
            "ForeignKey": "foreignkey",
            "DateField": "date",
            "DateTimeField": "datetime",
        }

        BOOLEAN_CHOICES = [
            {"value": "True", "label": "Yes"},
            {"value": "False", "label": "No"},
        ]
        if self.filterset_class:
            exclude_fields = getattr(self.filterset_class.Meta, "exclude", [])
        model_fields = []
        is_bulk_update_trigger = False
        trigger_name = self.request.headers.get("Hx-Trigger-Name")
        is_operator_trigger = trigger_name == "operator"
        trigger = self.request.GET.get("hx_trigger")
        is_filter_form_trigger = trigger == "filter-form"
        bulk_update_trigger_value = self.request.META.get("HTTP_HX_TRIGGER")
        if bulk_update_trigger_value:
            if bulk_update_trigger_value.startswith("bulk-update-btn"):
                is_bulk_update_trigger = True

        has_filterset = bool(self.filterset_class)
        value_field = self.request.GET.get("value", "")
        for field in self.model._meta.fields:

            if field.auto_created or field.name == "id":
                continue
            if self.filterset_class:
                if field.name in exclude_fields:
                    continue

            field_class_name = field.__class__.__name__
            choices = []
            related_model = None
            related_app_label = None
            related_model_name = None

            # Handle field type determination with early returns
            if field.choices:
                field_type = "choice"
                choices = [
                    {"value": val, "label": label} for val, label in field.choices
                ]
            elif field_class_name == "ForeignKey":
                related_model = field.related_model
                related_app_label = related_model._meta.app_label
                related_model_name = related_model.__name__

                field_type = "foreignkey"
                if (
                    is_operator_trigger
                    or is_bulk_update_trigger
                    or is_filter_form_trigger
                    or value_field
                ):
                    related_objects = field.related_model.objects.all().order_by("id")
                    paginator = Paginator(related_objects, 10)

                    try:
                        paginated_objects = paginator.page(1)
                    except PageNotAnInteger:
                        paginated_objects = paginator.page(1)
                    except EmptyPage:
                        paginated_objects = paginator.page(paginator.num_pages)
                    choices = [
                        {"value": str(obj.pk), "label": str(obj)}
                        for obj in paginated_objects
                    ]
                    if value_field:
                        try:
                            value_obj = field.related_model.objects.get(pk=value_field)
                            value_choice = {
                                "value": str(value_obj.pk),
                                "label": str(value_obj),
                            }
                            if value_choice not in choices:
                                choices.append(value_choice)
                        except (field.related_model.DoesNotExist, ValueError):
                            # Ignore invalid value_field (e.g., non-existent PK or invalid format)
                            pass
            elif field_class_name == "DateTimeField":
                field_type = "datetime"
            elif field_class_name == "DateField":
                field_type = "date"
            else:
                field_type = FIELD_TYPE_MAP.get(field_class_name, "other")
                if field_type == "boolean":
                    choices = BOOLEAN_CHOICES

            operators = []
            if has_filterset:
                operators = self.filterset_class.get_operators_for_field(field_type)

            model_fields.append(
                {
                    "name": field.name,
                    "type": field_type,
                    "verbose_name": field.verbose_name,
                    "choices": choices,
                    "operators": operators,
                    "model": related_model_name,
                    "app_label": related_app_label,
                }
            )

        property_labels = getattr(self.model, "PROPERTY_LABELS", None)
        if not property_labels:
            property_labels = {
                name.replace("get_", "", 1): name.replace("get_", "", 1)
                .replace("_", " ")
                .title()
                for name, member in inspect.getmembers(
                    self.model, predicate=lambda x: isinstance(x, property)
                )
            }

        for name, member in inspect.getmembers(
            self.model, predicate=lambda x: isinstance(x, property)
        ):
            label_key = name.replace("get_", "", 1) if name.startswith("get_") else name
            if label_key in property_labels:
                model_fields.append(
                    {
                        "name": name,
                        "type": "text",
                        "verbose_name": property_labels[label_key],
                        "choices": [],
                        "operators": [],
                        "is_property": True,
                    }
                )

        self._model_fields_cache = model_fields
        return model_fields

    def handle_field_change(self, request, field_name, row_id):
        """Handle field change to update operators dropdown."""
        # Get field metadata from cache
        field_info = next(
            (
                field
                for field in self._get_model_fields()
                if field["name"] == field_name
            ),
            None,
        )

        if not field_info:
            return HttpResponse("Field not found", status=404)

        # Get appropriate operators for field type
        field_type = field_info["type"]

        operators = self.filterset_class.get_operators_for_field(field_type)

        context = {
            "operators": operators,
            "field_name": field_name,
            "row_id": row_id,
            "search_url": self.search_url,
        }

        return render(request, "partials/operator_select.html", context)

    def handle_operator_change(self, request, field_name, operator, row_id):
        """Handle operator change to update value field."""
        # Get field metadata from cache
        field_info = next(
            (
                field
                for field in self._get_model_fields()
                if field["name"] == field_name
            ),
            None,
        )

        if not field_info:
            return HttpResponse("Field not found", status=404)

        context = {"field_info": field_info, "operator": operator, "row_id": row_id}

        return render(request, "partials/value_field.html", context)

    def _check_dependencies(self, record_ids):
        """
        Check for dependencies in related models for the given record IDs.
        Returns two lists: records that cannot be deleted (with dependencies) and records that can be deleted.
        """
        import time

        from django.db.models import Prefetch

        can_delete = []
        cannot_delete = []

        # Start timer for initial queryset
        str_fields = ["id"]
        queryset = self.model.objects.filter(id__in=record_ids).only(*str_fields)

        related_objects = self.model._meta.related_objects
        if not related_objects:
            for obj in queryset:
                can_delete.append({"id": obj.id, "name": str(obj)})
            return (cannot_delete, can_delete, {})

        prefetch_queries = []

        # Build prefetch queries
        for related in related_objects:
            related_model = related.related_model
            related_name = related.get_accessor_name()
            if related_name:
                manager = getattr(
                    related_model,
                    "objects",
                    getattr(related_model, "all_objects", None),
                )
                if manager is None:
                    raise AttributeError(
                        f"No manager ('objects' or 'all_objects') defined for {related_model.__name__}"
                    )
                prefetch_queries.append(
                    Prefetch(
                        related_name,
                        queryset=manager.all()[:10],
                        to_attr=f"prefetched_{related_name}",
                    )
                )

        # Execute prefetch
        try:
            queryset = queryset.prefetch_related(*prefetch_queries)
        except AttributeError as e:
            raise AttributeError(
                f"Invalid prefetch_related lookup. Check related_name for {self.model.__name__} relations."
            )

        # Loop over objects to collect dependencies
        for obj in queryset:
            dependencies = []
            for related in related_objects:
                related_model = related.related_model
                related_name = related.get_accessor_name()
                if related_name:
                    related_records = getattr(obj, f"prefetched_{related_name}", [])
                    if related_records:
                        # related_count = len(related_records)  # Faster: uses prefetched data
                        dependencies.append(
                            {
                                "model_name": related_model._meta.verbose_name_plural,
                                "count": len(related_records),
                                "records": [str(rec) for rec in related_records],
                            }
                        )

            if dependencies:
                cannot_delete.append(
                    {"id": obj.id, "name": str(obj), "dependencies": dependencies}
                )
            else:
                can_delete.append({"id": obj.id, "name": str(obj)})

        # Final dependency summary
        dependency_details = {
            item["id"]: item["dependencies"] for item in cannot_delete
        }

        return cannot_delete, can_delete, dependency_details

    def _delete_all_dependencies(self, item_id, selected_data):
        """
        Hard delete all dependencies of a single record, not the record itself.
        Returns the updated context for rendering the modal with remaining dependencies.
        """
        try:
            if isinstance(selected_data, int):
                selected_data = [selected_data]
            elif not isinstance(selected_data, (list, tuple)):
                selected_data = []

            # Ensure the current item is included in selected_data
            if item_id not in selected_data:
                selected_data.append(item_id)

            # Fetch the record
            record = self.model.objects.get(id=item_id)
            related_objects = self.model._meta.related_objects

            # Hard delete ALL dependencies
            total_deleted_count = 0
            deleted_models = []

            for related in related_objects:
                related_model = related.related_model
                field_name = related.field.name
                filter_kwargs = {field_name: record}
                manager = getattr(
                    related_model,
                    "objects",
                    getattr(related_model, "all_objects", None),
                )
                if manager is None:
                    raise AttributeError(
                        f"No manager ('objects' or 'all_objects') defined for {related_model.__name__}"
                    )
                dependent_records = manager.filter(**filter_kwargs)
                count = dependent_records.count()

                if count > 0:
                    for dep_record in dependent_records:
                        dep_record.delete()
                    total_deleted_count += count
                    deleted_models.append(
                        f"{related_model._meta.verbose_name_plural} ({count})"
                    )

            # Recalculate dependencies for ALL items
            cannot_delete, can_delete, dependency_details = self._check_dependencies(
                selected_data
            )

            # Calculate how many main records can now be deleted
            can_delete_count = len(can_delete)

            # Create success message
            if deleted_models:
                deleted_summary = ", ".join(deleted_models)
                success_message = f"Successfully hard deleted {total_deleted_count} dependencies: {deleted_summary}"
            else:
                success_message = f"No dependencies found for '{record}'"

            # Prepare context for rendering the updated modal
            self.object_list = self.get_queryset()
            context = self.get_context_data()
            context.update(
                {
                    "selected_ids": selected_data,
                    "selected_ids_json": json.dumps(selected_data),
                    "cannot_delete": cannot_delete,
                    "can_delete": can_delete,
                    "cannot_delete_count": len(cannot_delete),
                    "can_delete_count": can_delete_count,
                    "success_message": success_message,
                    "model_verbose_name": self.model._meta.verbose_name_plural,
                }
            )

            return context

        except self.model.DoesNotExist:
            logger.error(f"Record with ID {item_id} does not exist.")
            self.object_list = self.get_queryset()
            context = self.get_context_data()
            context.update(
                {
                    "selected_ids": selected_data,
                    "selected_ids_json": json.dumps(selected_data),
                    "error_message": f"Record with ID {item_id} does not exist.",
                }
            )
            return context
        except Exception as e:
            logger.error(f"Hard delete of all dependencies failed: {str(e)}")
            self.object_list = self.get_queryset()
            context = self.get_context_data()
            context.update(
                {
                    "selected_ids": selected_data,
                    "selected_ids_json": json.dumps(selected_data),
                    "error_message": f"Hard delete of all dependencies failed: {str(e)}",
                }
            )
            return context

    def _delete_item_with_dependencies(self, item_id, record_ids, selected_data):
        """
        Hard delete only the specified dependency of a single record, not the record itself.
        Returns the updated context for rendering the modal with remaining dependencies.
        """
        try:
            if isinstance(selected_data, int):
                selected_data = [selected_data]
            elif not isinstance(selected_data, (list, tuple)):
                selected_data = []

            # Ensure the current item is included in selected_data
            if item_id not in selected_data:
                selected_data.append(item_id)

            # Fetch the record
            record = self.model.objects.get(id=item_id)
            dep_model_name = self.request.POST.get("dep_model_name")
            related_objects = self.model._meta.related_objects

            # Hard delete only the specified dependency
            deleted_count = 0
            for related in related_objects:
                related_model = related.related_model
                if related_model._meta.verbose_name_plural == dep_model_name:
                    field_name = related.field.name
                    filter_kwargs = {field_name: record}
                    manager = getattr(
                        related_model,
                        "objects",
                        getattr(related_model, "all_objects", None),
                    )
                    if manager is None:
                        raise AttributeError(
                            f"No manager ('objects' or 'all_objects') defined for {related_model.__name__}"
                        )
                    dependent_records = manager.filter(**filter_kwargs)

                    for dep_record in dependent_records:
                        dep_record.delete()
                        deleted_count += 1

            # Recalculate dependencies for ALL items including the current one
            cannot_delete, can_delete, dependency_details = self._check_dependencies(
                selected_data
            )

            # Calculate how many main records can now be deleted
            can_delete_count = len(can_delete)

            # Prepare context for rendering the updated modal
            self.object_list = self.get_queryset()
            context = self.get_context_data()
            context.update(
                {
                    "selected_ids": selected_data,
                    "selected_ids_json": json.dumps(selected_data),
                    "cannot_delete": cannot_delete,
                    "can_delete": can_delete,
                    "cannot_delete_count": len(cannot_delete),
                    "can_delete_count": can_delete_count,
                    "success_message": f"Successfully hard deleted {deleted_count} '{dep_model_name}' dependencies of '{record}'.",
                    "model_verbose_name": self.model._meta.verbose_name_plural,
                }
            )

            return context

        except self.model.DoesNotExist:
            logger.error(f"Record with ID {item_id} does not exist.")
            self.object_list = self.get_queryset()
            context = self.get_context_data()
            context.update(
                {
                    "selected_ids": selected_data,
                    "selected_ids_json": json.dumps(selected_data),
                    "error_message": f"Record with ID {item_id} does not exist.",
                }
            )
            return context
        except Exception as e:
            logger.error(f"Hard delete of dependencies failed: {str(e)}")
            self.object_list = self.get_queryset()
            context = self.get_context_data()
            context.update(
                {
                    "selected_ids": selected_data,
                    "selected_ids_json": json.dumps(selected_data),
                    "error_message": f"Hard delete of dependencies failed: {str(e)}",
                }
            )
            return context

    def _perform_soft_delete(self, record_ids):
        """
        Perform soft deletion by moving records and their dependencies to RecycleBin model.
        Returns the number of records deleted (main records only).
        """
        try:
            queryset = self.model.objects.filter(id__in=record_ids)
            deleted_count = 0
            for obj in queryset:
                # Soft delete dependent records
                related_objects = self.model._meta.related_objects
                for related in related_objects:
                    related_model = related.related_model
                    field_name = related.field.name
                    filter_kwargs = {field_name: obj}
                    manager = getattr(
                        related_model,
                        "objects",
                        getattr(related_model, "all_objects", None),
                    )
                    if manager is None:
                        raise AttributeError(
                            f"No manager ('objects' or 'all_objects') defined for {related_model.__name__}"
                        )
                    dependent_records = manager.filter(**filter_kwargs)
                    for dep_record in dependent_records:
                        RecycleBin.create_from_instance(
                            dep_record, user=self.request.user
                        )
                        dep_record.delete()
                # Soft delete the main record
                RecycleBin.create_from_instance(obj, user=self.request.user)
                obj.delete()
                deleted_count += 1
            return deleted_count
        except Exception as e:
            logger.error(f"Soft delete failed: {str(e)}")
            raise

    def handle_custom_bulk_action(self, action, record_ids):
        """Handle custom bulk actions based on their configuration."""
        try:
            if action.get("handler"):
                # Call custom handler function if provided
                handler = getattr(self, action["handler"], None)
                if callable(handler):
                    return handler(record_ids, self.request)
                else:
                    return HttpResponse(
                        f"Handler {action['handler']} not found.", status=500
                    )

            # Default behavior: HTMX POST request
            url = action.get("url")
            if not url:
                return HttpResponse(
                    f"No URL provided for action {action['name']}.", status=400
                )

            context = self.get_context_data()
            context.update(
                {
                    "selected_ids": record_ids,
                    "selected_ids_json": json.dumps(record_ids),
                    "action_name": action["name"],
                }
            )

            # Prepare HTMX attributes
            hx_attrs = {
                (
                    "hx-post"
                    if action.get("method", "POST").upper() == "POST"
                    else "hx-get"
                ): url,
                "hx-target": action.get("target", "#modalBox"),
                "hx-swap": action.get("swap", "innerHTML"),
                "hx-vals": f'js:{{"selected_ids": JSON.stringify(selectedRecordIds("{self.view_id}")), "action": "{action["name"]}"}}',
            }
            if action.get("after_request"):
                hx_attrs["hx-on::after-request"] = action["after_request"]

            context["hx_attrs"] = hx_attrs
            return render(self.request, "list_view.html", context)

        except Exception as e:
            logger.error(f"Custom  action {action['name']} failed: {str(e)}")
            return HttpResponse(f"Action {action['name']} failed: {str(e)}", status=500)

    def post(self, request, *args, **kwargs):
        """
        Handle POST requests for exporting data.
        """

        record_ids = request.POST.get("record_ids")
        columns = [
            value
            for key, value in request.POST.items()
            if key.startswith("expo_avail_") or key.startswith("expo_add_")
        ]
        action = request.POST.get("action")
        # columns = request.POST.getlist("export_columns")
        export_format = request.POST.get("export_format")
        delete_type = request.POST.get("delete_type")

        # Handle custom bulk actions
        if action in [bulk["name"] for bulk in self.custom_bulk_actions]:
            try:
                record_ids = json.loads(record_ids) if record_ids else []
                bulk_action = next(
                    bulk for bulk in self.custom_bulk_actions if bulk["name"] == action
                )
                return self.handle_custom_bulk_action(bulk_action, record_ids)
            except json.JSONDecodeError as e:
                return HttpResponse("Invalid JSON data for record_ids", status=400)

        if action in [
            additional["name"] for additional in self.additional_action_button
        ]:
            try:
                record_ids = json.loads(record_ids) if record_ids else []
                bulk_action = next(
                    additional
                    for additional in self.additional_action_button
                    if additional["name"] == action
                )
                return self.handle_custom_bulk_action(bulk_action, record_ids)
            except json.JSONDecodeError as e:
                return HttpResponse("Invalid JSON data for record_ids", status=400)

        if request.POST.get("delete_mode_form") == "true":
            selected_ids = request.POST.get("selected_ids", "[]")
            try:
                selected_ids = json.loads(selected_ids)
                selected_ids = [int(id) for id in selected_ids if id.isdigit()]
                valid_ids = list(
                    self.get_queryset()
                    .filter(id__in=selected_ids)
                    .values_list("id", flat=True)
                )
                self.object_list = self.get_queryset()
                context = self.get_context_data()
                context["selected_ids"] = valid_ids
                context["selected_ids_json"] = json.dumps(valid_ids)
                if not valid_ids:
                    messages.error(request, "No rows selected for deletion.")
                    return HttpResponse("<script>$('#reloadButton').click();</script>")
                return render(request, "partials/delete_mode_form.html", context)
            except (json.JSONDecodeError, ValueError) as e:
                logger.error(f"Error processing selected_ids: {str(e)}")
                self.object_list = self.get_queryset()
                context = self.get_context_data()
                context["selected_ids"] = []
                context["selected_ids_json"] = json.dumps([])
                return render(request, "partials/delete_mode_form.html", context)

        if request.POST.get("bulk_update_form") == "true":
            selected_ids = request.POST.get("selected_ids", "[]")
            try:
                selected_ids = json.loads(selected_ids)
                selected_ids = [int(id) for id in selected_ids if id.isdigit()]
                valid_ids = list(
                    self.get_queryset()
                    .filter(id__in=selected_ids)
                    .values_list("id", flat=True)
                )
                self.object_list = self.get_queryset()
                context = self.get_context_data()
                context["selected_ids"] = selected_ids
                context["selected_ids_json"] = json.dumps(selected_ids)
                return render(request, "partials/bulk_update_form.html", context)
            except (json.JSONDecodeError, ValueError) as e:
                logger.error(f"Error processing selected_ids: {str(e)}")
                self.object_list = self.get_queryset()
                context = self.get_context_data()
                context["selected_ids"] = []
                context["selected_ids_json"] = json.dumps([])
                return render(request, "partials/bulk_update_form.html", context)

        # Handle bulk delete form rendering for hard delete
        if request.POST.get("bulk_delete_form") == "true":
            selected_ids = request.POST.get("selected_ids", "[]")
            try:
                selected_ids = json.loads(selected_ids)
                selected_ids = [int(id) for id in selected_ids if id.isdigit()]
                valid_ids = list(
                    self.get_queryset()
                    .filter(id__in=selected_ids)
                    .values_list("id", flat=True)
                )
                # Check dependencies for the bulk delete form
                cannot_delete, can_delete, dependency_details = (
                    self._check_dependencies(valid_ids)
                )
                self.object_list = self.get_queryset()
                context = self.get_context_data()
                context.update(
                    {
                        "selected_ids": valid_ids,
                        "selected_ids_json": json.dumps(valid_ids),
                        "cannot_delete": cannot_delete,
                        "can_delete": can_delete,
                        "cannot_delete_count": len(cannot_delete),
                        "can_delete_count": len(can_delete),
                        "model_verbose_name": self.model._meta.verbose_name_plural,
                    }
                )
                return render(request, "partials/bulk_delete_form.html", context)
            except (json.JSONDecodeError, ValueError) as e:
                logger.error(f"Error processing selected_ids: {str(e)}")
                self.object_list = self.get_queryset()
                context = self.get_context_data()
                context.update(
                    {
                        "selected_ids": [],
                        "selected_ids_json": json.dumps([]),
                        "cannot_delete": [],
                        "can_delete": [],
                        "cannot_delete_count": 0,
                        "can_delete_count": 0,
                        "error_message": "Invalid selected IDs provided.",
                        "model_verbose_name": self.model._meta.verbose_name_plural,
                    }
                )
                return render(request, "partials/bulk_delete_form.html", context)

        # Handle bulk delete form rendering for soft delete
        if request.POST.get("soft_delete_form") == "true":
            selected_ids = request.POST.get("selected_ids", "[]")
            try:
                selected_ids = json.loads(selected_ids)
                selected_ids = [int(id) for id in selected_ids if id.isdigit()]
                valid_ids = list(
                    self.get_queryset()
                    .filter(id__in=selected_ids)
                    .values_list("id", flat=True)
                )
                # Check dependencies for the bulk delete form
                cannot_delete, can_delete, dependency_details = (
                    self._check_dependencies(valid_ids)
                )
                self.object_list = self.get_queryset()
                context = self.get_context_data()
                context.update(
                    {
                        "selected_ids": valid_ids,
                        "selected_ids_json": json.dumps(valid_ids),
                        "cannot_delete": cannot_delete,
                        "can_delete": can_delete,
                        "cannot_delete_count": len(cannot_delete),
                        "can_delete_count": len(can_delete),
                        "model_verbose_name": self.model._meta.verbose_name_plural,
                    }
                )
                return render(request, "partials/soft_delete_form.html", context)
            except (json.JSONDecodeError, ValueError) as e:
                logger.error(f"Error processing selected_ids: {str(e)}")
                self.object_list = self.get_queryset()
                context = self.get_context_data()
                context.update(
                    {
                        "selected_ids": [],
                        "selected_ids_json": json.dumps([]),
                        "cannot_delete": [],
                        "can_delete": [],
                        "cannot_delete_count": 0,
                        "can_delete_count": 0,
                        "error_message": "Invalid selected IDs provided.",
                        "model_verbose_name": self.model._meta.verbose_name_plural,
                    }
                )
                return render(request, "partials/soft_delete_form.html", context)

        if action == "bulk_delete" and record_ids:
            try:
                record_ids = json.loads(record_ids)
                cannot_delete, can_delete, dependency_details = (
                    self._check_dependencies(record_ids)
                )

                if request.POST.get("confirm_delete") == "true":
                    try:
                        can_delete_ids = [item["id"] for item in can_delete]
                        individual_view_id = self.request.POST.get("view_id", "")
                        if delete_type == "soft":
                            deleted_count = self._perform_soft_delete(record_ids)
                            messages.success(
                                request,
                                f"Successfully soft deleted {deleted_count} records.",
                            )
                            return HttpResponse(
                                f"<script>$('#reloadButton').click();closeModal();$('#clear-select-btn-{individual_view_id}').click();</script>"
                            )
                        elif delete_type == "hard_non_dependent":  # Hard delete
                            deleted_count = self.model.objects.filter(
                                id__in=can_delete_ids
                            ).delete()[0]
                            messages.success(
                                request,
                                f"Successfully hard deleted {deleted_count} records.",
                            )
                            return HttpResponse(
                                f"<script>$('#reloadButton').click();$('#clear-select-btn-{individual_view_id}').click();</script>"
                            )
                    except Exception as e:
                        logger.error(f"Delete failed: {str(e)}")
                        messages.error(request, f"Delete failed: {str(e)}")
                        return HttpResponse(
                            "<script>$('#reloadButton').click();</script>"
                        )

                # Render the bulk delete form with dependency information
                self.object_list = self.get_queryset()
                context = self.get_context_data()
                context.update(
                    {
                        "selected_ids": record_ids,
                        "cannot_delete": cannot_delete,
                        "can_delete": can_delete,
                        "cannot_delete_count": len(cannot_delete),
                        "can_delete_count": len(can_delete),
                        "selected_ids_json": json.dumps(record_ids),
                        "model_verbose_name": self.model._meta.verbose_name_plural,
                    }
                )
                return render(request, "partials/bulk_delete_form.html", context)

            except json.JSONDecodeError as e:
                logger.error(f"JSON decode error: {e}")
                return HttpResponse("Invalid JSON data for record_ids", status=400)

        if action == "delete_item_with_dependencies" and request.POST.get("record_id"):

            try:
                item_id = int(request.POST.get("record_id"))
                # selected_ids = request.POST.getlist('selected_ids')
                # record_ids = json.loads(record_ids) if record_ids else []
                selected_ids = json.loads(request.POST.get("selected_ids", "[]"))
                selected_data = [int(id) for id in selected_ids] if selected_ids else []
                context = self._delete_item_with_dependencies(
                    item_id, record_ids, selected_data
                )
                return render(request, "partials/bulk_delete_form.html", context)

            except json.JSONDecodeError as e:
                logger.error(f"JSON decode error: {e}")
                self.object_list = self.get_queryset()
                context = self.get_context_data()
                context["error_message"] = "Invalid JSON data for record_ids."
                return render(request, "partials/bulk_delete_form.html", context)

        if action == "delete_all_dependencies" and request.POST.get("record_id"):
            try:
                item_id = int(request.POST.get("record_id"))
                selected_ids = json.loads(request.POST.get("selected_ids", "[]"))
                selected_data = [int(id) for id in selected_ids] if selected_ids else []
                context = self._delete_all_dependencies(item_id, selected_data)
                return render(request, "partials/bulk_delete_form.html", context)

            except json.JSONDecodeError as e:
                logger.error(f"JSON decode error: {e}")
                self.object_list = self.get_queryset()
                context = self.get_context_data()
                context["error_message"] = "Invalid JSON data for record_ids."
                return render(request, "partials/bulk_delete_form.html", context)
            except ValueError as e:
                logger.error(f"Value error: {e}")
                self.object_list = self.get_queryset()
                context = self.get_context_data()
                context["error_message"] = "Invalid record ID provided."
                return render(request, "partials/bulk_delete_form.html", context)

        if record_ids and export_format:
            try:
                record_ids = json.loads(record_ids)
                return self.handle_export(record_ids, columns, export_format)
            except json.JSONDecodeError as e:
                return HttpResponse("Invalid JSON data for record_ids", status=400)

        if record_ids:
            try:
                record_ids = json.loads(record_ids)
                # Collect all bulk update fields and values
                bulk_updates = {}
                for field in self.bulk_update_fields:
                    value = request.POST.get(f"bulk_update_value_{field}")
                    if value:  # Only include fields with non-empty values
                        bulk_updates[field] = value
                return self.handle_bulk_update(record_ids, bulk_updates)
            except json.JSONDecodeError as e:
                return HttpResponse("Invalid JSON data for record_ids", status=400)

        elif request.POST.get("bulk_update_field"):
            # Handle bulk update
            field_name = request.POST.get("bulk_update_field")
            new_value = request.POST.get("bulk_update_value")
            if record_ids and field_name and new_value:
                try:
                    record_ids = json.loads(record_ids)
                    return self.handle_bulk_update(record_ids, field_name, new_value)
                except json.JSONDecodeError as e:
                    return HttpResponse("Invalid JSON data for record_ids", status=400)
            return HttpResponse("Invalid request: Missing required fields", status=400)
        return HttpResponse("Invalid request: Missing required fields", status=400)

    def handle_export(self, record_ids, columns, export_format):
        """
        Handle the export of data in the specified format.
        """

        try:
            queryset = self.model.objects.filter(id__in=record_ids)
            model_fields = [
                (str(field.verbose_name), field.name, field)
                for field in self.model._meta.fields
            ]
            property_labels = getattr(self.model, "PROPERTY_LABELS", None)
            if not property_labels:
                property_labels = {
                    name.replace("get_", "", 1): name.replace("get_", "", 1)
                    .replace("_", " ")
                    .title()
                    for name, member in inspect.getmembers(
                        self.model, predicate=lambda x: isinstance(x, property)
                    )
                }

            for name, member in inspect.getmembers(
                self.model, predicate=lambda x: isinstance(x, property)
            ):
                label_key = (
                    name.replace("get_", "", 1) if name.startswith("get_") else name
                )
                if label_key in property_labels:
                    model_fields.append((str(property_labels[label_key]), name, None))

            for field in self.model._meta.fields:
                if field.choices:
                    method_name = f"get_{field.name}_display"
                    if hasattr(self.model, method_name):
                        model_fields.append(
                            (str(field.verbose_name), method_name, "method")
                        )

            # Get table columns from _get_columns
            table_columns = self._get_columns()
            table_column_names = [
                col[1] for col in table_columns
            ]  # Field names of table columns

            # Use selected columns if provided, otherwise use table columns
            if columns:
                column_headers = [
                    field[0] for field in model_fields if field[1] in columns
                ]
                selected_fields = [
                    field for field in model_fields if field[1] in columns
                ]
            else:
                # Use table columns instead of all model fields
                column_headers = [col[0] for col in table_columns]
                selected_fields = [
                    field for field in model_fields if field[1] in table_column_names
                ]

                # If no table columns are defined, log error and return
                if not table_columns:
                    return HttpResponse(
                        "No table columns defined for export", status=400
                    )

            data = []
            for obj in queryset:
                row = []
                for verbose_name, field_name, field in selected_fields:
                    try:
                        value = getattr(obj, field_name, "")
                        if field == "method" or callable(value):
                            value = value()
                        elif field is None:  # This is a @property
                            # Properties already computed by getattr, no further action needed
                            pass
                        elif isinstance(field, ForeignKey):
                            value = (
                                str(getattr(value, "username", value)) if value else ""
                            )
                        elif isinstance(field, ManyToManyField):
                            # Handle ManyToManyField (e.g., tags)
                            value = (
                                ", ".join(str(item) for item in value.all())
                                if value
                                else ""
                            )
                        elif callable(value):
                            value = value()
                        row.append(str(value) if value is not None else "")
                    except Exception as e:
                        row.append("")  # Fallback to empty string
                data.append(row)

            model_verbose_name = self.model._meta.verbose_name_plural.lower().replace(
                " ", "_"
            )
            document_title = f"Exported {self.model._meta.verbose_name_plural}"
            if export_format == "csv":
                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = (
                    f'attachment; filename="exported_{model_verbose_name}.csv"'
                )
                writer = csv.writer(response)
                writer.writerow(column_headers)
                for row in data:
                    writer.writerow(row)
                return response

            elif export_format == "xlsx":
                wb = Workbook()
                ws = wb.active

                # Append the header row
                ws.append([str(header) for header in column_headers])

                # ws.append(column_headers)

                # Style the header row
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal="center")
                header_fill = PatternFill(
                    start_color="eafb5b", end_color="eafb5b", fill_type="solid"
                )  # Light gray background

                for cell in ws[1]:  # First row (headers)
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.fill = header_fill

                # Set column widths (adjust the width as needed)
                for col in ws.columns:
                    column_letter = col[
                        0
                    ].column_letter  # Get the column letter (A, B, C, ...)
                    ws.column_dimensions[column_letter].width = (
                        25  # Set width to 20 (adjust as needed)
                    )

                # Append the data rows
                for row in data:
                    ws.append([str(cell) if cell is not None else "" for cell in row])

                # for row in data:
                #     ws.append(row)

                # Optionally, adjust row heights if needed (e.g., for better readability)
                for idx, row in enumerate(ws.rows, 1):
                    ws.row_dimensions[idx].height = 15  # Adjust row height (optional)

                response = HttpResponse(
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = (
                    f'attachment; filename="exported_{model_verbose_name}.xlsx"'
                )
                buffer = BytesIO()
                wb.save(buffer)
                response.write(buffer.getvalue())
                buffer.close()
                return response

            if export_format == "pdf":
                buffer = BytesIO()
                # Use landscape orientation for better width
                page_size = (letter[1], letter[0])  # 792 x 612 points (landscape)
                width, height = page_size

                c = canvas.Canvas(buffer, pagesize=page_size)

                # Set PDF metadata title
                c.setTitle(document_title)

                # Set fonts and adjust font size for readability
                title_font_size = 18
                header_font_size = 12
                data_font_size = 10

                # Table settings
                start_x = 50
                start_y = height - 100
                min_col_width = 120
                padding = 8
                max_rows_per_page = 7
                max_cols_per_page = 6
                extra_row_spacing = 10

                # Helper function to wrap text
                def wrap_text(text, max_chars):
                    text = str(text) if text is not None else ""
                    if len(text) <= max_chars:
                        return [text] if text else [""]
                    words = text.split()
                    lines = []
                    current_line = ""
                    for word in words:
                        if len(current_line) + len(word) + 1 <= max_chars:
                            current_line += word + " "
                        else:
                            lines.append(current_line.strip())
                            current_line = word + " "
                    if current_line:
                        lines.append(current_line.strip())
                    return lines if lines else [""]

                # Split columns into chunks
                column_chunks = [
                    column_headers[i : i + max_cols_per_page]
                    for i in range(0, len(column_headers), max_cols_per_page)
                ]
                field_chunks = [
                    selected_fields[i : i + max_cols_per_page]
                    for i in range(0, len(selected_fields), max_cols_per_page)
                ]

                for chunk_idx, (chunk_headers, chunk_fields) in enumerate(
                    zip(column_chunks, field_chunks)
                ):
                    # Calculate column width for this chunk
                    total_table_width = min(
                        len(chunk_headers) * min_col_width, width - 100
                    )
                    col_width = (
                        total_table_width / len(chunk_headers) if chunk_headers else 100
                    )
                    max_chars_per_line = int(col_width // (header_font_size * 0.5))

                    # Reset row counter for the chunk
                    rows_drawn = 0
                    row_start = 0

                    while row_start < len(data):
                        # Draw title with column range
                        c.setFont("Helvetica-Bold", title_font_size)
                        column_range = f"Columns {(chunk_idx * max_cols_per_page) + 1} to {min((chunk_idx + 1) * max_cols_per_page, len(column_headers))}"
                        c.drawCentredString(
                            width / 2, height - 50, f"{document_title} ({column_range})"
                        )

                        # Draw headers for this chunk
                        c.setFont("Helvetica-Bold", header_font_size)
                        c.setFillColor(colors.black)
                        header_y = start_y
                        max_header_lines = 1
                        for i, header in enumerate(chunk_headers):
                            wrapped_header = wrap_text(header, max_chars_per_line)
                            max_header_lines = max(
                                max_header_lines, len(wrapped_header)
                            )

                        # Draw header background
                        header_height = max_header_lines * (header_font_size + 2) + 15
                        c.setFillColor(colors.lightgrey)
                        c.rect(
                            start_x,
                            header_y - header_height + 5,
                            total_table_width,
                            header_height,
                            fill=1,
                            stroke=0,
                        )

                        # Redraw headers, centered vertically within the background
                        c.setFont("Helvetica-Bold", header_font_size)
                        c.setFillColor(colors.black)
                        for i, header in enumerate(chunk_headers):
                            x = start_x + i * col_width + padding
                            wrapped_header = wrap_text(header, max_chars_per_line)
                            total_text_height = len(wrapped_header) * (
                                header_font_size + 2
                            )
                            y_offset = (header_height - total_text_height) / 2 + 3
                            for line in wrapped_header:
                                c.drawString(x, header_y - y_offset, line)
                                y_offset += header_font_size + 2

                        # Draw data rows for this chunk
                        c.setFont("Helvetica", data_font_size)
                        y = header_y - header_height - 10
                        rows_drawn = 0

                        for row_idx in range(row_start, len(data)):
                            if rows_drawn >= max_rows_per_page:
                                break

                            row = []
                            for _, field_name, _ in chunk_fields:
                                try:
                                    index = [f[1] for f in selected_fields].index(
                                        field_name
                                    )
                                    row.append(data[row_idx][index])
                                except ValueError:
                                    row.append("")

                            # row = [
                            #     data[row_idx][selected_fields.index(field)]
                            #     for field in chunk_fields
                            # ]

                            # Calculate row height and center the text
                            max_lines_in_row = 1
                            for value in row:
                                wrapped_value = wrap_text(value, max_chars_per_line)
                                max_lines_in_row = max(
                                    max_lines_in_row, len(wrapped_value)
                                )
                            row_height = (
                                max_lines_in_row * (data_font_size + 2)
                                + extra_row_spacing
                            )
                            total_text_height = max_lines_in_row * (data_font_size + 2)
                            text_y_offset = (row_height - total_text_height) / 2 + 9

                            # Draw alternating row background
                            if rows_drawn % 2 == 0:
                                c.setFillColor(colors.whitesmoke)
                                c.rect(
                                    start_x,
                                    y - row_height,
                                    total_table_width,
                                    row_height,
                                    fill=1,
                                    stroke=0,
                                )

                            # Draw row data, centered vertically with adjustment
                            for i, value in enumerate(row):
                                wrapped_value = wrap_text(value, max_chars_per_line)
                                x = start_x + i * col_width + padding
                                y_offset = text_y_offset
                                for line in wrapped_value:
                                    c.setFillColor(colors.black)
                                    c.drawString(x, y - y_offset, line)
                                    y_offset += data_font_size + 2

                            y -= row_height
                            rows_drawn += 1

                        row_start += max_rows_per_page
                        c.showPage()

                c.save()
                response = HttpResponse(content_type="application/pdf")
                response["Content-Disposition"] = (
                    f'attachment; filename="exported_{model_verbose_name}.pdf"'
                )
                response.write(buffer.getvalue())
                buffer.close()
                return response

        except Exception as e:
            return HttpResponse(f"Export failed: {str(e)}", status=500)

    def handle_bulk_update(self, record_ids, bulk_updates):
        try:
            queryset = self.model.objects.filter(id__in=record_ids)
            field_infos = {field["name"]: field for field in self._get_model_fields()}

            update_dict = {}
            has_valid_values = False
            for field_name, new_value in bulk_updates.items():
                if new_value == "" or new_value is None:
                    continue

                field_info = field_infos.get(field_name)
                if not field_info:
                    return HttpResponse(f"Field {field_name} not found", status=400)

                field_type = field_info["type"]
                try:
                    if field_type == "boolean":
                        new_value = new_value.lower() in ("true", "yes", "1")
                    elif field_type in ("number", "integer"):
                        new_value = int(new_value)
                    elif field_type in ("float", "decimal"):
                        from decimal import Decimal

                        new_value = Decimal(new_value)
                    elif field_type in ("date", "datetime"):
                        from datetime import datetime

                        format = (
                            "%Y-%m-%d" if field_type == "date" else "%Y-%m-%dT%H:%M"
                        )
                        new_value = datetime.strptime(new_value, format)
                        if field_type == "date":
                            new_value = new_value.date()
                    elif field_type == "choice":
                        choices = [c["value"] for c in field_info.get("choices", [])]
                        if new_value not in choices:
                            return HttpResponse(
                                f"Invalid choice for {field_name}", status=400
                            )
                    elif field_type == "foreignkey":
                        if new_value == "":
                            new_value = None
                        elif new_value:
                            try:
                                new_value = int(new_value)
                            except ValueError:
                                pass
                    update_dict[field_name] = new_value
                    has_valid_values = True
                except ValueError as e:
                    return HttpResponse(
                        f"Invalid value for field {field_name}: {str(e)}", status=400
                    )

            if not has_valid_values:
                messages.info(
                    self.request, "No fields were updated as no values were provided."
                )
                return HttpResponse(
                    f"<script>$('#reloadButton').click();$('#clear-select-btn-{self.view_id}').click();</script>"
                )

            records_before = {obj.id: obj for obj in queryset}
            content_type = ContentType.objects.get_for_model(self.model)
            user = self.request.user if self.request.user.is_authenticated else None

            updated_count = queryset.update(**update_dict)

            if updated_count > 0:
                for record_id in record_ids:
                    if record_id not in records_before:
                        continue
                    record = records_before[record_id]
                    updated_record = self.model.objects.get(id=record_id)

                    changes = {}
                    for field_name, _ in update_dict.items():
                        old_value = getattr(record, field_name, None)
                        new_value = getattr(updated_record, field_name, None)
                        if old_value != new_value:
                            changes[field_name] = [
                                str(old_value) if old_value is not None else "--",
                                str(new_value) if new_value is not None else "--",
                            ]

                    if changes:
                        LogEntry.objects.create(
                            content_type=content_type,
                            object_id=record_id,
                            object_repr=str(updated_record),
                            action=LogEntry.Action.UPDATE,
                            actor=user,
                            timestamp=timezone.now(),
                            changes=changes,
                        )

            messages.success(
                self.request, f"Updated {updated_count} records successfully."
            )

            self.object_list = self.get_queryset()
            return HttpResponse(
                f"<script>$('#reloadButton').click();$('#clear-select-btn-{self.view_id}').click();</script>"
            )

        except Exception as e:
            return HttpResponse(f"Bulk update failed: {str(e)}", status=500)

    def get(self, request, *args, **kwargs):
        """
        Handle GET requests and ensure an HttpResponse is returned.
        """

        self.object_list = self.get_queryset()
        context = self.get_context_data()

        # Handle filter row addition
        if request.GET.get("add_filter_row") == "true":

            curr_row_id = int(request.GET.get("row_id"))
            new_row_id = curr_row_id + 1
            filter_rows = [{"row_id": new_row_id}]
            context["filter_rows"] = filter_rows
            return render(request, "partials/filter_row.html", context)

        if "remove_filter" in request.GET:
            return self.handle_remove_filter(request)

        if request.GET.get("clear_all_filters") == "true":
            return self.handle_clear_all_filters(request)

        if request.GET.get("remove_filter_field") == "true":
            return HttpResponse("")

        # Handle HTMX requests
        if request.headers.get("HX-Request") == "true":
            # Handle field change
            if request.GET.get("field_change") and not request.GET.get(
                "operator_change"
            ):
                field_name = request.GET.get("field")
                row_id = request.GET.get("row_id")
                return self.handle_field_change(request, field_name, row_id)

            # Handle operator change
            if request.GET.get("operator_change"):
                field_name = request.GET.get("field")
                operator = request.GET.get("operator")
                row_id = request.GET.get("row_id")
                return self.handle_operator_change(
                    request, field_name, operator, row_id
                )

            return render(request, self.template_name, context)

        return self.render_to_response(context)

    def handle_remove_filter(self, request):
        """Handle removing a specific filter or the search parameter while preserving other query parameters."""
        # Get the filter to remove
        remove_filter = request.GET.get("remove_filter", "")  # Default to empty string

        query_params = request.GET.copy()

        new_fields = []
        new_operators = []
        new_values = []
        new_start_values = []
        new_end_values = []
        search_value = query_params.get("search", "")  # Treat search as a single value

        # Handle filter removal or search removal
        if remove_filter == "search":
            fields = [f for f in query_params.getlist("field") if f.strip()]
            operators = [o for o in query_params.getlist("operator") if o.strip()]
            values = [v for v in query_params.getlist("value") if v.strip()]
            start_values = [
                sv for sv in query_params.getlist("start_value") if sv.strip()
            ]
            end_values = [ev for ev in query_params.getlist("end_value") if ev.strip()]

            new_fields = fields
            new_operators = operators
            new_values = values
            new_start_values = start_values
            new_end_values = end_values
            search_value = ""
        else:
            # Handle filter removal by index
            filter_index = int(remove_filter) if remove_filter.isdigit() else -1
            fields = query_params.getlist("field")
            operators = query_params.getlist("operator")
            values = query_params.getlist("value")
            start_values = query_params.getlist("start_value")
            end_values = query_params.getlist("end_value")

            # Create new lists without the removed filter, ignoring empty values
            for i in range(len(fields)):
                if i != filter_index and fields[i].strip():
                    new_fields.append(fields[i])
                    if i < len(operators) and operators[i].strip():
                        new_operators.append(operators[i])
                    if i < len(values) and values[i].strip():
                        new_values.append(values[i])
                    if i < len(start_values) and start_values[i].strip():
                        new_start_values.append(start_values[i])
                    if i < len(end_values) and end_values[i].strip():
                        new_end_values.append(end_values[i])

        # Initialize new query parameters
        new_query_params = QueryDict("", mutable=True)

        # Preserve all other query parameters except specific ones
        for key, values_list in query_params.lists():
            if key not in [
                "field",
                "operator",
                "value",
                "start_value",
                "end_value",
                "remove_filter",
                "page",
                "apply_filter",
                "hx_trigger",
                "search",
            ]:
                for value in values_list:
                    if value.strip():  # Only include non-empty values
                        new_query_params.appendlist(key, value)

        # Append filter-related parameters
        for field in new_fields:
            new_query_params.appendlist("field", field)
        for operator in new_operators:
            new_query_params.appendlist("operator", operator)
        for value in new_values:
            new_query_params.appendlist("value", value)
        for start_value in new_start_values:
            new_query_params.appendlist("start_value", start_value)
        for end_value in new_end_values:
            new_query_params.appendlist("end_value", end_value)

        # Only include search if it has a non-empty value
        if search_value:
            new_query_params["search"] = search_value

        # Only set apply_filter if there are non-empty filters
        if new_fields:
            new_query_params["apply_filter"] = "true"

        # Update request.GET with new query parameters
        request.GET = new_query_params
        self.object_list = self.get_queryset()
        context = self.get_context_data()
        response = render(request, self.template_name, context)

        # Update URL for HX-Push-Url
        new_query_string = new_query_params.urlencode()
        current_path = self.main_url
        url = f"{current_path}?{new_query_string}" if new_query_string else current_path
        response["HX-Push-Url"] = url

        return response

    def handle_clear_all_filters(self, request):
        """Handle clearing all applied filters while preserving non-filter query parameters."""
        query_params = request.GET.copy()
        filter_params = [
            "field",
            "operator",
            "value",
            "start_value",
            "end_value",
            "apply_filter",
            "clear_all_filters",
            "page",
            "search",
        ]
        new_query_params = QueryDict(mutable=True)
        for key, values in query_params.lists():
            if key not in filter_params:
                for value in values:
                    new_query_params.appendlist(key, value)
        self.object_list = self.get_queryset()
        context = self.get_context_data()
        response = render(request, self.template_name, context)
        new_query_string = new_query_params.urlencode()
        url = f"{self.main_url}" + (f"?{new_query_string}" if new_query_string else "")
        response["HX-Push-Url"] = url

        return response

    def get_context_data(self, **kwargs):
        """Enhance context with column and filtering information."""
        context = super().get_context_data(**kwargs)
        if self.store_ordered_ids:
            context["ordered_ids_key"] = self.ordered_ids_key
            context["ordered_ids"] = self.request.session.get(self.ordered_ids_key, [])

        filter_fields = self._get_model_fields()
        view_type = self.request.GET.get("view_type") or self.get_default_view_type()
        context["saved_list_name"] = None  # default

        if view_type and view_type.startswith("saved_list_"):
            try:
                saved_list_id = int(view_type.split("_")[2])
                saved_list = self.request.user.saved_filter_lists.filter(
                    id=saved_list_id
                ).first()
                if saved_list:
                    context["saved_list_name"] = saved_list.name
            except (IndexError, ValueError):
                pass
        context["view_type"] = view_type
        context["filter_fields"] = filter_fields
        context["filter_push_url"] = self.filter_url_push
        context["model_verbose_name"] = self.model._meta.verbose_name_plural
        context["model_name"] = self.model.__name__
        context["no_record_add_button"] = self.no_record_add_button or {}
        context["no_record_section"] = self.no_record_section
        context["no_record_msg"] = self.no_record_msg
        context["bulk_update_two_column"] = self.bulk_update_two_column
        header_attrs_dict = {}
        for item in self.header_attrs:
            for col_name, attrs in item.items():
                header_attrs_dict[col_name] = attrs

        col_attrs_dict = {}
        visible_columns = self._get_columns()

        if not visible_columns and self.columns:
            visible_columns = [[col[0], col[1]] for col in self.columns]

        if self.col_attrs and visible_columns:
            first_column_field = visible_columns[0][1]
            for item in self.col_attrs:
                for col_name, attrs in item.items():
                    col_attrs_dict[first_column_field] = attrs
                    break
                break

        context["header_attrs"] = header_attrs_dict
        context["col_attrs"] = col_attrs_dict

        field_operators = {}
        field_types = {}
        choices = {}

        operator_display = {
            "exact": "Equals",
            "iexact": "Equals (case insensitive)",
            "icontains": "Contains",
            "gt": "Greater than",
            "lt": "Less than",
            "gte": "Greater than or equal to",
            "lte": "Less than or equal to",
            "startswith": "Starts with",
            "endswith": "Ends with",
            "date_range": "Between",
            "isnull": "Is empty",
        }
        context["operator_display"] = operator_display
        context["pinned_view"] = PinnedView.all_objects.filter(
            user=self.request.user, model_name=self.model.__name__
        ).first()
        field_verbose_names = {}
        for field in filter_fields:
            field_operators[field["name"]] = field.get("operators", [])
            field_types[field["name"]] = field.get("type", [])
            choices[field["name"]] = field.get("choices", [])
            field_verbose_names[field["name"]] = field.get("verbose_name", "")

        context["field_verbose_names"] = field_verbose_names
        context["columns"] = self._get_columns()
        context["raw_attrs"] = self.raw_attrs
        context["view_id"] = self.view_id
        context["action_method"] = self.action_method
        context["bulk_export_option"] = self.bulk_export_option
        context["current_sort"] = self.request.GET.get("sort", self.default_sort_field)
        context["exclude_columns_from_sorting"] = self.exclude_columns_from_sorting
        context["current_direction"] = self.request.GET.get(
            "direction", self.default_sort_direction
        )
        context["current_query"] = self.request.GET.urlencode()
        context["is_htmx_request"] = self.request.headers.get("HX-Request") == "true"
        context["has_next"] = False
        context["next_page"] = None
        if "page_obj" in context and context["page_obj"] is not None:
            context["has_next"] = context["page_obj"].has_next()
            if context["has_next"]:
                context["next_page"] = context["page_obj"].next_page_number()
        context["search_url"] = self.search_url or self.request.path
        context["main_url"] = self.main_url or self.request.path
        query_params = {
            item: self.request.GET.getlist(item) for item in self.request.GET
        }
        context["query_params"] = query_params

        filter_rows = []
        if (
            query_params.get("field")
            and self.request.GET.get("add_filter_row") != "true"
        ):
            for i, field in enumerate(query_params["field"]):
                field_info = next((f for f in filter_fields if f["name"] == field), {})
                raw_value = (
                    query_params.get("value", [None])[i]
                    if i < len(query_params.get("value", []))
                    else None
                )

                # Convert ForeignKey ID to display value
                display_value = raw_value
                if field_info.get("type") == "foreignkey" and raw_value:
                    try:
                        # Get the related model field
                        model_field = self.model._meta.get_field(field)
                        related_model = model_field.related_model

                        # Fetch the related object and get its string representation
                        related_obj = related_model.objects.get(pk=raw_value)
                        display_value = str(related_obj)
                    except Exception:
                        # If anything fails, keep the raw value
                        display_value = raw_value
                elif field_info.get("type") == "choice" and raw_value:
                    try:
                        # Get the field and its choices
                        field_obj = self.model._meta.get_field(field)
                        if field_obj.choices:
                            choices_dict = dict(field_obj.choices)
                            display_value = choices_dict.get(raw_value, raw_value)
                        else:
                            display_value = raw_value
                    except Exception:
                        display_value = raw_value
                row = {
                    "row_id": i,
                    "field": field,
                    "operator": (
                        query_params.get("operator", [None])[i]
                        if i < len(query_params.get("operator", []))
                        else None
                    ),
                    "value": raw_value,
                    "raw_value": display_value,
                    "start_value": (
                        query_params.get("start_value", [None])[i]
                        if i < len(query_params.get("start_value", []))
                        else None
                    ),
                    "end_value": (
                        query_params.get("end_value", [None])[i]
                        if i < len(query_params.get("end_value", []))
                        else None
                    ),
                    "operators": field_operators.get(field, []),
                    "type": field_types.get(field, []),
                    "choices": choices.get(field, []),
                    "model": field_info.get("model", None),  # Use field_info
                    "app_label": field_info.get("app_label", None),
                    "verbose_name": field_verbose_names.get(field, field),
                    "operator_display": operator_display.get(
                        (
                            query_params.get("operator", [None])[i]
                            if i < len(query_params.get("operator", []))
                            else None
                        ),
                        (
                            query_params.get("operator", [None])[i]
                            if i < len(query_params.get("operator", []))
                            else None
                        ),
                    ),
                }
                filter_rows.append(row)
        else:
            filter_rows = [
                {
                    "row_id": 0,
                    "field": None,
                    "operator": None,
                    "value": None,
                    "operators": [],
                }
            ]

        context["filter_rows"] = filter_rows
        context["last_row_id"] = len(filter_rows) - 1

        if hasattr(self, "filterset"):
            context["filterset"] = self.filterset

        if self.actions and len(self.actions) > self.max_visible_actions:
            context["visible_actions"] = self.actions[: self.max_visible_actions]
            context["dropdown_actions"] = self.actions[self.max_visible_actions :]
            context["use_dropdown"] = True
        else:
            context["visible_actions"] = self.actions
            context["dropdown_actions"] = []
            context["use_dropdown"] = False

        context["model_name"] = self.model.__name__
        context["app_label"] = self.model._meta.app_label
        context["total_records_count"] = self.get_queryset().count()
        context["selected_ids"] = list(self.get_queryset().values_list("id", flat=True))
        context["selected_ids_json"] = json.dumps(context["selected_ids"])
        context["custom_bulk_actions"] = self.custom_bulk_actions
        context["additional_action_button"] = self.additional_action_button
        bulk_update_fields_metadata = [
            field for field in filter_fields if field["name"] in self.bulk_update_fields
        ]
        context["bulk_update_fields"] = bulk_update_fields_metadata
        context["clear_session_button_enabled"] = self.clear_session_button_enabled
        context["bulk_select_option"] = self.bulk_select_option
        context["bulk_update_option"] = self.bulk_update_option
        context["enable_sorting"] = self.enable_sorting
        context["sorting_target"] = self.sorting_target
        context["bulk_delete_enabled"] = self.bulk_delete_enabled
        queryset_ids = list(self.get_queryset().values_list("id", flat=True))
        session_key = f"list_view_queryset_ids_{self.model._meta.model_name}"
        self.request.session[session_key] = queryset_ids
        query_params = self.request.GET.copy()
        if "page" in query_params:
            del query_params["page"]
        context["search_params"] = query_params.urlencode()
        # context["bulk_delete_url"] = reverse("horilla_generics:generic_bulk_delete")
        context["filter_set_class"] = self.filterset_class
        context["table_width"] = self.table_width
        context["table_class"] = self.table_class
        context["table_height_as_class"] = self.table_height_as_class
        context["table_height"] = self.table_height
        context["save_to_list_option"] = self.save_to_list_option
        return context


@method_decorator(htmx_required, name="dispatch")
class HorillaKanbanView(HorillaListView):

    template_name = "kanban_view.html"
    group_by_field = None
    paginate_by = 30
    filterset_module = "filters"
    kanban_attrs: str = None
    height_kanban = None

    _view_registry = {}

    def __init_subclass__(cls, **kwargs):
        super().__init_subclass__(**kwargs)
        if hasattr(cls, "model") and cls.model:
            HorillaKanbanView._view_registry[cls.model] = cls

    def dispatch(self, request, *args, **kwargs):
        if not self.request.user.is_authenticated:
            login_url = f"{reverse_lazy('horilla_core:login')}?next={request.path}"
            return redirect(login_url)
        app_label = kwargs.get("app_label")
        app_label = app_label.split(".")[-1] if app_label else ""
        model_name = kwargs.get("model_name") or request.POST.get("model_name")
        if model_name:
            try:
                self.model = apps.get_model(app_label=app_label, model_name=model_name)
            except Exception as e:
                raise HorillaHttp404(
                    f"Invalid app_label/model_name: {app_label}/{model_name}"
                )

        if self.model is None:
            raise ImproperlyConfigured("Model must be specified via URL or POST data.")

        return super().dispatch(request, *args, **kwargs)

    def can_user_modify_item(self, item):
        """
        Check if the user has permission to modify the item.
        Returns True if user can modify, False otherwise.
        """
        user = self.request.user
        model = self.model
        app_label = model._meta.app_label
        model_name = model._meta.model_name

        # Check if user has global change permission
        change_perm = f"{app_label}.change_{model_name}"
        if user.has_perm(change_perm):
            return True

        # Check for change_own permission
        change_own_perm = f"{app_label}.change_own_{model_name}"
        if user.has_perm(change_own_perm):
            # Get owner fields from model
            owner_fields = getattr(model, "OWNER_FIELDS", [])

            # Check if user owns this item
            for owner_field in owner_fields:
                try:
                    owner_value = getattr(item, owner_field, None)
                    if owner_value == user:
                        return True
                except AttributeError:
                    continue

        return False

    def post(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        if request.POST.get("item_id") and request.POST.get("new_column"):
            return self.update_kanban_item(request)
        elif request.POST.get("column_order"):
            return self.update_kanban_column_order(request)
        return super().post(request, *args, **kwargs)

    def update_kanban_item(self, request):
        item_id = request.POST.get("item_id")
        new_column = request.POST.get("new_column")
        app_label = request.POST.get("app_label")
        model_name = request.POST.get("model_name")
        class_name = request.POST.get("class_name")

        if not all([item_id, new_column, app_label, model_name, class_name]):
            return HttpResponse(status=400, content="Missing required parameters")

        try:
            view_class = HorillaKanbanView._view_registry.get(self.model)
            if not view_class:
                return HttpResponse(
                    status=404, content=f"View class {class_name} not found"
                )

            # Instantiate the view class
            view = view_class()
            view.request = request

            # Initialize model
            try:
                view.model = apps.get_model(
                    app_label=app_label.split(".")[-1], model_name=model_name
                )
            except LookupError as e:
                messages.error(
                    request, f"Invalid app_label/model_name: {app_label}/{model_name}"
                )
                return HttpResponse("<script>$('#reloadButton').click();")

            group_by = view.get_group_by_field()
            try:
                item = view.model.objects.get(pk=item_id)
                if not view.can_user_modify_item(item):
                    messages.error(
                        request, "You don't have permission to modify this item"
                    )
                    return HttpResponse("<script>$('#reloadButton').click();</script>")

                field = view.model._meta.get_field(group_by)

                if hasattr(field, "choices") and field.choices:
                    valid_choices = dict(field.choices)
                    reverse_choices = {v: k for k, v in valid_choices.items()}
                    if new_column in reverse_choices:
                        setattr(item, group_by, reverse_choices[new_column])
                    elif new_column in valid_choices:
                        setattr(item, group_by, new_column)
                    else:
                        return HttpResponse(
                            status=400, content=f"Invalid column value: {new_column}"
                        )

                elif isinstance(field, ForeignKey):
                    if new_column.lower() == "none":
                        setattr(item, group_by, None)
                    else:
                        related_model = field.related_model
                        try:
                            related_obj = related_model.objects.get(pk=new_column)
                            setattr(item, group_by, related_obj)
                        except related_model.DoesNotExist:
                            return HttpResponse(
                                status=400,
                                content=f"Invalid related object: {new_column}",
                            )

                item.save()

            except view.model.DoesNotExist:
                messages.error(request, f"Item Not found")
                return HttpResponse("<script>$('#reloadButton').click();")

            # Reconstruct query parameters
            query_params = QueryDict(mutable=True)
            for key, values in request.POST.lists():
                if key not in [
                    "item_id",
                    "new_column",
                    "app_label",
                    "model_name",
                    "class_name",
                    "csrfmiddlewaretoken",
                ]:
                    for value in values:
                        query_params.appendlist(key, value)

            # FIXED: Use the complete get_queryset logic instead of basic filtering
            view.request.GET = query_params

            # Apply the full queryset logic from HorillaListView
            view.object_list = view.get_queryset()

            # Get context
            context = view.get_context_data()
            context["app_label"] = app_label
            context["model_name"] = model_name
            context["class_name"] = class_name

            rendered_content = render_to_string(
                "partials/kanban_blocks.html", context, request=request
            )

            main_url = getattr(
                view, "main_url", f"/horilla_generics/kanban/{app_label}/{model_name}/"
            )
            response = HttpResponse(rendered_content)
            new_query_string = query_params.urlencode()
            url = main_url + (f"?{new_query_string}" if new_query_string else "")
            response["HX-Push-Url"] = url
            return response

        except Exception as e:
            return HttpResponse(status=500, content=f"Error: {str(e)}")

    def update_kanban_column_order(self, request):
        app_label = request.POST.get("app_label")
        model_name = request.POST.get("model_name")
        class_name = request.POST.get("class_name")
        column_order = request.POST.get("column_order")

        # Validate required parameters

        try:
            # Dynamically import the view class
            view_class = HorillaKanbanView._view_registry.get(self.model)
            if not view_class:
                return HttpResponse(
                    status=404, content=f"View class {class_name} not found"
                )

            # Instantiate the view class
            view = view_class()
            view.request = request
            view.model = apps.get_model(
                app_label=app_label.split(".")[-1], model_name=model_name
            )
            main_url = getattr(view, "main_url")

            group_by = view.get_group_by_field()
            try:
                field = view.model._meta.get_field(group_by)
                if not isinstance(field, ForeignKey):
                    return HttpResponse(
                        status=400,
                        content="Column ordering is only supported for ForeignKey fields.",
                    )

                related_model = field.related_model
                if "order" not in [f.name for f in related_model._meta.get_fields()]:
                    return HttpResponse(
                        status=400,
                        content=f"Related model {related_model.__name__} does not support ordering",
                    )
            except Exception as e:
                return HttpResponse(
                    status=400,
                    content=f"Invalid group_by field: {group_by}",
                )

            try:
                column_order = json.loads(column_order)
                if not isinstance(column_order, list):
                    raise ValueError("column_order must be a list")
            except json.JSONDecodeError as e:
                return HttpResponse(status=400, content="Invalid column_order format")

            try:
                with transaction.atomic():
                    max_order = (
                        related_model.objects.aggregate(Max("order"))["order__max"] or 0
                    )
                    temp_offset = max_order + 1000
                    valid_pks = []
                    for index, column_key in enumerate(column_order):
                        if column_key == "None":
                            continue
                        try:
                            related_obj = related_model.objects.get(pk=column_key)
                            related_obj.order = temp_offset + index
                            related_obj.save()
                            valid_pks.append(column_key)
                        except related_model.DoesNotExist:
                            continue

                    for index, column_key in enumerate(valid_pks):
                        related_obj = related_model.objects.get(pk=column_key)
                        related_obj.order = index
                        related_obj.save()

            except IntegrityError as e:
                return HttpResponse(
                    status=400,
                    content="Failed to update column order due to a unique constraint violation.",
                )
            except Exception as e:
                return HttpResponse(status=500, content=f"Error: {str(e)}")

            # Reconstruct query parameters
            query_params = QueryDict(mutable=True)
            for key, values in request.POST.lists():
                if key not in [
                    "column_order",
                    "app_label",
                    "model_name",
                    "class_name",
                    "csrfmiddlewaretoken",
                ]:
                    for value in values:
                        query_params.appendlist(key, value)

            view.request.GET = query_params
            view.object_list = view.get_queryset()

            context = view.get_context_data()
            context["app_label"] = app_label
            context["apps_label"] = app_label.split(".")[-1] if app_label else ""
            context["model_name"] = model_name
            context["class_name"] = class_name

            # Render response
            rendered_content = render_to_string(
                "partials/kanban_blocks.html", context, request=request
            )
            if not rendered_content.strip():
                return HttpResponse(
                    status=500, content="Error: Empty template response"
                )

            response = HttpResponse(rendered_content)
            new_query_string = query_params.urlencode()
            url = main_url + (f"?{new_query_string}" if new_query_string else "")
            response["HX-Push-Url"] = url
            return response

        except Exception as e:
            return HttpResponse(status=500, content=f"Error: {str(e)}")

    def get_group_by_field(self):
        model_name = self.model.__name__
        app_label = self.model._meta.app_label
        default_group = KanbanGroupBy.all_objects.filter(
            model_name=model_name, app_label=app_label, user=self.request.user
        ).first()
        return default_group.field_name if default_group else self.group_by_field

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if not hasattr(self, "object_list"):
            self.object_list = self.get_queryset()

        queryset = self.object_list
        group_by = self.get_group_by_field()

        app_label = self.model.__module__.rsplit(".", 1)[0] if self.model else ""
        model_name = self.model.__name__ if self.model else ""
        context["app_label"] = app_label
        context["apps_label"] = app_label.split(".")[-1] if app_label else ""
        context["model_name"] = model_name
        context["kanban_attrs"] = self.kanban_attrs
        context["class_name"] = self.__class__.__name__
        context["height_kanban"] = self.height_kanban
        if not group_by:
            context["error"] = "No grouping field specified."
            return context

        try:
            field = self.model._meta.get_field(group_by)
            if not (
                (hasattr(field, "choices") and field.choices)
                or isinstance(field, ForeignKey)
            ):
                context["error"] = (
                    f"Field '{group_by}' is not a Choice field or ForeignKey field."
                )
                return context

            allow_column_reorder = False
            has_colour_field = False
            if isinstance(field, ForeignKey):
                related_model = field.related_model
                related_fields = [f.name for f in related_model._meta.fields]
                allow_column_reorder = "order" in related_fields
                has_colour_field = "color" in related_fields

            context["group_by_field"] = group_by
            context["group_by_label"] = field.verbose_name
            context["allow_column_reorder"] = allow_column_reorder

            grouped_items = {}
            paginated_groups = {}

            if hasattr(field, "choices") and field.choices:
                num_columns = len(field.choices)
                for value, label in field.choices:
                    grouped_items[value] = {
                        "label": label,
                        "items": queryset.filter(**{group_by: value}),
                    }
                existing_values = set(queryset.values_list(group_by, flat=True))
                for value in existing_values:
                    if value not in grouped_items:
                        grouped_items[value] = {
                            "label": f"Unknown ({value})",
                            "items": queryset.filter(**{group_by: value}),
                        }

                sorted_items = {}
                for value, _ in field.choices:
                    if value in grouped_items:
                        sorted_items[value] = grouped_items[value]
                for key, group in grouped_items.items():
                    if key not in sorted_items:
                        sorted_items[key] = group

                for key, group in sorted_items.items():
                    total_count = group["items"].count()
                    ordered_items = group["items"].order_by(
                        "id"
                    )  # Use 'id' or 'created_at'
                    paginator = Paginator(ordered_items, self.paginate_by)
                    page = self.request.GET.get(f"page_{key}", 1)
                    try:
                        page_obj = paginator.page(page)
                    except PageNotAnInteger:
                        page_obj = paginator.page(1)
                    except EmptyPage:
                        page_obj = paginator.page(paginator.num_pages)
                    paginated_groups[key] = {
                        "label": group["label"],
                        "items": page_obj.object_list,
                        "page_obj": page_obj,
                        "has_next": page_obj.has_next(),
                        "next_page": (
                            page_obj.next_page_number() if page_obj.has_next() else None
                        ),
                        "total_count": total_count,
                    }

            elif isinstance(field, ForeignKey):
                queryset = queryset.prefetch_related(group_by)
                related_model = field.related_model
                if "order" in [f.name for f in related_model._meta.fields]:
                    related_items = related_model.objects.all().order_by("order")
                else:
                    related_items = related_model.objects.all().order_by("pk")

                for related_item in related_items:
                    grouped_items[related_item.pk] = {
                        "label": str(related_item),
                        "items": queryset.filter(
                            **{f"{group_by}__pk": related_item.pk}
                        ),
                        "color": (
                            getattr(related_item, "color", None)
                            if has_colour_field
                            else None
                        ),
                    }

                if field.null:
                    grouped_items[None] = {
                        "label": "None",
                        "items": queryset.filter(**{f"{group_by}__isnull": True}),
                        "color": None,
                    }
                    num_columns = len(related_items) + 1
                else:
                    num_columns = len(related_items)

                if None in grouped_items and not grouped_items[None]["items"].exists():
                    del grouped_items[None]
                    if field.null:
                        num_columns -= 1

                sorted_items = {}
                for related_item in related_items:
                    if related_item.pk in grouped_items:
                        sorted_items[related_item.pk] = grouped_items[related_item.pk]
                if None in grouped_items:
                    sorted_items[None] = grouped_items[None]

                for key, group in sorted_items.items():
                    total_count = group["items"].count()
                    ordered_items = group["items"].order_by(
                        "id"
                    )  # Use 'id' or 'created_at'
                    paginator = Paginator(ordered_items, self.paginate_by)
                    page = self.request.GET.get(f"page_{key}", 1)
                    try:
                        page_obj = paginator.page(page)
                    except PageNotAnInteger:
                        page_obj = paginator.page(1)
                    except EmptyPage:
                        page_obj = paginator.page(paginator.num_pages)
                    paginated_groups[key] = {
                        "label": group["label"],
                        "items": page_obj.object_list,
                        "page_obj": page_obj,
                        "has_next": page_obj.has_next(),
                        "next_page": (
                            page_obj.next_page_number() if page_obj.has_next() else None
                        ),
                        "total_count": total_count,
                        "colour": group["color"],
                    }

            display_columns = []
            for verbose_name, field_name in self.columns:
                if field_name != group_by:
                    display_columns.append({"name": field_name, "label": verbose_name})
            for key, group in paginated_groups.items():
                group["count"] = len(group["items"])
                for item in group["items"]:
                    item.can_drag = self.can_user_modify_item(item)
                    item.display_columns = []
                    for column in display_columns:
                        field_name = column["name"]
                        value = None
                        if hasattr(item, field_name):
                            value = getattr(item, field_name)
                            if callable(value):
                                value = value()
                        item.display_columns.append(
                            {
                                "name": field_name,
                                "label": column["label"],
                                "value": value,
                            }
                        )

            context.update(
                {
                    "grouped_items": paginated_groups,
                    "display_columns": display_columns,
                    "num_columns": num_columns,
                    "model_name": model_name,
                    "app_label": app_label,
                    "apps_label": app_label.split(".")[-1] if app_label else "",
                    "columns": self.columns,
                    "actions": self.actions,
                    "filter_class": self.filterset_class.__name__,
                    "group_by_field": group_by,
                    "kanban_attrs": self.kanban_attrs,
                }
            )
        except FieldError as e:
            context["error"] = f"Invalid grouping field '{group_by}': {str(e)}"
        except Exception as e:
            context["error"] = f"Error grouping by field '{group_by}': {str(e)}"
        return context

    def load_more_items(self, request, *args, **kwargs):
        column_key = request.GET.get("column_key")
        page = request.GET.get("page")
        group_by = self.get_group_by_field()

        if not page or not group_by:
            return HttpResponse(status=400, content="Missing required parameters")

        try:
            field = self.model._meta.get_field(group_by)
            if column_key == "None":
                column_key = None
            elif isinstance(field, ForeignKey) and column_key and column_key.isdigit():
                column_key = int(column_key)

            queryset = self.get_queryset()

            if hasattr(field, "choices") and field.choices:
                items = queryset.filter(**{group_by: column_key}).order_by(
                    "id"
                )  # Ensure ordering
            elif isinstance(field, ForeignKey):
                if column_key is None:
                    items = queryset.filter(**{f"{group_by}__isnull": True}).order_by(
                        "id"
                    )
                else:
                    items = queryset.filter(**{f"{group_by}__pk": column_key}).order_by(
                        "id"
                    )

            paginate_by = getattr(self, "paginate_by", 10)
            paginator = Paginator(items, paginate_by)
            try:
                page_obj = paginator.page(page)
            except PageNotAnInteger:
                page_obj = paginator.page(1)
            except EmptyPage:
                return HttpResponse("")  # Return empty response for no more items

            display_columns = []
            for verbose_name, field_name in self.columns:
                if field_name != group_by:
                    display_columns.append({"name": field_name, "label": verbose_name})

            for item in page_obj.object_list:
                item.can_drag = self.can_user_modify_item(item)
                item.display_columns = []
                for column in display_columns:
                    field_name = column["name"]
                    try:
                        value = getattr(item, field_name)
                        if callable(value):
                            value = value()
                    except AttributeError:
                        value = None
                    item.display_columns.append(
                        {
                            "name": field_name,
                            "label": column["label"],
                            "value": str(value) if value is not None else "N/A",
                        }
                    )

            context = {
                "group": {
                    "items": page_obj.object_list,
                    "has_next": page_obj.has_next(),
                    "next_page": (
                        page_obj.next_page_number() if page_obj.has_next() else None
                    ),
                    "label": str(column_key) if column_key else "None",
                },
                "actions": getattr(self, "actions", []),
                "column_key": column_key,
                "class_name": self.__class__.__name__,
                "app_label": (
                    self.model.__module__.rsplit(".", 1)[0] if self.model else ""
                ),
                "apps_label": self.model.__module__.split(".")[1],
                "model_name": self.model.__name__ if self.model else "",
                "key": column_key,
                "kanban_attrs": self.kanban_attrs,
            }

            return HttpResponse(
                render_to_string("partials/kanban_items.html", context, request=request)
            )
        except Exception as e:
            return HttpResponse(status=500, content=f"Error: {str(e)}")


class HorillaDetailView(DetailView):

    template_name = "detail_view.html"
    context_object_name = "obj"
    body: list = []
    excluded_fields = []
    pipeline_field = ""
    breadcrumbs = []
    actions = []
    tab_url: str = ""
    final_stage_action = {}

    _view_registry = {}

    def __init_subclass__(cls, **kwargs):
        """
        Automatically register child classes with their models.
        This allows the parent to find the correct child class dynamically.
        """
        super().__init_subclass__(**kwargs)
        if hasattr(cls, "model") and cls.model:
            HorillaDetailView._view_registry[cls.model] = cls

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect_to_login(request.get_full_path())
        try:
            model_name = request.POST.get("model_name") or request.GET.get("model_name")
            app_label = request.POST.get("app_label") or request.GET.get("app_label")

            if model_name and app_label:
                self.model = apps.get_model(app_label=app_label, model_name=model_name)
            self.object = self.get_object()

        except Exception as e:
            if request.headers.get("HX-Request") == "true":
                messages.error(self.request, e)
                return HttpResponse(headers={"HX-Refresh": "true"})
            raise HorillaHttp404(e)
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        if not self.request.user.is_authenticated:
            login_url = f"{reverse_lazy('horilla_core:login')}?next={request.path}"
            return redirect(login_url)
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        if not self.model:
            raise HorillaHttp404("Model not found")
        return super().get_queryset()

    def get_body(self):
        """Return a normalized list of (verbose_name, field_name) pairs."""
        normalized_body = []
        if not self.body:
            return normalized_body

        field_permissions = get_field_permissions_for_model(
            self.request.user, self.model
        )

        instance = self.model()
        for field in self.body:
            field_name = field[1] if isinstance(field, tuple) else field

            field_perm = field_permissions.get(field_name, "readwrite")
            if field_perm == "hidden":
                continue

            if isinstance(field, tuple):
                normalized_body.append(field)
            else:
                try:
                    model_field = instance._meta.get_field(field)
                    normalized_body.append((model_field.verbose_name, field))
                except FieldDoesNotExist:
                    pass
        return normalized_body

    def check_update_permission(self):
        """
        Check if the current user has permission to update the pipeline field.
        Returns True if user has permission, False otherwise.
        """
        user = self.request.user
        current_obj = self.get_object()
        model_name = self.model._meta.model_name
        app_label = self.model._meta.app_label

        # Superuser always has permission
        if user.is_superuser:
            return True

        # Check if user is the owner
        is_owner = False
        owner_fields = getattr(self.model, "OWNER_FIELDS", [])

        for owner_field in owner_fields:
            try:
                field_value = getattr(current_obj, owner_field, None)
                if field_value:
                    # Handle ManyToMany fields
                    if hasattr(field_value, "all"):
                        if user in field_value.all():
                            is_owner = True
                            break
                    # Handle ForeignKey fields
                    elif field_value == user:
                        is_owner = True
                        break
            except Exception:
                continue

        # Check change_own permission if user is owner
        if is_owner:
            change_own_perm = f"{app_label}.change_own_{model_name}"
            if user.has_perm(change_own_perm):
                return True

        # Check regular change permission
        change_perm = f"{app_label}.change_{model_name}"
        if user.has_perm(change_perm):
            return True

        return False

    def get_pipeline_choices(self):
        """
        Generate pipeline data for the specified pipeline_field.
        Returns a list of tuples: (display_name, value, is_completed, is_current).
        - For choice fields: Use choices defined in the model.
        - For foreign keys: Use related objects, ordered by the 'order' field.
        - is_completed: True if the stage's order is < the current value's order.
        - is_current: True if this is the current stage.
        """
        if not self.pipeline_field:
            return []
        try:
            obj = self.get_object()
        except Http404:
            return render(self.request, "error/403.html")
        field = self.model._meta.get_field(self.pipeline_field)
        current_value = getattr(obj, self.pipeline_field)

        pipeline = []
        if hasattr(field, "choices") and field.choices:
            current_choice_index = None
            for i, (value, display_name) in enumerate(field.choices):
                if value == current_value:
                    current_choice_index = i
                    break

            for i, (value, display_name) in enumerate(field.choices):
                is_completed = (
                    current_choice_index is not None and i < current_choice_index
                )
                is_current = value == current_value
                is_final = False
                pipeline.append(
                    (display_name, value, is_completed, is_current, is_final)
                )

        elif isinstance(field, ForeignKey):
            related_model = field.related_model
            order_field = None
            try:
                order_field = related_model._meta.get_field("order")
            except:
                pass
            queryset = related_model.objects.all()
            if order_field:
                queryset = queryset.order_by("order")

            current_order = (
                getattr(current_value, "order", None) if current_value else None
            )
            current_id = current_value.id if current_value else None

            for related_obj in queryset:
                is_completed = False
                is_current = related_obj.id == current_id
                is_final = getattr(related_obj, "is_final", False)
                if current_order is not None:
                    related_order = getattr(related_obj, "order", None)
                    is_completed = (
                        related_order is not None and related_order < current_order
                    )
                pipeline.append(
                    (
                        str(related_obj),
                        related_obj.id,
                        is_completed,
                        is_current,
                        is_final,
                    )
                )
        else:
            return []

        return pipeline

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["body"] = self.get_body()
        context["pipeline_choices"] = self.get_pipeline_choices()
        current_obj = self.get_object()
        current_id = current_obj.id
        context["tab_url"] = self.tab_url
        field_permissions = get_field_permissions_for_model(
            self.request.user, self.model
        )
        context["can_update"] = self.check_update_permission()
        context["field_permissions"] = field_permissions
        if hasattr(self, "final_stage_action"):
            final_stage = self.final_stage_action
            if callable(final_stage):
                context["final_stage_action"] = final_stage()
            else:
                context["final_stage_action"] = final_stage
        else:
            context["final_stage_action"] = self.final_stage_action

        session_key = f"list_view_queryset_ids_{self.model._meta.model_name}"

        queryset_ids = self.request.session.get(session_key, [])
        if not queryset_ids:
            from horilla_generics.views import HorillaListView

            list_view = HorillaListView()
            list_view.request = self.request
            list_view.model = self.model
            queryset = list_view.get_queryset()
            queryset_ids = list(queryset.values_list("id", flat=True))
            self.request.session["list_view_queryset_ids"] = queryset_ids
        try:
            current_index = queryset_ids.index(current_id)
        except ValueError:
            current_index = -1
        context["has_previous"] = current_index > 0
        context["has_next"] = current_index < len(queryset_ids) - 1
        context["previous_id"] = (
            queryset_ids[current_index - 1] if context["has_previous"] else None
        )
        context["next_id"] = (
            queryset_ids[current_index + 1] if context["has_next"] else None
        )
        url = resolve(self.request.path)
        context["url_name"] = url.url_name
        context["app_label"] = self.model._meta.app_label

        # Session keys for storing breadcrumb state
        breadcrumbs_session_key = (
            f"detail_view_breadcrumbs_{self.model._meta.model_name}_{current_id}"
        )
        referer_session_key = (
            f"detail_referer_{self.model._meta.model_name}_{current_id}"
        )

        hx_current_url = self.request.headers.get("HX-Current-URL")
        http_referer = self.request.META.get("HTTP_REFERER")

        is_reload = False
        if hx_current_url:
            current_path = urlparse(hx_current_url).path
            is_reload = current_path == self.request.path

        if is_reload:
            stored_breadcrumbs = self.request.session.get(breadcrumbs_session_key)
            if stored_breadcrumbs:

                breadcrumbs_for_context = (
                    stored_breadcrumbs[:-1] if stored_breadcrumbs else []
                )
                breadcrumbs_for_context.append((current_obj, None))
                context["breadcrumbs"] = breadcrumbs_for_context
                context["actions"] = self.actions
                context["model_name"] = self.model._meta.model_name
                if self.pipeline_field:
                    context["pipeline_field"] = self.pipeline_field
                    context["pipeline_field_verbose_name"] = self.model._meta.get_field(
                        self.pipeline_field
                    ).verbose_name
                return context

        breadcrumbs = []
        stored_referer = self.request.session.get(referer_session_key)

        if hx_current_url and not is_reload:
            referer = hx_current_url
            referer_path = urlparse(referer).path
            if referer_path != self.request.path:
                self.request.session[referer_session_key] = referer
        elif stored_referer:
            referer = stored_referer
        else:
            referer = http_referer
            if referer:
                referer_path = urlparse(referer).path
                if referer_path != self.request.path:
                    self.request.session[referer_session_key] = referer

        dynamic_breadcrumbs = []
        if referer:
            referer_path = urlparse(referer).path
            if referer_path != self.request.path:
                try:
                    resolved = resolve(referer_path)

                    referer_view = (
                        resolved.func.view_class
                        if hasattr(resolved.func, "view_class")
                        else None
                    )
                    is_detail_view = referer_view and issubclass(
                        referer_view, HorillaDetailView
                    )

                    if is_detail_view:
                        session_breadcrumbs = self.request.session.get(
                            "detail_view_breadcrumbs", []
                        )
                        breadcrumbs.extend(session_breadcrumbs)
                    else:
                        label = (
                            resolved.url_name.replace("_", " ")
                            .replace("-", " ")
                            .title()
                            if resolved.url_name
                            else "Back"
                        )
                        for suffix in [
                            " View",
                            " Detail",
                            " List",
                            " Create",
                            " Update",
                            " Delete",
                        ]:
                            if label.endswith(suffix):
                                label = label[: -len(suffix)]
                                break
                        breadcrumbs.append((label, referer))
                except:
                    breadcrumbs.append(("Back", referer))

            dynamic_breadcrumbs = breadcrumbs.copy()

            referrer_app = self.request.GET.get("referrer_app")
            referrer_model = self.request.GET.get("referrer_model")
            referrer_id = self.request.GET.get("referrer_id")
            referrer_label = self.request.GET.get("referrer_label")
            referrer_url = self.request.GET.get("referrer_url")

            if referrer_app and referrer_model and referrer_id:
                if not (
                    referrer_model == self.model._meta.model_name
                    and str(referrer_id) == str(current_id)
                ):
                    try:
                        model_class = apps.get_model(
                            app_label=referrer_app, model_name=referrer_model
                        )
                        obj = model_class.objects.get(pk=referrer_id)
                        obj_title = (
                            str(obj)
                            if hasattr(obj, "__str__")
                            else referrer_label or f"{referrer_model} {referrer_id}"
                        )
                        if referrer_url:
                            try:
                                url = reverse(
                                    f"{referrer_app}:{referrer_url}",
                                    kwargs={"pk": referrer_id},
                                )

                                parsed_url = urlparse(url)
                                query_dict = parse_qs(parsed_url.query)

                                section_for_breadcrumb = None
                                if referrer_app and referrer_model:
                                    try:
                                        model_class = apps.get_model(
                                            app_label=referrer_app,
                                            model_name=referrer_model,
                                        )
                                        section_info = get_section_info_for_model(
                                            model_class
                                        )
                                        section_for_breadcrumb = section_info.get(
                                            "section"
                                        )
                                    except Exception as e:
                                        section_for_breadcrumb = None

                                if not section_for_breadcrumb:
                                    section_for_breadcrumb = self.request.GET.get(
                                        "section"
                                    )

                                if section_for_breadcrumb:
                                    query_dict["section"] = [section_for_breadcrumb]

                                new_query = urlencode(query_dict, doseq=True)
                                url = urlunparse(parsed_url._replace(query=new_query))

                            except Exception:
                                url = None
                        dynamic_breadcrumbs.append((obj_title, url))
                    except (LookupError, model_class.DoesNotExist, ValueError) as e:
                        if referrer_label and referrer_url:
                            dynamic_breadcrumbs.append((referrer_label, referrer_url))

        dynamic_breadcrumbs.append((current_obj, None))

        session_url_value = self.request.GET.get("session_url")
        if session_url_value:
            updated_breadcrumbs = []
            for label, url in dynamic_breadcrumbs:
                if url:
                    try:
                        parsed_url = urlparse(url)
                        query_dict = parse_qs(parsed_url.query)
                        query_dict["session_url"] = [session_url_value]
                        new_query = urlencode(query_dict, doseq=True)
                        url = urlunparse(parsed_url._replace(query=new_query))
                    except Exception:
                        pass  # Keep original URL if parsing fails
                updated_breadcrumbs.append((label, url))
            dynamic_breadcrumbs = updated_breadcrumbs

        self.request.session["detail_view_breadcrumbs"] = breadcrumbs

        serializable_breadcrumbs = []
        for label, url in dynamic_breadcrumbs:
            if hasattr(label, "_meta"):  # It's a model instance
                label = str(label)
            serializable_breadcrumbs.append((label, url))

        self.request.session[breadcrumbs_session_key] = serializable_breadcrumbs

        context["breadcrumbs"] = dynamic_breadcrumbs
        context["actions"] = self.actions
        context["model_name"] = self.model._meta.model_name
        if self.pipeline_field:
            context["pipeline_field"] = self.pipeline_field
            context["pipeline_field_verbose_name"] = self.model._meta.get_field(
                self.pipeline_field
            ).verbose_name
        return context

    def post(self, request, *args, **kwargs):
        """
        Handle POST requests for updating the pipeline field.
        """
        if request.POST.get("pipeline_update"):
            model_name = request.POST.get("model_name")
            app_label = request.POST.get("app_label")
            pipeline_field = request.POST.get("pipeline_field")

            try:
                model = apps.get_model(app_label, model_name)
            except Exception as e:
                messages.error(self.request, e)
                return HttpResponse("<script>$('#reloadButton').click();</script>")

            view_class = self._view_registry.get(model, self.__class__)

            if view_class != self.__class__:
                view_instance = view_class()
                view_instance.request = self.request
                view_instance.args = self.args
                view_instance.kwargs = self.kwargs
                view_instance.model = model
                view_instance.pipeline_field = pipeline_field
                return view_instance.update_pipeline(request, *args, **kwargs)

            self.model = model
            self.pipeline_field = pipeline_field
            return self.update_pipeline(request, *args, **kwargs)

        return HttpResponse(status=400)

    def update_pipeline(self, request, *args, **kwargs):
        """
        Handle HTMX POST request to update the pipeline field value.
        Re-render the Kanban choices template with updated data.
        """
        self.object = self.get_object()

        pipeline_value = request.POST.get("pipeline_value")
        if not pipeline_value:
            return HttpResponse(status=400)

        try:
            # Permission check
            user = request.user
            model_name = self.model._meta.model_name
            app_label = self.model._meta.app_label

            # Check if user is the owner
            is_owner = False
            owner_fields = getattr(self.model, "OWNER_FIELDS", [])

            for owner_field in owner_fields:
                try:
                    field_value = getattr(self.object, owner_field, None)
                    if field_value:
                        # Handle ManyToMany fields
                        if hasattr(field_value, "all"):
                            if user in field_value.all():
                                is_owner = True
                                break
                        # Handle ForeignKey fields
                        elif field_value == user:
                            is_owner = True
                            break
                except Exception:
                    continue

            # Check permissions
            has_permission = False

            if user.is_superuser:
                has_permission = True
            elif is_owner:
                # Check if user has change_own permission
                change_own_perm = f"{app_label}.change_own_{model_name}"
                if user.has_perm(change_own_perm):
                    has_permission = True

            # Check regular change permission if not owner or doesn't have change_own
            if not has_permission:
                change_perm = f"{app_label}.change_{model_name}"
                if user.has_perm(change_perm):
                    has_permission = True

            if not has_permission:
                messages.error(
                    self.request, _("You don't have permission to update this record.")
                )
                return HttpResponse("<script>$('#reloadButton').click();</script>")

            # Proceed with pipeline update
            field = self.model._meta.get_field(self.pipeline_field)
            if hasattr(field, "choices") and field.choices:
                # Validate for choice fields
                if pipeline_value not in [choice[0] for choice in field.choices]:
                    raise ValidationError("Invalid choice")
                setattr(self.object, self.pipeline_field, pipeline_value)
            elif isinstance(field, ForeignKey):
                # Validate for ForeignKey fields
                related_model = field.related_model
                try:
                    related_obj = related_model.objects.get(pk=pipeline_value)
                    setattr(self.object, self.pipeline_field, related_obj)
                except Exception as e:
                    logger.error(e)
            else:
                return HttpResponse(status=400)

            self.object.save()
            messages.success(
                self.request,
                _(f"{self.model._meta.verbose_name} Stage Updated Successfully"),
            )
            context = self.get_context_data(object=self.object)
            context["pipeline_update"] = True
            kanban_html = render_to_string(
                "partials/pipeline_choices.html", context, request=self.request
            )
            return HttpResponse(kanban_html)
        except Exception as e:
            messages.error(self.request, e)
            return HttpResponse("<script>$('#reloadButton').click();</script>")


@method_decorator(htmx_required, name="dispatch")
class HorillaTabView(TemplateView):
    """
    Generic TabView
    """

    view_id = ""
    template_name = "tab_view.html"
    tabs: list = []
    background_class = ""
    background_color = ""
    tab_class = ""

    def __init__(self, **kwargs: Any) -> None:
        super().__init__(**kwargs)
        request = getattr(_thread_local, "request", None)
        self.request = request

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user:
            active_tab = ActiveTab.objects.filter(
                created_by=self.request.user, path=self.request.path
            ).first()
            if active_tab:
                context["active_target"] = active_tab.tab_target
        context["tabs"] = self.tabs
        context["view_id"] = self.view_id
        context["background_class"] = self.background_class
        context["background_color"] = self.background_color
        context["tab_class"] = self.tab_class
        return context


@method_decorator(htmx_required, name="dispatch")
class HorillaDetailTabView(HorillaTabView):
    """
    Generic for tabs in detail views
    """

    view_id = "generic-details-tab-view"
    object_id = None
    urls = {}
    tab_class = "h-[calc(_100vh_-_475px_)] overflow-hidden vbvbvb"

    def __init__(self, **kwargs: Any) -> None:
        super().__init__(**kwargs)
        pipeline_field = self.request.GET.get("pipeline_field")
        user = self.request.user
        self.tabs = []
        if self.object_id:
            if "details" in self.urls:
                details_url = (
                    f"{reverse_lazy(self.urls['details'], kwargs={'pk': self.object_id})}?pipeline_field={pipeline_field}"
                    if pipeline_field
                    else reverse(self.urls["details"], kwargs={"pk": self.object_id})
                )
                self.tabs.append(
                    {
                        "title": _("Details"),
                        "url": details_url,
                        "target": "tab-details-content",
                        "id": "details",
                    }
                )
            if "activity" in self.urls:
                self.tabs.append(
                    {
                        "title": _("Activity"),
                        "url": reverse_lazy(
                            self.urls["activity"], kwargs={"pk": self.object_id}
                        ),
                        "target": "tab-activity-content",
                        "id": "activity",
                    }
                )

            if "related_lists" in self.urls:
                self.tabs.append(
                    {
                        "title": _("Related Lists"),
                        "url": f"{reverse_lazy(self.urls['related_lists'], kwargs={'pk': self.object_id})}",
                        "target": "tab-related-lists-content",
                        "id": "related-lists",
                    }
                )

            if "notes_attachments" in self.urls and (
                user.has_perm("horilla_core.view_horillaattachment")
                or user.has_perm("horilla_core.view_own_horillaattachment")
                or user.is_superuser
            ):

                self.tabs.append(
                    {
                        "title": _("Notes & Attachments"),
                        "url": f"{reverse_lazy(self.urls['notes_attachments'], kwargs={'pk': self.object_id})}",
                        "target": "tab-notes-attachments-content",
                        "id": "notes-attachments",
                    }
                )

            if "history" in self.urls:
                self.tabs.append(
                    {
                        "title": _("History"),
                        "url": reverse_lazy(
                            self.urls["history"], kwargs={"pk": self.object_id}
                        ),
                        "target": "tab-history-content",
                        "id": "history",
                    }
                )


@method_decorator(htmx_required, name="dispatch")
class HorillaDetailSectionView(DetailView):
    """
    A generic detail view that supports multiple tabs for displaying related objects.
    """

    template_name = "details_tab.html"
    context_object_name = "obj"
    body = []
    edit_field = True
    non_editable_fields = []
    excluded_fields = [
        "id",
        "created_at",
        "additional_info",
        "updated_at",
        "history",
        "is_active",
        "created_by",
        "updated_by",
    ]
    include_fields = []

    def get(self, request, *args, **kwargs):
        """
        Override get method to handle object not found before processing
        """
        try:
            self.object = self.get_object()
        except Exception as e:
            messages.error(self.request, e)
            return HttpResponse("<script>$('#reloadButton').click();</script>")

        context = self.get_context_data(object=self.object)
        return self.render_to_response(context)

    def get_default_body(self):
        """
        Dynamically generate body based on model fields.
        Exclude fields like 'id' or others you don't want to display.
        """
        excluded_fields = self.excluded_fields
        pipeline_field = self.request.GET.get("pipeline_field")
        if pipeline_field:
            excluded_fields.append(pipeline_field)

        if self.include_fields:
            return [
                (field.verbose_name, field.name)
                for field in self.model._meta.get_fields()
                if field.name in self.include_fields
                and field.name not in excluded_fields
                and hasattr(field, "verbose_name")
            ]
        return [
            (field.verbose_name, field.name)
            for field in self.model._meta.get_fields()
            if field.name not in excluded_fields and hasattr(field, "verbose_name")
        ]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["body"] = self.body or self.get_default_body()
        context["model_name"] = self.model._meta.model_name
        context["app_label"] = self.model._meta.app_label
        context["edit_field"] = self.edit_field
        context["non_editable_fields"] = self.non_editable_fields

        # field permissions context
        field_permissions = get_field_permissions_for_model(
            self.request.user, self.model
        )
        context["field_permissions"] = field_permissions
        context["can_update"] = HorillaDetailView.check_update_permission(self)
        pipeline_field = self.request.GET.get("pipeline_field")
        if pipeline_field:
            context["pipeline_field"] = pipeline_field
        return context


@method_decorator(htmx_required, name="dispatch")
class HorillaActivitySectionView(DetailView):
    """
    Generic Activity Tab View
    """

    template_name = "activity_tab.html"
    context_object_name = "obj"

    def dispatch(self, request, *args, **kwargs):
        try:
            self.object = self.get_object()
        except Exception as e:
            messages.error(self.request, e)
            return HttpResponse(headers={"HX-Refresh": "true"})
        return super().dispatch(request, *args, **kwargs)

    def add_task_button(self):
        return {
            "url": f"""{ reverse_lazy('activity:task_create_form')}""",
            "attrs": 'id="task-create"',
        }

    def add_meetings_button(self):
        return {
            "url": f"""{ reverse_lazy('activity:meeting_create_form')}""",
            "attrs": 'id="meeting-create"',
        }

    def add_call_button(self):
        return {
            "url": f"""{ reverse_lazy('activity:call_create_form')}""",
            "attrs": 'id="call-create"',
        }

    def add_email_button(self):
        return {
            "url": f"""{ reverse_lazy('horilla_mail:send_mail_view')}""",
            "attrs": 'id="email-create"',
            "title": _("Send Email"),
        }

    def add_event_button(self):
        return {
            "url": f"""{ reverse_lazy('activity:event_create_form')}""",
            "attrs": 'id="event-create"',
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pk = self.kwargs.get("pk")
        context["object_id"] = pk
        context["model_name"] = self.model._meta.model_name
        context["app_label"] = self.model._meta.app_label
        content_type = ContentType.objects.get_for_model(self.model)
        context["content_type_id"] = content_type.id
        context["add_task_button"] = self.add_task_button() or {}
        context["add_meetings_button"] = self.add_meetings_button() or {}
        context["add_call_button"] = self.add_call_button() or {}
        context["add_email_button"] = self.add_email_button() or {}
        context["add_event_button"] = self.add_event_button() or {}
        return context


@method_decorator(htmx_required, name="dispatch")
class HorillaRelatedListSectionView(DetailView):
    template_name = "related_list.html"
    context_object_name = "object"

    related_list_config = {}
    max_items_per_list = None
    excluded_related_lists = []

    _view_registry = {}

    def __init_subclass__(cls, **kwargs):
        """
        Automatically register child classes with their models.
        This allows the parent to find the correct child class dynamically.
        """
        super().__init_subclass__(**kwargs)
        if hasattr(cls, "model") and cls.model:
            HorillaRelatedListSectionView._view_registry[cls.model] = cls

    def get_related_lists_metadata(self):
        """
        Get metadata for related lists (for tab navigation), including custom related lists.
        """
        obj = self.get_object()
        related_lists = []

        related_config = getattr(self, "related_list_config", {})
        if isinstance(related_config, functools.cached_property):
            try:
                related_config = related_config.func(self)
            except Exception:
                related_config = {}
        related_config = related_config if isinstance(related_config, dict) else {}
        # Standard related fields
        for field in obj._meta.get_fields():
            if not self.is_valid_related_field(field):
                continue

            related_model = field.related_model
            config = related_config.get(field.name, {})
            default_title = related_model._meta.verbose_name_plural.title()

            related_lists.append(
                {
                    "name": field.name,
                    "title": config.get("title", default_title),
                    "model_name": related_model.__name__,
                    "app_label": related_model._meta.app_label,
                    "parent_model_name": obj._meta.model_name,
                    "is_custom": False,
                    "config": config,
                }
            )

        for custom_name, custom_config in related_config.get(
            "custom_related_lists", {}
        ).items():
            try:
                related_model = apps.get_model(
                    custom_config["app_label"], custom_config["model_name"]
                )
                default_title = related_model._meta.verbose_name_plural.title()

                related_lists.append(
                    {
                        "name": custom_name,
                        "title": custom_config.get("config").get(
                            "title", default_title
                        ),
                        "model_name": related_model.__name__,
                        "app_label": related_model._meta.app_label,
                        "parent_model_name": obj._meta.model_name,
                        "is_custom": True,
                        "custom_config": custom_config,
                    }
                )
            except LookupError:
                continue
        return related_lists

    def get_single_related_list(self, obj, field_name):
        """
        Get data for a single related list, handling both standard and custom related lists
        """
        related_config = getattr(self, "related_list_config", {})
        if isinstance(related_config, functools.cached_property):
            try:
                related_config = related_config.func(self)
            except Exception as e:
                related_config = {}
        elif not isinstance(related_config, dict):
            related_config = {}
        custom_related_lists = related_config.get("custom_related_lists", {})

        if field_name in custom_related_lists:
            return self.build_custom_related_list_data(
                obj, field_name, custom_related_lists[field_name]
            )

        for field in obj._meta.get_fields():
            if field.name == field_name and self.is_valid_related_field(field):
                return self.build_related_list_data(obj, field)
        return None

    def get_related_lists(self):
        """
        Dynamically discover all related models, including custom ones
        """
        obj = self.get_object()
        related_lists = []

        for field in obj._meta.get_fields():
            if self.is_valid_related_field(field):
                related_list_data = self.build_related_list_data(obj, field)
                if related_list_data:
                    related_lists.append(related_list_data)

        related_config = getattr(self, "related_list_config", {})

        if isinstance(related_config, functools.cached_property):
            try:
                related_config = related_config.func(self)
            except Exception as e:
                related_config = {}

        if not isinstance(related_config, dict):
            related_config = {}

        custom_related_lists = related_config.get("custom_related_lists", {})
        for custom_name, custom_config in custom_related_lists.items():
            related_list_data = self.build_custom_related_list_data(
                obj, custom_name, custom_config
            )
            if related_list_data:
                related_lists.append(related_list_data)

        return related_lists

    def is_valid_related_field(self, field):
        """
        Check if field should be included in related lists, respecting exclusions
        """
        excluded_fields = [
            "history",
            "logentry",
            "log_entries",
            "audit_log",
            "auditlog",
            "activity_log",
            "change_log",
            "revisions",
        ] + self.excluded_related_lists

        return (
            (
                field.one_to_many
                or field.many_to_many
                or isinstance(field, GenericRelation)
            )
            and not field.name.startswith("_")
            and field.name.lower() not in excluded_fields
        )

    def build_related_list_data(self, obj, field):
        """
        Build data structure for a standard related list using HorillaListView
        """
        try:
            related_manager = getattr(obj, field.name)
            if hasattr(related_manager, "all"):
                queryset = related_manager.all()
            else:
                return None
            total_count = queryset.count()
            related_model = field.related_model
            model_name = related_model.__name__
            config = self.related_list_config.get(field.name, {})
            dropdown_actions = config.get("dropdown_actions", [])
            custom_buttons = config.get("custom_buttons", [])
            default_title = related_model._meta.verbose_name_plural.title()

            list_view = self.create_generic_list_view_instance(
                model=related_model,
                queryset=queryset[: self.max_items_per_list],
                config=config,
                view_id=field.name,
            )

            rendered_html = self.render_generic_list_view(list_view)

            return {
                "name": field.name,
                "title": config.get("title", default_title),
                "model": related_model,
                "model_name": model_name,
                "app_label": related_model._meta.app_label,
                "total_count": total_count,
                "can_add": config.get("can_add", True),
                "add_url": config.get("add_url", ""),
                "button_name": config.get("button_name"),
                "field_obj": field,
                "rendered_content": rendered_html,
                "dropdown_actions": dropdown_actions,
                "custom_buttons": custom_buttons,
                "is_custom": False,
            }
        except Exception as e:
            return None

    def build_custom_related_list_data(self, obj, custom_name, custom_config):
        """
        Build data structure for a custom related list with proper company filtering.
        This method ensures CompanyFilteredManager is invoked by querying the related model directly.
        """
        try:

            related_model = apps.get_model(
                custom_config["app_label"], custom_config["model_name"]
            )
            default_title = related_model._meta.verbose_name_plural.title()
            config = custom_config.get("config", {})
            dropdown_actions = config.get("dropdown_actions", [])

            queryset = None

            # Handle custom queryset function
            if "queryset" in custom_config:
                queryset = custom_config["queryset"](obj)

            # Handle intermediate model pattern
            elif "related_field" in custom_config:
                related_field = custom_config["related_field"]
                intermediate_field = custom_config["intermediate_field"]
                intermediate_model_name = custom_config.get("intermediate_model")

                if intermediate_model_name:
                    # Find the intermediate model across different app labels
                    intermediate_model = self._find_intermediate_model(
                        intermediate_model_name,
                        obj._meta.app_label,
                        custom_config["app_label"],
                    )

                    if intermediate_model:
                        # Step 1: Filter intermediate model by the parent object
                        intermediate_qs = intermediate_model.objects.filter(
                            **{related_field: obj}
                        )

                        # Step 2: Find the field in intermediate model that points to the related model
                        related_obj_field = self._get_related_field_from_intermediate(
                            intermediate_model,
                            related_model,
                            exclude_field=related_field,
                        )

                        if related_obj_field:
                            related_ids = list(
                                intermediate_qs.values_list(
                                    f"{related_obj_field}_id", flat=True
                                ).distinct()
                            )
                            queryset = related_model.objects.filter(pk__in=related_ids)

                            columns = config.get("columns", [])
                            annotations = self._build_intermediate_annotations(
                                intermediate_model,
                                intermediate_field,
                                related_field,
                                related_obj_field,
                                obj,
                                columns,
                            )

                            if annotations:
                                queryset = queryset.annotate(**annotations)
                        else:
                            # Fallback: couldn't find the field
                            queryset = related_model.objects.filter(
                                **{f"{intermediate_field}__{related_field}": obj}
                            )
                    else:
                        # Fallback: couldn't find intermediate model
                        queryset = related_model.objects.filter(
                            **{f"{intermediate_field}__{related_field}": obj}
                        )
                else:
                    # No intermediate_model specified, use direct relationship
                    queryset = related_model.objects.filter(
                        **{f"{intermediate_field}__{related_field}": obj}
                    )

            if queryset is None:
                return None

            total_count = queryset.count()

            list_view = self.create_generic_list_view_instance(
                model=related_model,
                queryset=queryset,
                config=config,
                view_id=custom_name,
            )
            rendered_html = self.render_generic_list_view(list_view)

            return {
                "name": custom_name,
                "title": config.get("title", default_title),
                "model": related_model,
                "model_name": related_model.__name__,
                "app_label": related_model._meta.app_label,
                "total_count": total_count,
                "can_add": config.get("can_add", True),
                "add_url": config.get("add_url", ""),
                "button_name": config.get("button_name"),
                "field_obj": None,
                "rendered_content": rendered_html,
                "dropdown_actions": dropdown_actions,
                "custom_buttons": config.get("custom_buttons", ""),
                "is_custom": True,
            }
        except Exception as e:
            import logging

            logger = logging.getLogger(__name__)
            logger.error(
                f"Error building custom related list {custom_name}: {str(e)}",
                exc_info=True,
            )
            return None

    def _find_intermediate_model(
        self, intermediate_model_name, obj_app_label, related_app_label
    ):
        """Helper to find intermediate model across different app labels."""
        app_labels_to_try = [obj_app_label, related_app_label]

        # Add other app labels
        for app_config in apps.get_app_configs():
            if app_config.label not in app_labels_to_try:
                app_labels_to_try.append(app_config.label)

        for app_label in app_labels_to_try:
            try:
                return apps.get_model(app_label, intermediate_model_name)
            except LookupError:
                continue

        return None

    def _get_related_field_from_intermediate(
        self, intermediate_model, related_model, exclude_field=None
    ):
        """Find the field in intermediate model pointing to related model."""
        for field in intermediate_model._meta.get_fields():
            if (
                hasattr(field, "related_model")
                and field.related_model == related_model
                and field.name != exclude_field
            ):
                return field.name
        return None

    def _build_intermediate_annotations(
        self,
        intermediate_model,
        intermediate_field,
        related_field,
        related_obj_field,
        obj,
        columns,
    ):
        """Build annotations for fields from the intermediate model."""
        from django.db.models import OuterRef, Subquery

        annotations = {}

        for col_verbose, col_field in columns:
            if "__" in col_field and col_field.startswith(intermediate_field):
                field_parts = col_field.split("__", 1)

                if len(field_parts) >= 2:
                    intermediate_field_name = field_parts[1]

                    # Extract the actual field name (before any display methods)
                    if "__" in intermediate_field_name:
                        value_field = intermediate_field_name.split("__")[0]
                    else:
                        # Handle get_*_display methods
                        if intermediate_field_name.startswith(
                            "get_"
                        ) and intermediate_field_name.endswith("_display"):
                            # Extract field name: get_member_status_display -> member_status
                            value_field = intermediate_field_name.replace(
                                "get_", ""
                            ).replace("_display", "")
                        else:
                            value_field = intermediate_field_name

                    subquery = intermediate_model.objects.filter(
                        **{related_field: obj, related_obj_field: OuterRef("pk")}
                    ).values(value_field)[:1]

                    annotations[col_field] = Subquery(subquery)

        return annotations

    def create_generic_list_view_instance(self, model, queryset, config, view_id=None):
        """
        Create and configure HorillaListView instance
        """
        section_info = get_section_info_for_model(model)
        section = section_info.get("section", "")

        col_attrs = config.get("col_attrs", [])
        for col_attr in col_attrs:
            for field_name, attrs in col_attr.items():
                if isinstance(attrs, dict):
                    for key, value in attrs.items():
                        if key in [
                            "hx-get",
                            "hx-post",
                            "hx-delete",
                            "href",
                        ] and isinstance(value, str):
                            value = re.sub(r"([&?])section=[^&]*", "", value)
                            value = value.replace("?&", "?").rstrip("&").rstrip("?")
                            separator = "&" if "?" in value else "?"
                            attrs[key] = f"{value}{separator}section={section}"

        list_view = HorillaListView()
        list_view.model = model
        list_view.request = self.request
        list_view.queryset = queryset
        columns = self.get_columns_for_model(model, config)
        list_view.columns = columns
        actions = config.get("actions", [])
        actions_method = config.get("action_method")
        list_view.actions = actions
        list_view.action_method = actions_method
        list_view.col_attrs = col_attrs
        list_view.bulk_select_option = False
        list_view.clear_session_button_enabled = False
        list_view.filterset_class = None
        list_view.table_width = False
        list_view.table_class = False
        list_view.view_id = f"{view_id}-content" if view_id else None
        list_view.main_url = self.request.path
        list_view.search_url = self.request.path
        list_view.table_height = False
        list_view.table_height_as_class = "h-[calc(_100vh_-_500px_)]"
        list_view.owner_filtration = False
        return list_view

    def render_generic_list_view(self, list_view):
        """
        Render HorillaListView and return HTML string
        """
        try:
            sorted_queryset = list_view.get_queryset()
            # Set object_list to the sorted QuerySet
            list_view.object_list = sorted_queryset
            context = list_view.get_context_data()

            return render_to_string(
                list_view.template_name, context, request=self.request
            )
        except Exception as e:
            return ""

    def get_columns_for_model(self, model, config):
        """
        Get columns to display for a model in the format [(verbose_name, field_name), ...]
        """
        if "columns" in config:
            return config["columns"]

        columns = []
        exclude_fields = config.get("exclude", [])
        default_exclude = [
            "id",
            "created_at",
            "additional_info",
            "updated_at",
            "history",
            "is_active",
            "created_by",
            "updated_by",
            "company",
        ]
        try:
            for field in model._meta.fields:
                if (
                    field.name not in default_exclude
                    and field.name not in exclude_fields
                ):
                    columns.append((field.verbose_name, field.name))
                    if len(columns) == 5:
                        break
        except Exception as e:
            return []

        return columns

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["related_lists_metadata"] = self.get_related_lists_metadata()
        context["object_id"] = self.object.pk
        context["class_name"] = self.__class__.__name__
        return context


@method_decorator(htmx_required, name="dispatch")
class HorillaRelatedListContentView(LoginRequiredMixin, DetailView):
    """
    View to handle HTMX GenericSingleDetailedViewequests for individual related list content
    """

    template_name = "related_list_content.html"
    context_object_name = "object"

    def get_parent_view_class(self, model, class_name):
        """
        Dynamically resolve the parent view class for the given model
        """
        try:
            view_class = HorillaRelatedListSectionView._view_registry.get(model)
            if view_class:
                return view_class
        except (ImportError, AttributeError) as e:
            logger.error(f"Error resolving view {class_name} in {model}{str(e)}")

        return HorillaRelatedListSectionView

    def get_queryset(self):
        """
        Dynamically resolve the model and app_label from model_name query parameter
        """
        model_name = self.request.GET.get("model_name")
        if not model_name:
            raise HorillaHttp404("model_name parameter is required")
        try:
            content_type = ContentType.objects.get(model=model_name.lower())
            app_label = content_type.app_label
            model = apps.get_model(app_label=app_label, model_name=model_name)
            return model.objects.all()
        except Exception as e:
            messages.error(self.request, e)
            raise HorillaHttp404(e)

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        field_name = request.GET.get("field_name")
        class_name = request.GET.get("class_name")

        if not field_name:
            return HttpResponse("Field name required", status=400)

        model = self.get_queryset().model
        parent_view_class = self.get_parent_view_class(model, class_name)
        parent_view = parent_view_class()
        parent_view.request = request
        parent_view.model = model
        parent_view.excluded_related_lists = getattr(
            parent_view_class, "excluded_related_lists", []
        )
        related_list_data = parent_view.get_single_related_list(self.object, field_name)
        if not related_list_data:
            return HttpResponse(
                f"No valid related field found for field_name: {field_name}", status=404
            )

        context = {
            "related_list": related_list_data,
            "object": self.object,
            "class_name": class_name,
        }

        return render(request, self.template_name, context)


@method_decorator(htmx_required, name="dispatch")
@method_decorator(
    permission_required_or_denied(
        [
            "horilla_core.view_horillaattachment",
            "horilla_core.view_own_horillaattachment",
        ]
    ),
    name="dispatch",
)
class HorillaNotesAttachementSectionView(DetailView):

    template_name = "notes_attachments.html"
    context_object_name = "obj"

    @cached_property
    def columns(self):
        """
        Define columns like in LeadListView
        """
        instance = HorillaAttachment()
        return [
            (instance._meta.get_field("title").verbose_name, "title"),
            (instance._meta.get_field("created_by").verbose_name, "created_by"),
            (instance._meta.get_field("created_at").verbose_name, "created_at"),
        ]

    def get_actions(self):
        """
        Return actions based on user permissions.
        """
        user = self.request.user
        actions = []

        if (
            user.is_superuser
            or user.has_perm("horilla_core.view_own_horillaattachment")
            or user.has_perm("horilla_core.view_horillaattachment")
        ):
            actions.append(
                {
                    "action": "View",
                    "src": "assets/icons/eye1.svg",
                    "img_class": "w-4 h-4",
                    "attrs": """
                            hx-get="{get_detail_view_url}"
                            hx-target="#contentModalBox"
                            hx-swap="innerHTML"
                            onclick="openContentModal()"
                            """,
                }
            )

        if self.check_change_attachment_permission():
            actions.append(
                {
                    "action": "Edit",
                    "src": "assets/icons/edit.svg",
                    "img_class": "w-4 h-4",
                    "attrs": """
                            hx-get="{get_edit_url}"
                            hx-target="#modalBox"
                            hx-swap="innerHTML"
                            hx-on:click="openModal();"
                            """,
                }
            )

        # Delete action - check delete_horillaattachment permission
        if user.is_superuser or user.has_perm("horilla_core.delete_horillaattachment"):
            actions.append(
                {
                    "action": "Delete",
                    "src": "assets/icons/a4.svg",
                    "img_class": "w-4 h-4",
                    "attrs": """
            hx-post="{get_delete_url}"
            hx-target="#deleteModeBox"
            hx-swap="innerHTML"
            hx-trigger="click"
            hx-vals='{{"check_dependencies": "true"}}'
            onclick="openDeleteModeModal()"
        """,
                }
            )

        return actions

    def check_change_attachment_permission(self):
        """
        Check if user has permission to edit attachments.
        Checks change_horillaattachment permission or change_own if user is owner.

        Returns:
            bool: True if user has permission, False otherwise
        """
        user = self.request.user
        if user.has_perm("horilla_core.change_horillaattachment"):
            return True

        if user.has_perm("horilla_core.change_own_horillaattachment"):
            return True

        return False

    def check_attachment_add_permission(self):
        """
        Check if user has permission to add attachments.
        Requires:
        1. Add permission on HorillaAttachment model
        2. Add or Change permission on the related object (or change_own if owner)

        Returns:
            bool: True if user has permission, False otherwise
        """
        user = self.request.user

        related_object = self.get_object()
        related_model = related_object.__class__
        model_name = related_model._meta.model_name
        app_label = related_model._meta.app_label

        # Check if user is the owner of the related object
        is_owner = False
        owner_fields = getattr(related_model, "OWNER_FIELDS", [])

        for owner_field in owner_fields:
            try:
                field_value = getattr(related_object, owner_field, None)
                if field_value:
                    # Handle ManyToMany fields
                    if hasattr(field_value, "all"):
                        if user in field_value.all():
                            is_owner = True
                            break
                    # Handle ForeignKey fields
                    elif field_value == user:
                        is_owner = True
                        break
            except Exception:
                continue

        if is_owner:
            change_own_perm = f"{app_label}.change_own_{model_name}"
            if user.has_perm(change_own_perm) and user.has_perm(
                "horilla_core.add_horillaattachment"
            ):
                return True

        change_perm = f"{app_label}.change_{model_name}"

        if user.has_perm(change_perm) and user.has_perm(
            "horilla_core.add_horillaattachment"
        ):
            return True

        return False

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        object_id = self.kwargs.get("pk")

        try:
            content_type = ContentType.objects.get_for_model(model=self.model)
        except ContentType.DoesNotExist:
            from django.http import HttpResponseNotFound

            return HttpResponseNotFound("Model not found")

        queryset = HorillaAttachment.objects.filter(
            content_type=content_type, object_id=object_id
        )

        list_view = HorillaListView(model=HorillaAttachment)

        list_view.request = self.request
        list_view.queryset = queryset
        list_view.columns = self.columns
        list_view.view_id = f"attachments_{content_type.model}_{object_id}"
        list_view.bulk_select_option = False
        list_view.list_column_visibility = False
        list_view.actions = self.get_actions()
        list_view.table_height = False
        list_view.table_height_as_class = "h-[calc(_100vh_-_500px_)]"
        list_view.table_width = False
        context = list_view.get_context_data(object_list=queryset)
        context.update(super().get_context_data())
        context["can_add_attachment"] = self.check_attachment_add_permission()
        return render(request, self.template_name, context)


@method_decorator(htmx_required, name="dispatch")
@method_decorator(
    permission_required_or_denied(
        [
            "horilla_core.view_horillaattachment",
            "horilla_core.view_own_horillaattachment",
        ]
    ),
    name="dispatch",
)
class HorillaNotesAttachementDetailView(DetailView):

    template_name = "notes_attachments_detail.html"
    context_object_name = "obj"
    model = HorillaAttachment

    def get(self, request, *args, **kwargs):
        try:
            self.object = self.get_object()
        except Http404:
            messages.error(self.request, "The requested attachment does not exist.")
            return HttpResponse(
                "<script>$('#reloadButton').click();$('#reloadMessagesButton').click();closeContentModal();</script>"
            )

        context = self.get_context_data()
        return self.render_to_response(context)


@method_decorator(htmx_required, name="dispatch")
class HorillaNotesAttachmentCreateView(LoginRequiredMixin, FormView):
    template_name = "forms/notes_attachment_form.html"
    form_class = HorillaAttachmentForm
    model = HorillaAttachment

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form_url"] = reverse_lazy("horilla_generics:notes_attachment_create")
        pk = self.kwargs.get("pk")
        if pk:
            context["form_url"] = reverse_lazy(
                "horilla_generics:notes_attachment_edit", kwargs={"pk": pk}
            )
        return context

    def get_object(self):
        """Return object if pk exists (for edit mode)."""
        pk = self.kwargs.get("pk")
        if pk:
            obj = get_object_or_404(HorillaAttachment, pk=pk)
            return obj
        return None

    def get_form(self, form_class=None):
        """Bind instance if editing."""
        form_class = self.get_form_class()
        obj = self.get_object()
        return form_class(instance=obj, **self.get_form_kwargs())

    def check_related_object_permission(self, related_object, permission_type="add"):
        """
        Check if user has permission to add/change notes on the related object.

        Args:
            related_object: The object to which the attachment is related
            permission_type: 'add' or 'change'

        Returns:
            bool: True if user has permission, False otherwise
        """
        user = self.request.user

        related_model = related_object.__class__
        model_name = related_model._meta.model_name
        app_label = related_model._meta.app_label

        is_owner = False
        owner_fields = getattr(related_model, "OWNER_FIELDS", [])

        for owner_field in owner_fields:
            try:
                field_value = getattr(related_object, owner_field, None)
                if field_value:
                    if hasattr(field_value, "all"):
                        if user in field_value.all():
                            is_owner = True
                            break
                    # Handle ForeignKey fields
                    elif field_value == user:
                        is_owner = True
                        break
            except Exception:
                continue

        if is_owner:
            change_own_perm = f"{app_label}.change_own_{model_name}"
            if user.has_perm(change_own_perm) and user.has_perm(
                "horilla_core.add_horillaattachment"
            ):
                return True

        change_perm = f"{app_label}.change_{model_name}"
        if user.has_perm(change_perm) and user.has_perm(
            "horilla_core.add_horillaattachment"
        ):
            return True

        return False

    def dispatch(self, request, *args, **kwargs):
        """Check permissions before processing the request."""
        # For edit mode, check if attachment exists and user has permission
        pk = kwargs.get("pk")
        if pk:
            try:
                attachment = self.model.objects.get(pk=pk)
                related_object = attachment.related_object

                if related_object:
                    if not self.check_related_object_permission(
                        related_object, "change"
                    ):
                        messages.error(
                            request,
                            _("You don't have permission to edit this attachment."),
                        )
                        return HttpResponse(
                            "<script>$('#reloadButton').click();$('#reloadMessagesButton').click();closeModal();</script>"
                        )
            except self.model.DoesNotExist:
                messages.error(request, _("The requested attachment does not exist."))
                return HttpResponse(
                    "<script>$('#reloadButton').click();$('#reloadMessagesButton').click();closeModal();</script>"
                )

        # For create mode, check permission on the related object
        else:
            model_name = request.GET.get("model_name")
            object_id = request.GET.get("object_id")

            if model_name and object_id:
                try:
                    content_type = ContentType.objects.get(model=model_name.lower())
                    related_model = content_type.model_class()
                    related_object = related_model.objects.get(pk=object_id)

                    if not self.check_related_object_permission(related_object, "add"):
                        messages.error(
                            request,
                            _(
                                "You don't have permission to add attachments to this record."
                            ),
                        )
                        return HttpResponse(
                            "<script>$('#reloadButton').click();$('#reloadMessagesButton').click();closeModal();</script>"
                        )
                except (
                    ContentType.DoesNotExist,
                    related_model.DoesNotExist,
                    ValueError,
                ):
                    messages.error(request, _("Invalid related object."))
                    return HttpResponse(
                        "<script>$('#reloadButton').click();$('#reloadMessagesButton').click();closeModal();</script>"
                    )

        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        model_name = self.request.GET.get("model_name")
        pk = self.kwargs.get("pk")

        attachment = form.save(commit=False)
        if not pk:
            content_type = ContentType.objects.get(model=model_name.lower())
            attachment.created_by = self.request.user
            attachment.object_id = self.request.GET.get("object_id")
            attachment.content_type = content_type
            attachment.company = self.request.active_company
            messages.success(self.request, f"{attachment.title} created successfully")
        else:
            messages.success(self.request, f"{attachment.title} updated successfully")
        attachment.save()
        return HttpResponse(
            "<script>$('#tab-notes-attachments').click();closeModal();</script>"
        )

    def get(self, request, *args, **kwargs):
        pk = kwargs.get("pk")
        if pk:
            try:
                self.model.objects.get(pk=pk)
            except self.model.DoesNotExist:
                messages.error(request, "The requested attachment does not exist.")
                return HttpResponse(
                    "<script>$('#reloadButton').click();$('#reloadMessagesButton').click();closeModal();</script>"
                )

        return super().get(request, *args, **kwargs)


@method_decorator(htmx_required, name="dispatch")
class HorillaHistorySectionView(DetailView):
    template_name = "history_tab.html"
    context_object_name = "obj"
    paginate_by = 2
    filter_form_class = HorillaHistoryForm

    def dispatch(self, request, *args, **kwargs):
        try:
            self.object = self.get_object()
        except Exception as e:
            messages.error(self.request, e)
            return HttpResponse(headers={"HX-Refresh": "true"})
            # return HttpResponse(
            #     "<script>$('#reloadButton').click();</script>"
            # )
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["model_name"] = self.model._meta.model_name
        histories = self.get_object().full_histories

        history_by_date = []
        date_dict = {}
        for entry in histories:
            date_key = entry.timestamp.date()
            if date_key not in date_dict:
                date_dict[date_key] = []
            date_dict[date_key].append(entry)

        sorted_dates = sorted(date_dict.keys(), reverse=True)
        history_by_date = [(date, date_dict[date]) for date in sorted_dates]
        filter_form = self.filter_form_class(self.request.GET)
        filter_applied = False
        if self.request.GET:
            filter_applied = any(
                self.request.GET.get(field) not in [None, "", "all"]
                for field in filter_form.fields
            )

            if filter_form.is_valid() and filter_applied:
                history_by_date = filter_form.apply_filter(history_by_date)

        paginator = Paginator(history_by_date, self.paginate_by)
        page_number = self.request.GET.get("page", 1)
        page_obj = paginator.get_page(page_number)

        context["page_obj"] = page_obj
        context["actions"] = [str(entry).split()[0].lower() for entry in histories]
        context["filter_form"] = filter_form
        context["filter_applied"] = filter_applied

        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        context = self.get_context_data(**kwargs)
        if request.GET.get("show_filter") == "true" and request.headers.get(
            "HX-Request"
        ):
            return HttpResponse(
                render_to_string(
                    "partials/history_filter_form.html",
                    {"form": context["filter_form"], "request": request},
                    request=request,
                )
            )

        return self.render_to_response(context)


class HorillaMultiStepFormView(FormView):
    template_name = "form_view.html"
    form_class = None
    model = None
    success_url = None
    step_titles = {}
    total_steps = 4
    form_url_name = None
    form_title = None
    fullwidth_fields = []
    dynamic_create_fields = []
    dynamic_create_field_mapping = {}
    pk_url_kwarg = "pk"
    permission_required = None
    check_object_permission = True
    permission_denied_template = "error/403.html"
    skip_permission_check = False

    single_step_url_name = None

    def get_single_step_url(self):
        """Get the URL for single-step form"""
        if not self.single_step_url_name:
            return None

        pk = self.kwargs.get(self.pk_url_kwarg)
        if pk:
            # For edit mode, use edit URL
            if isinstance(self.single_step_url_name, dict):
                url_name = self.single_step_url_name.get("edit")
                return (
                    reverse(url_name, kwargs={self.pk_url_kwarg: pk})
                    if url_name
                    else None
                )
            return reverse(self.single_step_url_name, kwargs={self.pk_url_kwarg: pk})
        else:
            # For create mode, use create URL
            if isinstance(self.single_step_url_name, dict):
                url_name = self.single_step_url_name.get("create")
                return reverse(url_name) if url_name else None
            return reverse(self.single_step_url_name)

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.storage_key = f"{self.__class__.__name__}_form_data"
        self.object = None

    def dispatch(self, request, *args, **kwargs):
        if not self.skip_permission_check and not self.has_permission():
            return render(request, self.permission_denied_template, {"modal": True})
        pk = kwargs.get(self.pk_url_kwarg)
        if pk:
            try:
                self.object = get_object_or_404(self.model, pk=pk)
            except Exception as e:
                messages.error(request, e)
                return HttpResponse(
                    "<script>$('#reloadButton').click();closeModal();</script>"
                )
            self.storage_key = f"{self.__class__.__name__}_form_data_{pk}"
        return super().dispatch(request, *args, **kwargs)

    def get_auto_permissions(self):
        """
        Automatically generate the appropriate permission based on create/edit mode.
        """
        if not self.model:
            return []

        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name

        # Check if this is edit mode or create mode
        is_edit_mode = bool(self.kwargs.get("pk"))

        if is_edit_mode:
            return [f"{app_label}.change_{model_name}"]
        else:
            return [f"{app_label}.add_{model_name}"]

    def has_permission(self):
        """
        Check if the user has the required permissions.
        Automatically checks both model permissions and object-level permissions.
        """
        user = self.request.user

        permissions = self.permission_required or self.get_auto_permissions()

        if isinstance(permissions, str):
            permissions = [permissions]

        has_model_permission = any(user.has_perm(perm) for perm in permissions)

        if has_model_permission:
            return True

        if self.kwargs.get("pk") and self.model:
            app_label = self.model._meta.app_label
            model_name = self.model._meta.model_name
            change_own_perm = f"{app_label}.change_own_{model_name}"
            permissions.append(change_own_perm)

            if user.has_perm(change_own_perm):
                return self.has_object_permission()

        return False

    def has_object_permission(self):
        """
        Check object-level permissions (e.g., ownership) on self.model.
        Uses model's OWNER_FIELDS attribute to determine ownership.
        """
        if not self.kwargs.get("pk") or not self.model:
            return False

        try:
            obj = self.model.objects.get(pk=self.kwargs["pk"])

            if hasattr(obj, "is_owned_by"):
                return obj.is_owned_by(self.request.user)

            if hasattr(self.model, "OWNER_FIELDS"):
                owner_fields = self.model.OWNER_FIELDS
                for owner_field in owner_fields:
                    if hasattr(obj, owner_field):
                        owner = getattr(obj, owner_field)
                        if owner == self.request.user:
                            return True

            fallback_owner_fields = [
                f"{self.model._meta.model_name}_owner",  # e.g., campaign_owner
                "owner",
                "created_by",
                "user",
            ]

            for owner_field in fallback_owner_fields:
                if hasattr(obj, owner_field):
                    owner = getattr(obj, owner_field)
                    if owner == self.request.user:
                        return True

            return False

        except self.model.DoesNotExist:
            return False

    def cleanup_session_data(self):
        """Clean up session data"""
        keys_to_remove = [self.storage_key, f"{self.storage_key}_files"]
        for key in keys_to_remove:
            if key in self.request.session:
                del self.request.session[key]
        self.request.session.modified = True

    def get_form_class(self):
        if self.form_class is None and self.model is not None:

            class DynamicMultiStepForm(HorillaMultiStepForm):
                class Meta:
                    model = self.model
                    fields = "__all__"
                    exclude = [
                        "created_at",
                        "updated_at",
                        "created_by",
                        "updated_by",
                        "additional_info",
                    ]
                    widgets = {
                        field.name: forms.DateInput(attrs={"type": "date"})
                        for field in self.model._meta.fields
                        if isinstance(field, models.DateField)
                    }

            return DynamicMultiStepForm
        return super().get_form_class()

    def get_initial_step(self):
        """Get the initial step, ensuring it's valid and within bounds."""
        try:
            step = int(self.request.POST.get("step", 1))
            if step < 1 or step > self.total_steps:
                return 1
            return step
        except (ValueError, TypeError):
            return 1

    def encode_file_for_session(self, uploaded_file):
        """Encode file to store in session"""
        try:
            content = uploaded_file.read()
            uploaded_file.seek(0)
            return {
                "name": uploaded_file.name,
                "content": base64.b64encode(content).decode("utf-8"),
                "content_type": uploaded_file.content_type,
                "size": uploaded_file.size,
            }
        except Exception as e:
            logger.error(f"Error encoding file: {e}")
            return None

    def decode_file_from_session(self, file_data):
        """Decode file from session storage"""
        try:
            if not file_data or "content" not in file_data:
                return None
            content = base64.b64decode(file_data["content"])
            return SimpleUploadedFile(
                name=file_data["name"],
                content=content,
                content_type=file_data["content_type"],
            )
        except Exception as e:
            logger.error(f"Error decoding file: {e}")
            return None

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["request"] = self.request
        step = getattr(self, "current_step", self.get_initial_step())
        kwargs["step"] = step
        kwargs["full_width_fields"] = self.fullwidth_fields
        kwargs["dynamic_create_fields"] = self.dynamic_create_fields

        if self.object:
            kwargs["instance"] = self.object

        form_data = self.request.session.get(self.storage_key, {})
        files_data = self.request.session.get(f"{self.storage_key}_files", {})
        form_class = self.get_form_class()
        step_fields = getattr(form_class, "step_fields", {}).get(step, [])

        if self.request.method == "POST" and "reset" not in self.request.GET:
            post_data = self.request.POST.copy()
            boolean_fields = [
                field.name
                for field in self.model._meta.fields
                if isinstance(field, models.BooleanField)
            ]
            many_to_many_fields = [
                field.name
                for field in self.model._meta.get_fields()
                if isinstance(field, models.ManyToManyField)
            ]
            file_fields = [
                field.name
                for field in self.model._meta.fields
                if isinstance(field, (models.FileField, models.ImageField))
            ]

            for key in post_data:
                if key not in ["csrfmiddlewaretoken", "step", "previous"]:
                    if key in many_to_many_fields:
                        values = post_data.getlist(key)
                        form_data[key] = values if values else []
                        continue

                    try:
                        model_field = self.model._meta.get_field(key)
                        if isinstance(model_field, models.DateField) and not isinstance(
                            model_field, models.DateTimeField
                        ):
                            parsed_date = parse_date(
                                post_data[key].split("T")[0]
                                if "T" in post_data[key]
                                else post_data[key]
                            )
                            if parsed_date:
                                form_data[key] = parsed_date.isoformat()
                                continue
                        parsed_datetime = parse_datetime(post_data[key])
                        if parsed_datetime:
                            form_data[key] = parsed_datetime.isoformat()
                            continue
                        parsed_date = parse_date(post_data[key])
                        if parsed_date:
                            form_data[key] = parsed_date.isoformat()
                            continue
                        try:
                            decimal_value = Decimal(post_data[key])
                            form_data[key] = str(decimal_value)
                            continue
                        except (ValueError, TypeError, InvalidOperation):
                            pass
                    except:
                        pass
                    form_data[key] = post_data[key]

            for field_name in boolean_fields:
                if (
                    field_name in step_fields
                    and field_name not in post_data
                    and step == int(post_data.get("step", 1))
                ):
                    form_data[field_name] = False
            for field_name in file_fields:
                if field_name in self.request.FILES:
                    uploaded_file = self.request.FILES[field_name]
                    encoded_file = self.encode_file_for_session(uploaded_file)
                    if encoded_file:
                        files_data[field_name] = encoded_file
                        form_data[f"{field_name}_filename"] = uploaded_file.name
                        form_data[f"{field_name}_new_file"] = True
                        # Remove cleared flag if new file uploaded
                        form_data.pop(f"{field_name}_cleared", None)
                # Check if file was cleared
                elif (
                    f"{field_name}-clear" in post_data
                    and post_data[f"{field_name}-clear"] == "true"
                ):
                    # Mark file as cleared
                    form_data[f"{field_name}_cleared"] = True
                    # Remove file from session
                    files_data.pop(field_name, None)
                    form_data.pop(f"{field_name}_filename", None)
                    form_data.pop(f"{field_name}_new_file", None)

            self.request.session[self.storage_key] = form_data
            self.request.session[f"{self.storage_key}_files"] = files_data
            self.request.session.modified = True

        if form_data:
            if (
                self.request.method == "GET"
                and step == 1
                and "previous" not in self.request.POST
                and "new" in self.request.GET
                and not self.object
            ):
                pass
            else:
                kwargs["form_data"] = form_data
                kwargs["data"] = form_data

        files_dict = {}

        if self.request.FILES:
            files_dict.update(self.request.FILES)

        for field_name, file_data in files_data.items():
            if field_name not in files_dict:
                decoded_file = self.decode_file_from_session(file_data)
                if decoded_file:
                    files_dict[field_name] = decoded_file

        if files_dict:
            kwargs["files"] = files_dict

        # Updated here

        if self.request.method == "POST" and (
            "previous" in self.request.POST
            or (
                self.get_initial_step() < self.total_steps
                and "step" in self.request.POST
            )
        ):
            if "data" in kwargs:
                kwargs["data"] = None

        return kwargs

    def get_form_title(self):
        if self.model:
            action = _("Update") if self.object else _("Create")
            verbose = self.model._meta.verbose_name
            return f"{action} {verbose}"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        self.current_step = getattr(self, "current_step", self.get_initial_step())
        context["step_titles"] = self.step_titles
        context["total_steps"] = self.total_steps
        context["current_step"] = self.current_step
        context["form_title"] = self.form_title or self.get_form_title()
        context["object"] = self.object
        context["is_edit"] = bool(self.object)
        context["full_width_fields"] = self.fullwidth_fields
        context["dynamic_create_fields"] = self.dynamic_create_fields
        context["dynamic_create_field_mapping"] = self.dynamic_create_field_mapping

        if self.form_url_name:
            if self.object:
                context["form_url"] = reverse(
                    self.form_url_name, kwargs={self.pk_url_kwarg: self.object.pk}
                )
            else:
                context["form_url"] = reverse(self.form_url_name)
        else:
            context["form_url"] = self.request.path

        related_models_info = {}
        if self.dynamic_create_fields:
            for field_name in self.dynamic_create_fields:
                try:
                    field = self.model._meta.get_field(field_name)
                    if isinstance(field, (models.ForeignKey, models.ManyToManyField)):
                        related_model = field.related_model
                        related_models_info[field_name] = {
                            "model_name": related_model._meta.model_name,
                            "app_label": related_model._meta.app_label,
                            "verbose_name": related_model._meta.verbose_name.title(),
                        }
                except:
                    pass
        context["related_models_info"] = related_models_info

        context["stored_form_data"] = self.request.session.get(self.storage_key, {})
        context["stored_files_data"] = self.request.session.get(
            f"{self.storage_key}_files", {}
        )

        form_data = self.request.session.get(self.storage_key, {})
        files_data = self.request.session.get(f"{self.storage_key}_files", {})
        file_field_states = {}

        for field in self.model._meta.fields:
            if isinstance(field, (models.FileField, models.ImageField)):
                field_name = field.name
                # Check if file was cleared
                if form_data.get(f"{field_name}_cleared"):
                    file_field_states[field_name] = {
                        "has_file": False,
                        "filename": None,
                        "is_cleared": True,
                    }
                elif field_name in files_data or form_data.get(
                    f"{field_name}_new_file"
                ):
                    filename = form_data.get(f"{field_name}_filename")
                    file_field_states[field_name] = {
                        "has_file": True,
                        "filename": filename,
                        "is_new": True,
                    }
                # Use instance file if exists and not modified
                elif self.object and hasattr(self.object, field_name):
                    instance_file = getattr(self.object, field_name, None)
                    if instance_file and instance_file.name:
                        file_field_states[field_name] = {
                            "has_file": True,
                            "filename": instance_file.name.split("/")[-1],
                            "is_existing": True,
                        }

        context["file_field_states"] = file_field_states

        form = context.get("form")
        if form and hasattr(self, "fullwidth_fields"):
            for field_name, field in form.fields.items():
                if field_name in self.fullwidth_fields:
                    field.widget.attrs["fullwidth"] = True

        context["single_step_url"] = self.get_single_step_url()
        return context

    def form_valid(self, form):
        step = self.get_initial_step()

        if step < self.total_steps:
            self.current_step = step + 1
            form_kwargs = self.get_form_kwargs()

            files_data = self.request.session.get(f"{self.storage_key}_files", {})
            final_files = {}

            if self.request.FILES:
                final_files.update(self.request.FILES)

            for field_name, file_data in files_data.items():
                if field_name not in final_files:
                    decoded_file = self.decode_file_from_session(file_data)
                    if decoded_file:
                        final_files[field_name] = decoded_file

            next_step_form_kwargs = {
                "step": self.current_step,
                "form_data": form_kwargs.get("form_data", {}),
                "instance": self.object if self.object else None,
                "full_width_fields": self.fullwidth_fields,
                "dynamic_create_fields": self.dynamic_create_fields,
                "request": self.request,
            }

            if final_files:
                next_step_form_kwargs["files"] = final_files

            next_step_form = self.get_form_class()(**next_step_form_kwargs)

            try:
                next_step_form = self.get_form_class()(**next_step_form_kwargs)
                next_step_form.errors.clear()
                next_step_form.is_bound = False
            except Exception as e:
                logger.error(f"Error creating next step form: {e}")
                next_step_form = self.get_form_class()(**next_step_form_kwargs)

            return self.render_to_response(self.get_context_data(form=next_step_form))

        try:
            form_data = self.request.session.get(self.storage_key, {})
            files_data = self.request.session.get(f"{self.storage_key}_files", {})

            for key, value in self.request.POST.items():
                if key not in ["csrfmiddlewaretoken", "step", "previous"]:
                    if key in [
                        field.name
                        for field in self.model._meta.get_fields()
                        if isinstance(field, models.ManyToManyField)
                    ]:
                        form_data[key] = self.request.POST.getlist(key)
                    else:
                        form_data[key] = value

            final_files = {}

            if self.request.FILES:
                final_files.update(self.request.FILES)

            for field_name, file_data in files_data.items():
                if field_name not in final_files:
                    decoded_file = self.decode_file_from_session(file_data)
                    if decoded_file:
                        final_files[field_name] = decoded_file

            final_form_kwargs = {
                "data": form_data,
                "full_width_fields": self.fullwidth_fields,
                "dynamic_create_fields": self.dynamic_create_fields,
                "request": self.request,
            }

            if final_files:
                final_form_kwargs["files"] = final_files

            if self.object:
                final_form_kwargs["instance"] = self.object

            final_form = self.get_form_class()(**final_form_kwargs)

            if final_form.is_valid():
                try:

                    instance = final_form.save(commit=False)
                    instance.company = (
                        getattr(_thread_local, "request", None).active_company
                        if hasattr(_thread_local, "request")
                        else self.request.user.company
                    )
                    for field in self.model._meta.get_fields():
                        if isinstance(field, (models.FileField, models.ImageField)):
                            if form_data.get(f"{field.name}_cleared"):
                                setattr(instance, field.name, None)
                    instance.save()

                    for field in self.model._meta.get_fields():
                        if (
                            isinstance(field, models.ManyToManyField)
                            and field.name in form_data
                        ):
                            values = form_data[field.name]
                            if values:
                                getattr(instance, field.name).set(values)
                            else:
                                getattr(instance, field.name).clear()

                    self.cleanup_session_data()

                    action = "updated" if self.object else "created"
                    messages.success(
                        self.request,
                        f"{self.model.__name__} was successfully {action}.",
                    )
                    return HttpResponse(
                        "<script>$('#reloadButton').click();closeModal();</script>"
                    )
                except Exception as e:
                    final_form.add_error(None, e)

                    error_form_kwargs = {
                        "data": form_data,
                        "step": self.total_steps,
                        "form_data": form_data,
                        "full_width_fields": self.fullwidth_fields,
                        "dynamic_create_fields": self.dynamic_create_fields,
                        "request": self.request,
                    }

                    if final_files:
                        error_form_kwargs["files"] = final_files

                    if self.object:
                        error_form_kwargs["instance"] = self.object

                    error_form = self.get_form_class()(**error_form_kwargs)

                    # copy over the error into the form
                    for field_name, errors in final_form.errors.items():
                        if field_name == "__all__":
                            for error in errors:
                                error_form.add_error(None, error)
                        else:
                            for error in errors:
                                error_form.add_error(field_name, error)

                    self.current_step = self.total_steps
                    return self.render_to_response(
                        self.get_context_data(form=error_form)
                    )
            else:
                return self.render_to_response(self.get_context_data(form=final_form))

        except Exception as e:
            action = "updating" if self.object else "creating"
            messages.error(
                self.request, f"Error {action} {self.model.__name__}: {str(e)}"
            )
            logger.error(f"Exception in form_valid: {str(e)}")
            import traceback

            traceback.print_exc()
            return self.render_to_response(self.get_context_data(form=form))

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))

    def post(self, request, *args, **kwargs):
        if "previous" in request.POST:
            step = self.get_initial_step()
            if step > 1:
                self.current_step = step - 1

                files_data = self.request.session.get(f"{self.storage_key}_files", {})
                form_data = self.request.session.get(self.storage_key, {})

                final_files = {}
                for field_name, file_data in files_data.items():
                    decoded_file = self.decode_file_from_session(file_data)
                    if decoded_file:
                        final_files[field_name] = decoded_file

                form_kwargs = {
                    "step": self.current_step,
                    "form_data": form_data,
                    "instance": self.object if self.object else None,
                    "full_width_fields": self.fullwidth_fields,
                    "dynamic_create_fields": self.dynamic_create_fields,
                    "request": self.request,
                    "data": form_data,
                }

                if final_files:
                    form_kwargs["files"] = final_files

                form = self.get_form_class()(**form_kwargs)

                form.errors.clear()

                return self.render_to_response(self.get_context_data(form=form))

        return super().post(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        if "reset" in request.GET or ("new" in request.GET and not self.object):
            self.cleanup_session_data()
        elif self.object:
            step = int(request.GET.get("step", 1))
            if step == 1 and "previous" not in request.POST:
                self.cleanup_session_data()
        self.current_step = 1
        form = self.get_form()
        return self.render_to_response(self.get_context_data(form=form))


class HorillaSingleFormView(FormView):

    template_name = "single_form_view.html"
    model = None
    form_class = None
    success_url = None
    object = None
    fields = None
    exclude = None
    full_width_fields = None
    form_url = None
    dynamic_create_fields = None
    modal_height = True
    form_title = None
    hidden_fields = []
    view_id = ""
    condition_fields = None
    condition_model = None
    condition_field_choices = None
    condition_field_title = None
    header = True
    modal_height_class = None
    hx_attrs: dict = {}
    permission_required = None
    check_object_permission = True
    permission_denied_template = "error/403.html"
    skip_permission_check = False

    multi_step_url_name = None
    duplicate_mode = False

    def get_multi_step_url(self):
        """Get the URL for multi-step form"""
        if not self.multi_step_url_name:
            return None

        pk = self.kwargs.get("pk")
        if pk:
            # For edit mode, use edit URL
            if isinstance(self.multi_step_url_name, dict):
                url_name = self.multi_step_url_name.get("edit")
                return reverse(url_name, kwargs={"pk": pk}) if url_name else None
            return reverse(self.multi_step_url_name, kwargs={"pk": pk})
        else:
            # For create mode, use create URL
            if isinstance(self.multi_step_url_name, dict):
                url_name = self.multi_step_url_name.get("create")
                return reverse(url_name) if url_name else None
            return reverse(self.multi_step_url_name)

    def dispatch(self, request, *args, **kwargs):
        if "pk" in self.kwargs:
            self.duplicate_mode = (
                request.GET.get("duplicate", "false").lower() == "true"
            )

        if not self.skip_permission_check and not self.has_permission():
            return render(request, self.permission_denied_template, {"modal": True})

        if request.headers.get("HX-Request") and "add_condition_row" in request.GET:
            return self.add_condition_row(request)

        if "pk" in self.kwargs:
            try:
                self.object = get_object_or_404(self.model, pk=self.kwargs["pk"])
            except Exception as e:
                messages.error(request, e)
                return HttpResponse(
                    "<script>$('#reloadButton').click();closeModal();</script>"
                )
        return super().dispatch(request, *args, **kwargs)

    def get_auto_permissions(self):
        """
        Automatically generate the appropriate permission based on create/edit mode.
        """
        if not self.model:
            return []

        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name

        if self.duplicate_mode:
            return [f"{app_label}.add_{model_name}"]

        # Check if this is edit mode or create mode
        is_edit_mode = bool(self.kwargs.get("pk"))

        if is_edit_mode:
            return [f"{app_label}.change_{model_name}"]
        else:
            return [f"{app_label}.add_{model_name}"]

    def has_permission(self):
        """
        Check if the user has the required permissions.
        Automatically checks both model permissions and object-level permissions.
        """
        user = self.request.user

        permissions = self.permission_required or self.get_auto_permissions()

        if isinstance(permissions, str):
            permissions = [permissions]

        has_model_permission = any(user.has_perm(perm) for perm in permissions)

        if has_model_permission:
            return True

        if self.kwargs.get("pk") and self.model:
            app_label = self.model._meta.app_label
            model_name = self.model._meta.model_name
            change_own_perm = f"{app_label}.change_own_{model_name}"
            permissions.append(change_own_perm)

            if user.has_perm(change_own_perm):
                return self.has_object_permission()
        return False

    def has_object_permission(self):
        """
        Check object-level permissions (e.g., ownership) on self.model.
        Uses model's OWNER_FIELDS attribute to determine ownership.
        """
        if not self.kwargs.get("pk") or not self.model:
            return False

        try:
            obj = self.model.objects.get(pk=self.kwargs["pk"])

            if hasattr(obj, "is_owned_by"):
                return obj.is_owned_by(self.request.user)

            if hasattr(self.model, "OWNER_FIELDS"):
                owner_fields = self.model.OWNER_FIELDS
                for owner_field in owner_fields:
                    if hasattr(obj, owner_field):
                        owner = getattr(obj, owner_field)
                        if owner == self.request.user:
                            return True

            fallback_owner_fields = [
                f"{self.model._meta.model_name}_owner",  # e.g., campaign_owner
                "owner",
                "created_by",
                "user",
            ]

            for owner_field in fallback_owner_fields:
                if hasattr(obj, owner_field):
                    owner = getattr(obj, owner_field)
                    if owner == self.request.user:
                        return True

            return False

        except self.model.DoesNotExist:
            return False

    def get(self, request, *args, **kwargs):
        if self.kwargs.get("pk"):
            for key in self.session_keys_to_clear_on_edit:
                if key in request.session:
                    del request.session[key]
            request.session.modified = True

            existing_conditions = self.get_existing_conditions()
            if existing_conditions is not None:
                request.session["condition_row_count"] = len(existing_conditions)
                request.session.modified = True
        return super().get(request, *args, **kwargs)

    def get_existing_conditions(self):
        """Retrieve existing conditions for the current object in edit mode."""
        if self.kwargs.get("pk") and hasattr(self, "object") and self.object:
            existing_conditions = getattr(self.object, "team_members", None) or getattr(
                self.object, "conditions", None
            )
            if existing_conditions and hasattr(existing_conditions, "all"):
                return existing_conditions.all().order_by("created_at")
        return None

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.session_keys_to_clear_on_edit = ["condition_row_count"]

    def get_submitted_condition_data(self):
        """Extract condition field data from submitted form data"""
        condition_data = {}
        if self.condition_fields and self.request.method == "POST":
            for key, value in self.request.POST.items():
                # Check if this is a condition field with row_id
                for field_name in self.condition_fields:
                    if key.startswith(f"{field_name}_") and key != field_name:
                        # Extract row_id from the key
                        try:
                            row_id = key.replace(f"{field_name}_", "")
                            if row_id not in condition_data:
                                condition_data[row_id] = {}
                            condition_data[row_id][field_name] = value
                        except:
                            continue
        return condition_data

    def add_condition_row(self, request):
        row_id = request.GET.get("row_id", "0")

        new_row_id = "0"
        if row_id == "next":
            current_count = request.session.get("condition_row_count", 0)
            current_count += 1
            request.session["condition_row_count"] = current_count
            new_row_id = str(current_count)
        else:
            try:
                new_row_id = str(int(row_id) + 1)
            except ValueError:
                new_row_id = "1"

        form_kwargs = self.get_form_kwargs()
        form_kwargs["row_id"] = new_row_id

        if "pk" in self.kwargs:
            try:
                instance = self.model.objects.get(pk=self.kwargs["pk"])
                form_kwargs["instance"] = instance
            except self.model.DoesNotExist:
                pass

        form = self.get_form_class()(**form_kwargs)

        context = {
            "form": form,
            "condition_fields": self.condition_fields or [],
            "row_id": new_row_id,
            "submitted_condition_data": self.get_submitted_condition_data(),  # Add this
        }
        html = render_to_string("partials/condition_row.html", context, request=request)
        return HttpResponse(html)

    def get_form_class(self):
        if self.form_class is None and self.model is not None:
            full_width_fields = self.full_width_fields or []
            dynamic_create_fields = self.dynamic_create_fields or []
            hidden_fields = getattr(self, "hidden_fields", [])
            condition_fields = self.condition_fields or []
            condition_model = self.condition_model
            condition_field_choices = self.condition_field_choices or {}

            class DynamicForm(OwnerQuerysetMixin, HorillaModelForm):
                class Meta:
                    model = self.model
                    fields = self.fields if self.fields is not None else "__all__"
                    exclude = (
                        (
                            self.exclude
                            + [
                                "created_at",
                                "updated_at",
                                "created_by",
                                "updated_by",
                                "additional_info",
                            ]
                        )
                        if self.exclude is not None
                        else [
                            "created_at",
                            "updated_at",
                            "created_by",
                            "updated_by",
                            "additional_info",
                        ]
                    )
                    widgets = {
                        field.name: forms.DateInput(attrs={"type": "date"})
                        for field in self.model._meta.fields
                        if isinstance(field, models.DateField)
                    }

                def __init__(self, *args, **kwargs):
                    kwargs["dynamic_create_fields"] = dynamic_create_fields
                    kwargs["full_width_fields"] = full_width_fields
                    kwargs["hidden_fields"] = hidden_fields
                    kwargs["condition_fields"] = condition_fields
                    kwargs["condition_model"] = condition_model
                    kwargs["condition_field_choices"] = condition_field_choices
                    super().__init__(*args, **kwargs)

            return DynamicForm
        return super().get_form_class()

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["full_width_fields"] = self.full_width_fields or []
        kwargs["dynamic_create_fields"] = self.dynamic_create_fields or []
        kwargs["condition_fields"] = self.condition_fields or []
        kwargs["condition_model"] = self.condition_model
        kwargs["condition_field_choices"] = self.condition_field_choices or {}
        kwargs["hidden_fields"] = getattr(self, "hidden_fields", [])
        if self.object and not self.duplicate_mode:
            kwargs["instance"] = self.object
        elif self.object and self.duplicate_mode:
            # In duplicate mode, populate initial data from the object
            initial = kwargs.get("initial", {})
            for field in self.object._meta.fields:
                if field.name not in [
                    "id",
                    "pk",
                    "created_at",
                    "updated_at",
                    "created_by",
                    "updated_by",
                ]:
                    field_value = getattr(self.object, field.name)
                    if field_value is not None:
                        if field.get_internal_type() in ["CharField", "TextField"]:
                            initial[field.name] = f"{field_value} (Copy)"
                        else:
                            initial[field.name] = field_value

            # Handle ManyToMany fields
            for field in self.object._meta.many_to_many:
                m2m_value = getattr(self.object, field.name).all()
                if m2m_value.exists():
                    initial[field.name] = list(m2m_value)

            kwargs["initial"] = initial
        kwargs["request"] = self.request
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form_title"] = (
            self.form_title
            or f"{'Duplicate' if self.duplicate_mode else 'Update' if self.kwargs.get('pk') and not self.duplicate_mode else 'Create'} {self.model._meta.verbose_name}"
        )
        context["duplicate_mode"] = self.duplicate_mode
        context["full_width_fields"] = self.full_width_fields or []
        context["condition_fields"] = self.condition_fields or []
        context["condition_fields_tiltle"] = self.condition_field_title
        context["form_url"] = self.get_form_url()
        context["add_condition_url"] = (
            self.get_add_condition_url() if self.condition_fields else None
        )
        context["dynamic_create_fields"] = self.dynamic_create_fields or []
        context["dynamic_create_field_mapping"] = getattr(
            self, "dynamic_create_field_mapping", {}
        )
        context["modal_height"] = self.modal_height
        self.view_id = self.view_id or f"{self.model._meta.model_name}-form-view"
        context["view_id"] = self.view_id
        context["form_class_name"] = self.get_form_class().__name__
        context["model_name"] = (
            self.model._meta.model_name if self.model != None else ""
        )
        context["app_label"] = self.model._meta.app_label if self.model != None else ""
        context["submitted_condition_data"] = self.get_submitted_condition_data()

        if self.request.method == "POST" and context["submitted_condition_data"]:
            max_row_id = max(
                [
                    int(row_id)
                    for row_id in context["submitted_condition_data"].keys()
                    if row_id.isdigit()
                ]
                + [0]
            )
            context["condition_row_count"] = max_row_id + 1
        else:
            context["condition_row_count"] = self.request.session.get(
                "condition_row_count", 0
            )

        related_models_info = {}
        if self.dynamic_create_fields:
            for field_name in self.dynamic_create_fields:
                try:
                    field = self.model._meta.get_field(field_name)
                    if isinstance(field, (models.ForeignKey, models.ManyToManyField)):
                        related_model = field.related_model
                        related_models_info[field_name] = {
                            "model_name": related_model._meta.model_name,
                            "app_label": related_model._meta.app_label,
                            "verbose_name": related_model._meta.verbose_name.title(),
                        }
                except:
                    pass

        query_string = ""
        if self.request.GET:
            query_string = f"?{self.request.GET.urlencode()}"

        default_hx_attrs = {
            "hx-post": f"{self.form_url}{query_string}",
            "hx-swap": "outerHTML",
            "hx-target": f"#{self.view_id}-container",
            "enctype": "multipart/form-data",
        }
        context["related_models_info"] = related_models_info
        context["header"] = self.header
        context["modal_height_class"] = self.modal_height_class
        context["hx_attrs"] = {**default_hx_attrs, **(self.hx_attrs or {})}
        context["multi_step_url"] = self.get_multi_step_url()
        return context

    def get_form_url(self):
        return self.form_url or self.request.path

    def get_add_condition_url(self):
        if not self.condition_fields:
            return None
        params = self.request.GET.copy()
        params["add_condition_row"] = "1"
        form_url = self.get_form_url()
        return f"{form_url}{'&' if '?' in str(form_url) else '?'}{params.urlencode()}"

    def form_valid(self, form):
        if not self.request.user.is_authenticated:
            messages.error(
                self.request, "You must be logged in to perform this action."
            )
            return self.form_invalid(form)

        self.object = form.save(commit=False)

        for field_name, field in form.fields.items():
            if isinstance(field, forms.FileField) or isinstance(
                field, forms.ImageField
            ):
                clear_flag = self.request.POST.get(f"id_{field_name}_clear", "false")
                if clear_flag == "true":
                    setattr(self.object, field_name, None)

        if self.kwargs.get("pk") and not self.duplicate_mode:
            self.object.updated_at = timezone.now()
            self.object.updated_by = self.request.user
        else:
            self.object.created_at = timezone.now()
            self.object.created_by = self.request.user
            self.object.updated_at = timezone.now()
            self.object.updated_by = self.request.user
        self.object.company = (
            getattr(_thread_local, "request", None).active_company
            if hasattr(_thread_local, "request")
            else self.request.user.company
        )
        self.object.save()
        form.save_m2m()
        self.request.session["condition_row_count"] = 0
        self.request.session.modified = True
        messages.success(
            self.request,
            f"{self.model._meta.verbose_name.title()} {'duplicated' if self.duplicate_mode else 'updated' if self.kwargs.get('pk') and not self.duplicate_mode else 'created'} successfully!",
        )
        return HttpResponse("<script>$('#reloadButton').click();closeModal();</script>")

    def form_invalid(self, form):
        print(form.errors)
        return super().form_invalid(form)

    def get_success_url(self):
        return self.success_url or reverse_lazy(f"{self.model._meta.model_name}-list")


@method_decorator(htmx_required, name="dispatch")
class HorillaDynamicCreateView(LoginRequiredMixin, FormView):
    """
    View to handle dynamic creation of related models
    """

    template_name = "dynamic_form_view.html"
    target_model = None
    field_names = None

    def get_model_and_fields(self):
        app_label = self.kwargs.get("app_label")
        model_name = self.kwargs.get("model_name")
        fields_param = self.request.GET.get("fields", "")

        field_names = None
        if fields_param and fields_param.lower() not in ["none", ""]:
            field_names = [f.strip() for f in fields_param.split(",") if f.strip()]

        try:
            model = apps.get_model(app_label, model_name)
            return model, field_names
        except LookupError:
            logger.warning(f"Model {app_label}.{model_name} not found")
            messages.error(self.request, f"Model {app_label}.{model_name} not found")
            return None, None

    def dispatch(self, request, *args, **kwargs):
        # Initialize model + fields once
        self.target_model, self.field_names = self.get_model_and_fields()

        if not self.target_model:
            messages.error(self.request, "Invalid model or fields")
            return HttpResponse(
                "<script>$('#reloadMessagesButton').click();closeDynamicModal();</script>"
            )

        return super().dispatch(request, *args, **kwargs)

    def get_form_class(self):
        target_model, field_names = self.target_model, self.field_names

        class DynamicCreateForm(HorillaModelForm):
            class Meta:
                model = target_model
                fields = (
                    field_names if field_names and field_names != [""] else "__all__"
                )
                exclude = [
                    "created_at",
                    "updated_at",
                    "created_by",
                    "updated_by",
                    "additional_info",
                ]
                widgets = {
                    field.name: forms.DateInput(attrs={"type": "date"})
                    for field in target_model._meta.fields
                    if isinstance(field, models.DateField)
                }

        return DynamicCreateForm

    def get_full_width_fields(self):
        """Get full width fields from URL parameter"""
        full_width_param = self.request.GET.get("full_width_fields", "")
        full_width_fields = []
        if full_width_param and full_width_param.lower() not in ["none", ""]:
            full_width_fields = [
                f.strip() for f in full_width_param.split(",") if f.strip()
            ]
        return full_width_fields

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.target_model:
            context["form_title"] = (
                f"Create {self.target_model._meta.verbose_name.title()}"
            )
            context["target_field"] = self.request.GET.get("target_field")
            context["form_url"] = self.request.path
            context["full_width_fields"] = self.get_full_width_fields()
        return context

    def form_valid(self, form):
        if not self.request.user.is_authenticated:
            messages.error(
                self.request, "You must be logged in to perform this action."
            )
            return self.form_invalid(form)

        instance = form.save(commit=False)
        instance.created_at = timezone.now()
        instance.updated_at = timezone.now()
        instance.created_by = self.request.user
        instance.updated_by = self.request.user
        instance.save()
        form.save_m2m()

        target_field = self.request.GET.get("target_field")

        return HttpResponse(
            f"""
                <script>
                    var targetSelect = document.querySelector('select[name="{target_field}"]');
                    if (targetSelect) {{
                        var newOption = new Option('{instance}', '{instance.pk}', true, true);
                        targetSelect.add(newOption);

                        // Trigger change event if using Select2
                        if (window.$ && $(targetSelect).hasClass('js-example-basic-single')) {{
                            $(targetSelect).trigger('change');
                        }}
                    }}

                    closeDynamicModal();
                </script>
            """
        )

    def form_invalid(self, form):
        messages.error(self.request, "Please correct the errors below.")
        return super().form_invalid(form)


class HorillaSingleDeleteView(DeleteView):

    template_name = None
    success_url = None
    success_message = "The record was deleted successfully."
    reassign_all_visibility = True
    reassign_individual_visibility = True
    hx_target = None
    excluded_dependency_model_labels = [
        "RecycleBinPolicy",
        "RecycleBin",
        "ActiveTab",
        "KanbanGroupBy",
        "ListColumnVisibility",
        "RecentlyViewed",
        "SavedFilterList",
        "PinnedView",
        "LogEntry",
        "LoginHistory",
    ]

    def get_queryset(self):
        """
        Dynamically get queryset based on app_label and model_name from URL.
        """
        if self.model:
            # If model is explicitly set (in subclass), use it
            return (
                self.model.all_objects.all()
                if hasattr(self.model, "all_objects")
                else self.model.objects.all()
            )

        # Otherwise, get from URL parameters
        app_label = self.kwargs.get("app_label")
        model_name = self.kwargs.get("model_name")

        if not app_label or not model_name:
            raise ImproperlyConfigured(
                "HorillaSingleDeleteView requires either a 'model' attribute "
                "or 'app_label' and 'model_name' in URL kwargs."
            )

        try:
            self.model = apps.get_model(app_label, model_name)
            return (
                self.model.all_objects.all()
                if hasattr(self.model, "all_objects")
                else self.model.objects.all()
            )
        except LookupError:
            return render(self.request, "error/403.html", {"modal": True})

    def _get_excluded_models(self):
        """
        Resolve model names to model classes by searching all Django apps.
        Returns a list of model classes that match the names in excluded_dependency_model_labels.
        """
        excluded = []
        for model_name in self.excluded_dependency_model_labels:
            found_models = []
            for app_config in apps.get_app_configs():
                try:
                    model = app_config.get_model(model_name)
                    found_models.append(model)
                except LookupError:
                    continue

            if len(found_models) > 1:
                excluded.append(found_models[0])
            elif len(found_models) == 1:
                excluded.append(found_models[0])
            else:
                logger.warning(
                    f"Model '{model_name}' could not be resolved in any app."
                )

        return excluded

    def _get_paginated_individual_records(self, record_id, page=1, per_page=8):
        """
        Get paginated individual records for infinite scrolling in individual reassign form.
        """
        try:
            obj = self.model.objects.get(id=record_id)
            related_objects = self.model._meta.related_objects

            all_records = []
            available_targets = self.model.all_objects.exclude(id=record_id)
            is_nullable = False

            for related in related_objects:
                related_model = related.related_model
                related_name = related.get_accessor_name()
                if related_name:
                    if hasattr(related_model, "all_objects"):
                        related_records = related_model.all_objects.filter(
                            **{related.field.name: obj}
                        )
                    else:
                        related_records = getattr(obj, related_name).all()
                    all_records.extend(related_records)
                    is_nullable = self._is_field_nullable(related_model)

            total_count = len(all_records)
            offset = (page - 1) * per_page
            paginated_records = all_records[offset : offset + per_page]
            has_more = offset + per_page < total_count

            return {
                "records": paginated_records,
                "has_more": has_more,
                "next_page": page + 1 if has_more else None,
                "total_count": total_count,
                "available_targets": available_targets,
                "is_nullable": is_nullable,
            }

        except Exception as e:
            logger.error(f"Error getting paginated individual records: {str(e)}")
            return {
                "records": [],
                "has_more": False,
                "next_page": None,
                "total_count": 0,
                "available_targets": self.model.all_objects.exclude(id=record_id),
                "is_nullable": False,
            }

    def _check_dependencies(self, record_id, get_all=False):
        """
        Check for dependencies in related models for the given record ID, excluding specified models.
        Returns: cannot_delete (list), can_delete (list), dependency_details (dict).
        """
        cannot_delete = []
        can_delete = []
        dependency_details = {}

        try:
            obj = self.model.all_objects.filter(id=record_id).only("id").first()
            if not obj:
                logger.warning(
                    f"No record found with id {record_id} for model {self.model.__name__}"
                )
                return cannot_delete, can_delete, dependency_details

            related_objects = self.model._meta.related_objects
            if not related_objects:
                can_delete.append({"id": obj.id, "name": str(obj)})
                return cannot_delete, can_delete, dependency_details

            dependencies = []
            total_individual_records = 0
            excluded_models = self._get_excluded_models()

            for related in related_objects:
                related_model = related.related_model
                if related_model in excluded_models:
                    continue

                related_name = related.get_accessor_name()
                if related_name:
                    if hasattr(related_model, "all_objects"):
                        fk_field_name = related.field.name
                        all_related_records = related_model.all_objects.filter(
                            **{fk_field_name: obj}
                        )
                        total_count = all_related_records.count()

                        if get_all:
                            related_records = list(all_related_records)
                        else:
                            related_records = list(all_related_records[:10])
                    else:
                        related_records_qs = getattr(obj, related_name).all()
                        total_count = related_records_qs.count()
                        if get_all:
                            related_records = list(related_records_qs)
                        else:
                            related_records = list(related_records_qs[:10])

                    total_individual_records += total_count

                    if related_records or total_count > 0:
                        dependencies.append(
                            {
                                "model_name": related_model._meta.verbose_name_plural,
                                "count": total_count,
                                "records": [str(rec) for rec in related_records],
                                "related_model": related_model,
                                "related_name": related_name,
                                "related_records": related_records,
                                "has_more": (
                                    total_count > len(related_records)
                                    if not get_all
                                    else False
                                ),
                            }
                        )

            if dependencies:
                cannot_delete.append(
                    {
                        "id": obj.id,
                        "name": str(obj),
                        "dependencies": dependencies,
                        "total_individual_records": total_individual_records,  # Add this for individual form
                    }
                )
            else:
                can_delete.append({"id": obj.id, "name": str(obj)})

            dependency_details = {
                item["id"]: item["dependencies"] for item in cannot_delete
            }
            return cannot_delete, can_delete, dependency_details
        except Exception as e:
            logger.error(f"Error checking dependencies: {str(e)}")
            return cannot_delete, can_delete, dependency_details

    def _get_paginated_dependencies(self, record_id, related_name, page=1, per_page=8):
        """
        Get paginated dependencies for infinite scrolling.
        """
        try:
            obj = self.model.objects.get(id=record_id)
            related_objects = self.model._meta.related_objects
            excluded_models = self._get_excluded_models()

            for related in related_objects:
                related_model = related.related_model
                if related_model in excluded_models:
                    continue

                if related.get_accessor_name() == related_name:
                    related_model = related.related_model

                    if hasattr(related_model, "all_objects"):
                        fk_field_name = related.field.name
                        queryset = related_model.all_objects.filter(
                            **{fk_field_name: obj}
                        )
                    else:
                        queryset = getattr(obj, related_name).all()

                    total_count = queryset.count()
                    offset = (page - 1) * per_page
                    records = queryset[offset : offset + per_page]
                    has_more = offset + per_page < total_count

                    return {
                        "records": records,
                        "has_more": has_more,
                        "next_page": page + 1 if has_more else None,
                        "total_count": total_count,
                        "related_model": related_model,
                    }

            return {
                "records": [],
                "has_more": False,
                "next_page": None,
                "total_count": 0,
            }
        except Exception as e:
            logger.error(f"Error getting paginated dependencies: {str(e)}")
            return {
                "records": [],
                "has_more": False,
                "next_page": None,
                "total_count": 0,
            }

    def _is_field_nullable(self, related_model):
        """
        Check if the foreign key field to self.model in related_model is nullable.
        Returns: bool
        """
        try:
            field_name = [
                f.name
                for f in related_model._meta.fields
                if f.related_model == self.model
            ][0]
            field = related_model._meta.get_field(field_name)
            return field.null
        except IndexError:
            return False

    def _perform_bulk_reassign(self, record_id, new_target_id):
        """
        Reassign all dependent records to a new target.
        Returns the number of reassigned records.
        """
        try:
            obj = self.model.all_objects.get(id=record_id)
            new_target = self.model.all_objects.get(id=new_target_id)
            related_objects = self.model._meta.related_objects
            reassigned_count = 0
            excluded_models = self._get_excluded_models()

            for related in related_objects:
                related_model = related.related_model
                if related_model in excluded_models:
                    continue

                related_name = related.get_accessor_name()
                if related_name:
                    if hasattr(related_model, "all_objects"):
                        fk_field_name = related.field.name
                        related_records = related_model.all_objects.filter(
                            **{fk_field_name: obj}
                        )
                    else:
                        related_records = getattr(obj, related_name).all()

                    for rec in related_records:
                        field_name = [
                            f.name
                            for f in related_model._meta.fields
                            if f.related_model == self.model
                        ][0]
                        setattr(rec, field_name, new_target)
                        rec.save()
                        reassigned_count += 1
            return reassigned_count
        except ObjectDoesNotExist:
            raise ValueError(f"Target with id {new_target_id} does not exist")
        except Exception as e:
            raise

    def _perform_individual_action(self, record_id, actions):
        """
        Handle individual reassign, set null, or delete actions for dependent records.
        Actions is a dict with record IDs as keys and dicts {action: 'reassign/set_null/delete', new_target_id: ID} as values.
        Returns the number of processed records.
        """
        try:
            obj = self.model.all_objects.get(id=record_id)
            related_objects = self.model._meta.related_objects
            processed_count = 0
            excluded_models = self._get_excluded_models()

            for related in related_objects:
                related_model = related.related_model
                if related_model in excluded_models:
                    continue

                related_name = related.get_accessor_name()
                if related_name:
                    if hasattr(related_model, "all_objects"):
                        fk_field_name = related.field.name
                        related_records = related_model.objects.filter(
                            **{fk_field_name: obj}
                        )
                    else:
                        related_records = getattr(obj, related_name).all()
                    for rec in related_records:
                        if str(rec.id) in actions:
                            action = actions[str(rec.id)]
                            if action["action"] == "reassign" and action.get(
                                "new_target_id"
                            ):
                                try:
                                    new_target = self.model.all_objects.get(
                                        id=action["new_target_id"]
                                    )
                                    field_name = [
                                        f.name
                                        for f in related_model._meta.fields
                                        if f.related_model == self.model
                                    ][0]
                                    setattr(rec, field_name, new_target)
                                    rec.save()
                                    processed_count += 1
                                except ObjectDoesNotExist:
                                    continue
                            elif action[
                                "action"
                            ] == "set_null" and self._is_field_nullable(related_model):
                                field_name = [
                                    f.name
                                    for f in related_model._meta.fields
                                    if f.related_model == self.model
                                ][0]
                                setattr(rec, field_name, None)
                                rec.save()
                                processed_count += 1
                            elif action["action"] == "delete":
                                rec.delete()
                                processed_count += 1
            return processed_count
        except ObjectDoesNotExist:
            raise ValueError("Invalid target ID provided in individual actions")
        except Exception as e:
            raise

    def _delete_main_object(self, delete_mode, user=None):
        """
        Delete the main object based on the delete mode.
        """
        if delete_mode == "main_soft":
            RecycleBin.create_from_instance(self.object, user=user if user else None)
        self.object.delete()

    def get(self, request, *args, **kwargs):
        """
        Handle GET requests for delete view, including dependency check and form rendering.
        """
        if not self.request.user.is_authenticated:
            login_url = f"{reverse_lazy('horilla_core:login')}?next={request.path}"
            return redirect(login_url)
        try:

            self.object = self.get_object()
            record_id = self.object.id
            action = request.GET.get("action")
            view_id = request.GET.get("view_id", f"delete_{record_id}")
            delete_mode = request.GET.get(
                "delete_mode", "hard"
            )  # Get delete mode from request

            # Handle infinite scroll pagination for dependency records
            if action == "load_more_dependencies":
                related_name = request.GET.get("related_name")
                page = int(request.GET.get("page", 1))
                per_page = int(request.GET.get("per_page", 8))

                pagination_data = self._get_paginated_dependencies(
                    record_id, related_name, page, per_page
                )

                context = {
                    "records": pagination_data["records"],
                    "has_more": pagination_data["has_more"],
                    "next_page": pagination_data["next_page"],
                    "related_name": related_name,
                    "record_id": record_id,
                    "per_page": per_page,
                    "delete_mode": delete_mode,
                }
                return render(
                    request,
                    "partials/Single_delete/delete_dependency_partial.html",
                    context,
                )

            # Handle infinite scroll pagination for individual records
            if action == "load_more_individual_records":
                page = int(request.GET.get("page", 1))
                per_page = int(request.GET.get("per_page", 8))

                pagination_data = self._get_paginated_individual_records(
                    record_id, page, per_page
                )

                context = {
                    "records": pagination_data["records"],
                    "has_more": pagination_data["has_more"],
                    "next_page": pagination_data["next_page"],
                    "main_record_id": record_id,
                    "per_page": per_page,
                    "available_targets": pagination_data["available_targets"],
                    "is_nullable": pagination_data["is_nullable"],
                    "search_url": request.path,
                    "delete_mode": delete_mode,
                }
                return render(
                    request,
                    "partials/Single_delete/individual_reassign_partial.html",
                    context,
                )

            cannot_delete, can_delete, dependency_details = self._check_dependencies(
                record_id
            )
            available_targets = self.model.all_objects.exclude(id=record_id)
            dependent_records = []
            related_model = None
            is_nullable = False
            has_more_individual_records = False

            if cannot_delete:
                all_dependent_records = []
                for dep in cannot_delete[0]["dependencies"]:
                    related_model = dep["related_model"]
                    all_dependent_records.extend(dep["related_records"])
                    is_nullable = self._is_field_nullable(related_model)

                # For individual form, show first 8 records and check if there are more
                dependent_records = all_dependent_records[:8]
                has_more_individual_records = len(all_dependent_records) > 8

            context = {
                "object": self.object,
                "cannot_delete": cannot_delete,
                "can_delete": can_delete,
                "cannot_delete_count": len(cannot_delete),
                "can_delete_count": len(can_delete),
                "model_verbose_name": self.model._meta.verbose_name_plural,
                "search_url": request.path,
                "view_id": view_id,
                "record_id": record_id,
                "related_model": related_model,
                "dependent_records": dependent_records,
                "available_targets": available_targets,
                "is_nullable": is_nullable,
                "has_more_individual_records": has_more_individual_records,
                "delete_mode": delete_mode,
            }

            # Handle specific action requests
            if action == "show_bulk_reassign":
                return render(
                    request, "partials/Single_delete/bulk_reassign_form.html", context
                )
            elif action == "show_individual_reassign":
                return render(
                    request,
                    "partials/Single_delete/individual_reassign_form.html",
                    context,
                )
            elif action == "show_delete_confirmation":
                related_objects = self.model._meta.related_objects
                related_model = None
                related_verbose_name_plural = "records"
                for rel in related_objects:
                    related_model = rel.related_model
                    related_verbose_name_plural = (
                        related_model._meta.verbose_name_plural
                    )

                context["model_verbose_name"] = self.model._meta.verbose_name
                context["related_model"] = related_model if cannot_delete else None
                context["related_verbose_name_plural"] = related_verbose_name_plural
                context["hx_target"] = self.hx_target
                return render(
                    request, "partials/Single_delete/delete_all_confirm.html", context
                )

            if not request.GET.get("delete_mode"):
                return render(
                    request, "partials/Single_delete/delete_mode_modal.html", context
                )
            else:
                return render(
                    request,
                    "partials/Single_delete/delete_dependency_modal.html",
                    context,
                )

        except Exception as e:
            logger.error(f"Error in get method: {str(e)}")
            raise HorillaHttp404(e)

    def post(self, request, *args, **kwargs):
        try:
            self.object = self.get_object()
        except Exception as e:
            messages.error(self.request, _(str(e)))
            return HttpResponse(
                "<script>$('#reloadButton').click();$('#reloadMessagesButton').click();closeDeleteModeModal();</script>"
            )

        return super().post(request, *args, **kwargs)

    def delete(self, request, *args, **kwargs):
        """
        Handle POST requests for delete actions with dependency handling.
        """
        try:
            self.object = self.object
            record_id = self.object.id
            delete_mode = request.POST.get("delete_mode")
            action = request.POST.get("action")

            check_dependencies = request.POST.get("check_dependencies", "true")

            cannot_delete = []
            can_delete = []
            dependency_details = {}

            if not delete_mode and action != "check_dependencies_with_mode":
                context = {
                    "object": self.object,
                    "model_verbose_name": self.model._meta.verbose_name,
                    "search_url": request.path,
                    "view_id": request.GET.get("view_id", f"delete_{record_id}"),
                    "record_id": record_id,
                    "check_dependencies": check_dependencies,
                }
                return render(
                    request, "partials/Single_delete/delete_mode_modal.html", context
                )

            if check_dependencies == "false" and delete_mode:
                try:
                    with transaction.atomic():
                        self._delete_main_object(
                            delete_mode,
                            request.user if hasattr(request, "user") else None,
                        )

                    messages.success(request, self.get_success_message())
                    return self.get_post_delete_response()
                except Exception as e:
                    from django.utils.translation import gettext_lazy as _

                    logger.error(
                        f"Simple delete error for {self.model.__name__} id {record_id}: {str(e)}"
                    )
                    messages.info(
                        self.request,
                        _(
                            "Selected record is not associated with any company. Activate a company to proceed with deletion."
                        ),
                    )
                    return HttpResponse(
                        "<script>$('#reloadButton').click();$('#reloadMessagesButton').click();closeDeleteModeModal();</script>"
                    )

            if action == "check_dependencies_with_mode":
                cannot_delete, can_delete, dependency_details = (
                    self._check_dependencies(record_id)
                )

                available_targets = self.model.all_objects.exclude(id=record_id)
                dependent_records = []
                related_model = None
                is_nullable = False
                has_more_individual_records = False

                if cannot_delete:
                    all_dependent_records = []
                    for dep in cannot_delete[0]["dependencies"]:
                        related_model = dep["related_model"]
                        all_dependent_records.extend(dep["related_records"])
                        is_nullable = self._is_field_nullable(related_model)

                    dependent_records = all_dependent_records[:8]
                    has_more_individual_records = len(all_dependent_records) > 8

                context = {
                    "object": self.object,
                    "cannot_delete": cannot_delete,
                    "can_delete": can_delete,
                    "cannot_delete_count": len(cannot_delete),
                    "can_delete_count": len(can_delete),
                    "model_verbose_name": self.model._meta.verbose_name_plural,
                    "search_url": request.path,
                    "view_id": request.GET.get("view_id", f"delete_{record_id}"),
                    "record_id": record_id,
                    "related_model": related_model,
                    "dependent_records": dependent_records,
                    "available_targets": available_targets,
                    "is_nullable": is_nullable,
                    "has_more_individual_records": has_more_individual_records,
                    "delete_mode": delete_mode,
                    "reassign_all_visibility": self.reassign_all_visibility,
                    "reassign_individual_visibility": self.reassign_individual_visibility,
                    "hx_target": self.hx_target,
                }
                return render(
                    request,
                    "partials/Single_delete/delete_dependency_modal.html",
                    context,
                )

            if action == "bulk_reassign" and request.POST.get("new_target_id"):
                try:
                    with transaction.atomic():
                        new_target_id = int(request.POST.get("new_target_id"))
                        reassigned_count = self._perform_bulk_reassign(
                            record_id, new_target_id
                        )

                        self._delete_main_object(
                            delete_mode,
                            request.user if hasattr(request, "user") else None,
                        )

                    messages.success(
                        request,
                        f"Successfully reassigned {reassigned_count} records and deleted the {self.model._meta.verbose_name}.",
                    )
                    return HttpResponse(
                        "<script>htmx.trigger('#reloadButton','click');closeDeleteModal();closeModal();closeDeleteModeModal();</script>"
                    )
                except Exception as e:
                    logger.error(f"Bulk reassign error: {str(e)}")
                    return HttpResponse(
                        f"<script>alert('Error: {str(e)}');</script>", status=500
                    )

            elif action == "individual_action":
                try:
                    with transaction.atomic():
                        actions = {}
                        reassigned_count = 0
                        total_dependencies = 0

                        # Count total dependencies
                        cannot_delete, _, _ = self._check_dependencies(record_id)
                        if cannot_delete:
                            for dep in cannot_delete[0]["dependencies"]:
                                total_dependencies += dep["count"]

                        # Process actions
                        for key, value in request.POST.items():
                            if key.startswith("action_"):
                                record_id_key = key.replace("action_", "")
                                action_type = value
                                new_target_id = request.POST.get(
                                    f"new_target_{record_id_key}"
                                )
                                if action_type in ["reassign", "set_null", "delete"]:
                                    actions[record_id_key] = {
                                        "action": action_type,
                                        "new_target_id": (
                                            new_target_id
                                            if action_type == "reassign"
                                            and new_target_id
                                            else None
                                        ),
                                    }
                                    if action_type == "reassign" and new_target_id:
                                        try:
                                            self.model.objects.get(id=new_target_id)
                                            reassigned_count += 1
                                        except ObjectDoesNotExist:
                                            return HttpResponse(
                                                f"<script>alert('Invalid target ID');</script>",
                                                status=500,
                                            )

                        processed_count = self._perform_individual_action(
                            record_id, actions
                        )

                        remaining_cannot_delete, remaining_can_delete, _ = (
                            self._check_dependencies(record_id)
                        )

                        if not remaining_cannot_delete:
                            self._delete_main_object(
                                delete_mode,
                                request.user if hasattr(request, "user") else None,
                            )
                            if reassigned_count > 0:
                                messages.success(
                                    request,
                                    f"Reassigned {reassigned_count} records and deleted {self.object}",
                                )
                            else:
                                messages.success(
                                    request,
                                    f"Processed dependency records and deleted {self.object}",
                                )
                        elif processed_count > 0:
                            if reassigned_count > 0:
                                messages.success(
                                    request, f"Reassigned {reassigned_count} records"
                                )
                            else:
                                messages.success(
                                    request, f"Processed dependency records"
                                )

                    return HttpResponse(
                        "<script>htmx.trigger('#reloadButton','click');closeModal();closeDeleteModal();closeDeleteModeModal();</script>"
                    )
                except Exception as e:
                    return HttpResponse(
                        f"<script>alert('Error: {str(e)}');</script>", status=500
                    )

            elif action == "soft_delete_record":
                record_id_to_delete = request.POST.get("record_id")
                main_record_id = request.POST.get("main_record_id")

                if not record_id_to_delete or not main_record_id:
                    return HttpResponse(
                        "No record ID or main record ID provided", status=400
                    )

                try:
                    with transaction.atomic():
                        record_to_delete = None
                        obj = self.model.all_objects.get(id=main_record_id)
                        related_objects = self.model._meta.related_objects
                        excluded_models = self._get_excluded_models()

                        for related in related_objects:
                            related_model = related.related_model
                            if related_model in excluded_models:
                                continue

                            related_name = related.get_accessor_name()
                            if related_name:
                                try:
                                    record_to_delete = related_model.all_objects.get(
                                        id=record_id_to_delete
                                    )
                                    break
                                except ObjectDoesNotExist:
                                    continue

                        if record_to_delete:
                            RecycleBin.create_from_instance(
                                record_to_delete,
                                user=request.user if hasattr(request, "user") else None,
                            )
                            record_to_delete.delete()
                            messages.success(
                                request,
                                f"Successfully soft deleted {str(record_to_delete)}.",
                            )
                            return HttpResponse("")
                        else:
                            return HttpResponse("Record not found", status=404)

                except Exception as e:
                    logger.error(f"Soft delete error: {str(e)}")
                    return HttpResponse(
                        f"<script>alert('Error: {str(e)}');</script>", status=500
                    )

            elif action == "delete_single_record":
                record_id_to_delete = request.POST.get("record_id")
                main_record_id = request.POST.get("main_record_id")
                if not record_id_to_delete or not main_record_id:
                    return HttpResponse(
                        "No record ID or main record ID provided", status=400
                    )
                try:
                    with transaction.atomic():
                        record_to_delete = None
                        obj = self.model.all_objects.get(id=main_record_id)
                        related_objects = self.model._meta.related_objects
                        excluded_models = self._get_excluded_models()

                    for related in related_objects:
                        related_model = related.related_model
                        if related_model in excluded_models:
                            continue

                        related_name = related.get_accessor_name()
                        if related_name:
                            try:
                                record_to_delete = related_model.all_objects.get(
                                    id=record_id_to_delete
                                )
                                break
                            except ObjectDoesNotExist:
                                continue
                    if record_to_delete:
                        record_to_delete.delete()
                        messages.success(
                            request, f"Successfully deleted {str(record_to_delete)}."
                        )
                        return HttpResponse("")
                    else:
                        return HttpResponse("Record not found", status=404)

                except Exception as e:
                    return HttpResponse(
                        f"<script>alert('Error: {str(e)}');</script>", status=500
                    )

            elif action == "bulk_delete":
                try:
                    with transaction.atomic():
                        related_objects = self.model._meta.related_objects
                        excluded_models = self._get_excluded_models()

                        for related in related_objects:
                            related_model = related.related_model
                            if related_model in excluded_models:
                                continue

                            related_name = related.get_accessor_name()
                            if related_name:
                                if hasattr(related_model, "all_objects"):
                                    fk_field_name = related.field.name
                                    related_model.all_objects.filter(
                                        **{fk_field_name: self.object}
                                    ).delete()
                                else:
                                    getattr(self.object, related_name).all().delete()

                        self._delete_main_object(
                            delete_mode,
                            request.user if hasattr(request, "user") else None,
                        )

                    messages.success(
                        request,
                        f"Successfully deleted the {self.model._meta.verbose_name} and all its related records.",
                    )
                    return HttpResponse(
                        "<script>htmx.trigger('#reloadButton','click');closeDeleteModeModal();CloseDeleteConfirmModal();closeModal();</script>"
                    )
                except Exception as e:
                    logger.error(f"Bulk delete error: {str(e)}")
                    return HttpResponse(
                        f"<script>alert('Error: {str(e)}');</script>", status=500
                    )

            elif action == "simple_delete":
                try:
                    with transaction.atomic():
                        self._delete_main_object(
                            delete_mode,
                            request.user if hasattr(request, "user") else None,
                        )

                    messages.success(request, self.get_success_message())
                    return self.get_post_delete_response()
                except Exception as e:
                    from django.utils.translation import gettext_lazy as _

                    logger.error(
                        f"Simple delete error for {self.model.__name__} id {record_id}: {str(e)}"
                    )
                    messages.info(
                        self.request,
                        _(
                            "Selected record is not associated with any company. Activate a company to proceed with deletion."
                        ),
                    )
                    return HttpResponse(
                        "<script>$('#reloadButton').click();$('#reloadMessagesButton').click();closeDeleteModeModal();</script>"
                    )

            elif action == "set_null_action":
                record_id_to_update = request.POST.get("record_id")
                main_record_id = request.POST.get("main_record_id")
                if record_id_to_update and main_record_id:
                    try:
                        with transaction.atomic():
                            obj = self.model.all_objects.get(id=main_record_id)
                            related_objects = self.model._meta.related_objects
                            excluded_models = self._get_excluded_models()

                            updated = False
                            for related in related_objects:
                                related_model = related.related_model
                                if related_model in excluded_models:
                                    continue

                                related_name = related.get_accessor_name()

                                if related_name and self._is_field_nullable(
                                    related_model
                                ):
                                    try:
                                        record_to_update = (
                                            related_model.all_objects.get(
                                                id=record_id_to_update
                                            )
                                        )

                                        field_name = None
                                        for field in related_model._meta.fields:
                                            if field.related_model == self.model:
                                                field_name = field.name
                                                break

                                        if field_name:
                                            current_value = getattr(
                                                record_to_update, field_name
                                            )
                                            if current_value:
                                                setattr(
                                                    record_to_update, field_name, None
                                                )
                                                record_to_update.save()
                                                updated = True
                                                logger.info(
                                                    f"Set {field_name} to null for {related_model.__name__} id {record_id_to_update}"
                                                )
                                            break
                                    except ObjectDoesNotExist:
                                        continue

                            if updated:
                                messages.success(
                                    request, f"Successfully set record to null."
                                )

                            cannot_delete, can_delete, dependency_details = (
                                self._check_dependencies(main_record_id)
                            )

                            dependent_records = []
                            related_model = None
                            is_nullable = False
                            available_targets = self.model.all_objects.exclude(
                                id=main_record_id
                            )

                            if cannot_delete:
                                all_dependent_records = []
                                for dep in cannot_delete[0]["dependencies"]:
                                    related_model = dep["related_model"]
                                    all_dependent_records.extend(dep["related_records"])
                                    is_nullable = self._is_field_nullable(related_model)
                                dependent_records = all_dependent_records[:8]

                            context = {
                                "object": self.object,
                                "cannot_delete": cannot_delete,
                                "can_delete": can_delete,
                                "cannot_delete_count": len(cannot_delete),
                                "can_delete_count": len(can_delete),
                                "model_verbose_name": self.model._meta.verbose_name_plural,
                                "search_url": request.path,
                                "view_id": request.GET.get(
                                    "view_id", f"delete_{main_record_id}"
                                ),
                                "record_id": main_record_id,
                                "related_model": related_model,
                                "dependent_records": dependent_records,
                                "available_targets": available_targets,
                                "is_nullable": is_nullable,
                                "has_more_individual_records": (
                                    len(dependent_records) > 8
                                    if cannot_delete
                                    else False
                                ),
                                "delete_mode": delete_mode,
                                "hx_target": self.hx_target,
                            }

                            if cannot_delete:
                                return render(
                                    request,
                                    "partials/Single_delete/individual_reassign_form.html",
                                    context,
                                )
                            else:
                                return render(
                                    request,
                                    "partials/Single_delete/delete_dependency_modal.html",
                                    context,
                                )

                    except Exception as e:
                        logger.error(f"Set null action error: {str(e)}")
                        return HttpResponse(
                            f"<script>alert('Error: {str(e)}');</script>", status=500
                        )
                else:
                    return HttpResponse("No record ID provided", status=400)

            cannot_delete, can_delete, dependency_details = self._check_dependencies(
                record_id
            )
            if not cannot_delete:
                context = {
                    "object": self.object,
                    "cannot_delete": cannot_delete,
                    "can_delete": can_delete,
                    "cannot_delete_count": len(cannot_delete),
                    "can_delete_count": len(can_delete),
                    "model_verbose_name": self.model._meta.verbose_name_plural,
                    "search_url": request.path,
                    "view_id": request.GET.get("view_id", f"delete_{record_id}"),
                    "record_id": record_id,
                    "related_model": None,
                    "dependent_records": [],
                    "available_targets": self.model.all_objects.exclude(id=record_id),
                    "delete_mode": delete_mode,
                    "reassign_all_visibility": self.reassign_all_visibility,
                    "reassign_individual_visibility": self.reassign_individual_visibility,
                    "hx_target": self.hx_target,
                }
                return render(
                    request,
                    "partials/Single_delete/delete_dependency_modal.html",
                    context,
                )

            # return HttpResponse("Invalid request", status=400)
            messages.error(self.request, "Error in delete method")
            return HttpResponse(
                "<script>$('#reloadButton').click();closeDeleteModeModal();closeModal();</script>"
            )

        except Exception as e:
            logger.error(f"Error in delete method: {str(e)}")
            messages.error(self.request, f"Error in delete method: {str(e)}")
            return HttpResponse(
                "<script>$('#reloadButton').click();closeDeleteModeModal();</script>"
            )

    def get_post_delete_response(self):
        """
        Default post-delete behavior.
        """
        try:
            resolved_url = self.success_url or self.get_success_url()
            if resolved_url:
                return redirect(resolved_url)
        except Exception as e:
            logger.error(f"Error getting success URL: {str(e)}")
            return HttpResponse(
                f"<script>alert('Error: {str(e)}');</script>", status=500
            )
        return HttpResponse(
            "<script>htmx.trigger('#reloadButton','click');closeDeleteModeModal();</script>"
        )

    def form_valid(self, form):
        """
        Handle form submission by calling the delete method.
        """
        return self.delete(self.request, *self.args, **self.kwargs)

    def get_success_message(self):
        """
        Return the success message for deletion.
        """
        return self.success_message


@method_decorator(htmx_required, name="dispatch")
@method_decorator(
    permission_required_or_denied("horilla_core.delete_horillaattachment", modal=True),
    name="dispatch",
)
class HorillaNotesAttachmentDeleteView(LoginRequiredMixin, HorillaSingleDeleteView):
    model = HorillaAttachment

    def get_post_delete_response(self):
        return HttpResponse("<script>htmx.trigger('#reloadButton','click');</script>")


class HorillaModalDetailView(DetailView):
    """
    HorillDetailedView
    """

    title = "Detailed View"
    template_name = "single_detail_view.html"
    header: dict = {
        "title": "Horilla",
        "subtitle": "Horilla Detailed View",
        "avatar": "",
    }
    body: list = []

    action_method: list = []
    actions: list = []
    cols: dict = {}
    instance = None
    empty_template = None

    ids_key: str = "instance_ids"

    def get_queryset(self):
        """
        Filter queryset based on instance_ids from session.
        """
        queryset = super().get_queryset()
        instance_ids = self.request.session.get(self.ordered_ids_key, [])
        if instance_ids:
            queryset = queryset.filter(pk__in=instance_ids)
        return queryset

    def get_object(self, queryset=None):
        if queryset is None:
            queryset = self.get_queryset()
        try:
            self.instance = super().get_object(queryset)
        except Exception as e:
            logger.error(f"Error getting object: {e}")
        return self.instance

    def get(self, request, *args, **kwargs):
        if not self.request.GET.get(self.ids_key) and not self.request.session.get(
            self.ordered_ids_key
        ):
            self.request.session[self.ordered_ids_key] = []
        response = super().get(request, *args, **kwargs)
        if not self.instance and self.empty_template:
            return render(request, self.empty_template, context=self.get_context_data())
        elif not self.instance:
            messages.error(request, "The requested record does not exist.")
            return HttpResponse("<script>$('#reloadButton').click();</script>")
        return response

    def __init__(self, **kwargs: Any) -> None:
        super().__init__(**kwargs)
        self.ordered_ids_key = f"ordered_ids_{self.model.__name__.lower()}"
        request = getattr(_thread_local, "request", None)
        self.request = request
        # update_initial_cache(request, CACHE, HorillaDetailedView)

    def get_context_data(self, **kwargs: Any):
        context = super().get_context_data(**kwargs)
        obj = context.get("object")

        if not obj:
            return context

        pk = obj.pk
        instance_ids = self.request.session.get(self.ordered_ids_key, [])
        url_info = resolve(self.request.path)
        url_name = url_info.url_name
        key = next(iter(url_info.kwargs), "pk")

        context["instance"] = obj
        context["title"] = self.title
        context["header"] = self.header
        context["body"] = self.body
        context["actions"] = self.actions
        context["action_method"] = self.action_method
        context["cols"] = self.cols

        if instance_ids:
            prev_id, next_id = closest_numbers(instance_ids, pk)

            full_url_name = (
                f"{url_info.namespaces[0]}:{url_name}"
                if url_info.namespaces
                else url_name
            )
            context.update(
                {
                    "instance_ids": str(instance_ids),
                    "ids_key": self.ids_key,
                    "next_url": reverse_lazy(full_url_name, kwargs={key: next_id}),
                    "previous_url": reverse_lazy(full_url_name, kwargs={key: prev_id}),
                }
            )

            # Filter out instance_ids key from GET params
            get_params = self.request.GET.copy()
            get_params.pop(self.ids_key, None)
            context["extra_query"] = get_params.urlencode()
        else:
            context["extra_query"] = ""

        return context
