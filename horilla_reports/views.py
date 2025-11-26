import copy
import csv
import io
import json
import logging
from datetime import datetime
from functools import cached_property
from urllib.parse import urlencode, urlparse

import openpyxl
import pandas as pd
from django import forms
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import redirect_to_login
from django.contrib.contenttypes.models import ContentType
from django.db.models import ForeignKey, Q
from django.http import Http404, HttpResponse, QueryDict
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils.decorators import method_decorator
from django.utils.translation import gettext_lazy as _
from django.views import View
from django.views.decorators.http import require_POST
from django.views.generic import DetailView
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from horilla.exceptions import HorillaHttp404
from horilla_core.decorators import (
    htmx_required,
    permission_required,
    permission_required_or_denied,
)
from horilla_generics.forms import HorillaModelForm
from horilla_generics.mixins import RecentlyViewedMixin
from horilla_generics.views import (
    HorillaListView,
    HorillaNavView,
    HorillaSingleDeleteView,
    HorillaSingleFormView,
)
from horilla_reports.filters import ReportFilter
from horilla_reports.forms import ChangeChartReportForm, ReportForm
from horilla_reports.models import Report, ReportFolder
from horilla_utils.methods import get_section_info_for_model
from horilla_utils.middlewares import _thread_local

logger = logging.getLogger(__name__)


@method_decorator(htmx_required, name="dispatch")
@method_decorator(
    permission_required(
        ["horilla_reports.view_report", "horilla_reports.view_own_report"]
    ),
    name="dispatch",
)
class ReportNavbar(LoginRequiredMixin, HorillaNavView):

    search_url = reverse_lazy("horilla_reports:reports_list_view")
    main_url = reverse_lazy("horilla_reports:reports_list_view")
    filterset_class = ReportFilter
    one_view_only = True
    filter_option = False
    reload_option = False
    gap_enabled = False
    model_name = "Report"
    model_app_label = "reports"
    search_option = False
    all_view_types = False
    enable_actions = True

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        request = getattr(_thread_local, "request", None)
        title = request.GET.get("title")
        if title == "Reports":
            self.all_view_types = True
        else:
            self.all_view_types = False

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        title = self.request.GET.get("title")
        context["nav_title"] = title
        return context

    @cached_property
    def new_button(self):
        if self.request.user.has_perm("horilla_reports.add_report"):
            return {
                "title": "New Report",
                "url": f"""{ reverse_lazy('horilla_reports:create_report')}""",
                "attrs": {"id": "report-create"},
            }

    @cached_property
    def actions(self):
        """Actions for lead"""
        if self.request.user.has_perm("horilla_reports.view_report"):
            return [
                {
                    "action": "Load Default Reports",
                    "attrs": f"""
                            id="reports-load"
                            hx-get="{reverse_lazy('horilla_reports:load_default_reports')}"
                            hx-on:click="openModal();"
                            hx-target="#modalBox"
                            hx-swap="innerHTML"
                            """,
                },
            ]

    @cached_property
    def second_button(self):
        if self.request.user.has_perm("horilla_reports.add_reportfolder"):
            return {
                "title": "New Folder",
                "url": f"""{ reverse_lazy('horilla_reports:create_folder')}?pk={self.request.GET.get('pk', '')}""",
                "attrs": {"id": "report-folder-create"},
            }


@method_decorator(
    permission_required_or_denied(
        ["horilla_reports.view_report", "horilla_reports.view_own_report"]
    ),
    name="dispatch",
)
class ReportsListView(LoginRequiredMixin, HorillaListView):

    model = Report
    template_name = "report_list_view.html"
    view_id = "reports-list"
    filterset_class = ReportFilter
    search_url = reverse_lazy("horilla_reports:reports_list_view")
    main_url = reverse_lazy("horilla_reports:reports_list_view")
    table_width = False
    max_visible_actions = 5
    sorting_target = f"#tableview-{view_id}"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Reports"
        return context

    def no_record_add_button(self):
        if self.request.user.has_perm("horilla_reports.add_reports"):
            return {
                "url": f"""{ reverse_lazy('horilla_reports:load_default_reports')}?new=true""",
                "attrs": 'id="reports-load"',
                "title": "Load Default Reports",
            }

    columns = ["name", (_("Module"), "module_verbose_name"), "folder"]

    @cached_property
    def action_method(self):
        action_method = ""
        if self.request.user.has_perm(
            "horilla_reports.change_report"
        ) or self.request.user.has_perm("horilla_reports.delete_report"):
            action_method = "actions"
        return action_method

    @cached_property
    def col_attrs(self):
        query_params = {}
        if "section" in self.request.GET:
            query_params["section"] = self.request.GET.get("section")
        query_string = urlencode(query_params)
        attrs = {}
        if self.request.user.has_perm(
            "horilla_reports.view_report"
        ) or self.request.user.has_perm("horilla_reports.view_own_report"):
            attrs = {
                "hx-get": f"{{get_detail_view_url}}?{query_string}",
                "hx-target": "#mainContent",
                "hx-swap": "outerHTML",
                "hx-push-url": "true",
                "hx-select": "#mainContent",
                "style": "cursor:pointer",
                "class": "hover:text-primary-600",
            }
        return [
            {
                "name": {
                    **attrs,
                }
            }
        ]


@method_decorator(
    permission_required_or_denied(
        ["horilla_reports.view_report", "horilla_reports.view_own_report"]
    ),
    name="dispatch",
)
class FavouriteReportsListView(LoginRequiredMixin, HorillaListView):

    model = Report
    template_name = "favourite_report_list_view.html"
    view_id = "favourite-reports-list"
    filterset_class = ReportFilter
    search_url = reverse_lazy("horilla_reports:favourite_reports_list_view")
    main_url = reverse_lazy("horilla_reports:favourite_reports_list_view")
    table_width = False
    sorting_target = f"#tableview-{view_id}"

    @cached_property
    def action_method(self):
        action_method = ""
        if self.request.user.has_perm(
            "horilla_reports.change_report"
        ) or self.request.user.has_perm("horilla_reports.delete_report"):
            action_method = "actions"
        return action_method

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Favourite Reports"
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.filter(is_favourite=True)
        return queryset

    columns = ["name", (_("Module"), "module_verbose_name"), "folder"]

    @cached_property
    def col_attrs(self):
        query_params = {}
        if "section" in self.request.GET:
            query_params["section"] = self.request.GET.get("section")
        query_string = urlencode(query_params)
        attrs = {}
        if self.request.user.has_perm(
            "horilla_reports.view_report"
        ) or self.request.user.has_perm("horilla_reports.view_own_report"):
            attrs = {
                "hx-get": f"{{get_detail_view_url}}?{query_string}",
                "hx-target": "#mainContent",
                "hx-swap": "outerHTML",
                "hx-push-url": "true",
                "hx-select": "#mainContent",
                "style": "cursor:pointer",
                "class": "hover:text-primary-600",
            }
        return [
            {
                "name": {
                    **attrs,
                }
            }
        ]


@method_decorator(
    permission_required_or_denied(
        ["horilla_reports.view_report", "horilla_reports.view_own_report"]
    ),
    name="dispatch",
)
class ReportDetailView(RecentlyViewedMixin, LoginRequiredMixin, DetailView):
    model = Report
    template_name = "report_detail.html"
    context_object_name = "report"

    def col_attrs(self):
        """Define column attributes for clickable rows in the report list view."""
        query_params = {}
        report = self.object
        model_class = report.model_class
        section = get_section_info_for_model(model_class)
        section_value = section["section"]
        query_params["section"] = section_value
        query_params["session_url"] = False
        query_string = urlencode(query_params)
        attrs = {}

        if self.request.user.has_perm("horilla_reports.view_report"):
            attrs = {
                "hx-get": f"{{get_detail_url}}?{query_string}",
                "hx-target": "#mainContent",
                "hx-swap": "outerHTML",
                "hx-push-url": "true",
                "hx-select-oob": "#sideMenuContainer",
                "hx-select": "#mainContent",
                "style": "cursor:pointer",
                "class": "hover:text-primary-600",
            }

        columns_with_attrs = []

        for col in report.selected_columns_list:
            columns_with_attrs.append({col: {**attrs}})

        return columns_with_attrs

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if not self.model.objects.filter(
            report_owner_id=self.request.user, pk=self.kwargs["pk"]
        ).first() and not self.request.user.has_perm("horilla_reports.view_report"):
            return render(self.request, "error/403.html")
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        report = self.object

        session_key = f"report_preview_{report.pk}"
        preview_data = self.request.session.get(session_key, {})

        temp_report = self.create_temp_report(report, preview_data)

        aggregate_columns_dict = temp_report.aggregate_columns_dict
        if not isinstance(aggregate_columns_dict, list):
            aggregate_columns_dict = (
                [aggregate_columns_dict] if aggregate_columns_dict else []
            )

        # Get model data
        model_class = temp_report.model_class
        queryset = model_class.objects.all()

        # Apply filters
        filters = temp_report.filters_dict
        if filters:
            query = None
            for index, (field_name, filter_data) in enumerate(filters.items()):
                if not filter_data.get("value"):
                    continue  # Skip empty filters
                operator = filter_data.get("operator", "exact")
                value = filter_data.get("value")
                logic = (
                    filter_data.get("logic", "and") if index > 0 else "and"
                )  # Default to AND for first filter

                # Use original_field instead of field_name
                actual_field = filter_data.get("original_field", field_name)

                # Construct filter kwargs
                filter_kwargs = {}
                if operator == "exact":
                    filter_kwargs[f"{actual_field}"] = value
                elif operator == "icontains":
                    filter_kwargs[f"{actual_field}__icontains"] = value
                elif operator == "gt":
                    filter_kwargs[f"{actual_field}__gt"] = value
                elif operator == "lt":
                    filter_kwargs[f"{actual_field}__lt"] = value
                elif operator == "gte":
                    filter_kwargs[f"{actual_field}__gte"] = value
                elif operator == "lte":
                    filter_kwargs[f"{actual_field}__lte"] = value

                # Combine filters with AND or OR
                if not filter_kwargs:
                    continue
                current_query = Q(**filter_kwargs)

                if query is None:
                    query = current_query
                elif logic == "or":
                    query |= current_query
                else:  # logic == 'and'
                    query &= current_query

            if query:
                queryset = queryset.filter(query)

        # Convert queryset to DataFrame
        fields = []
        if temp_report.selected_columns_list:
            fields.extend(temp_report.selected_columns_list)
        if temp_report.row_groups_list:
            fields.extend(temp_report.row_groups_list)
        if temp_report.column_groups_list:
            fields.extend(temp_report.column_groups_list)
        for agg in aggregate_columns_dict:
            if agg.get("field"):
                fields.append(agg["field"])

        # Remove duplicates while preserving order
        fields = list(dict.fromkeys(fields))

        data = list(queryset.values(*fields)) if fields else list(queryset.values())
        df = pd.DataFrame(data)

        # Initialize context
        context["panel_open"] = bool(preview_data)
        context["hierarchical_data"] = []
        context["pivot_columns"] = []
        context["pivot_table"] = {}
        context["pivot_index"] = []
        context["aggregate_columns"] = []
        context["has_hierarchical_groups"] = len(temp_report.row_groups_list) > 1
        context["configuration_type"] = self.get_configuration_type(temp_report)
        panel_open = self.request.GET.get("panel_open") == "true" or bool(preview_data)
        context["panel_open"] = panel_open
        context["has_unsaved_changes"] = bool(preview_data)

        # Add verbose names for row and column groups
        context["row_group_verbose_names"] = [
            model_class._meta.get_field(field_name).verbose_name.title()
            for field_name in temp_report.row_groups_list
        ]
        context["column_group_verbose_names"] = [
            model_class._meta.get_field(field_name).verbose_name.title()
            for field_name in temp_report.column_groups_list
        ]

        # Handle different configurations
        row_count = len(temp_report.row_groups_list)
        col_count = len(temp_report.column_groups_list)

        if row_count == 0 and col_count == 0:
            self.handle_0_row_0_col(df, temp_report, context)
        elif row_count == 1 and col_count == 0:
            self.handle_1_row_0_col(df, temp_report, context)
        elif row_count == 1 and col_count == 1:
            self.handle_1_row_1_col(df, temp_report, context)
        elif row_count == 1 and col_count == 2:
            self.handle_1_row_2_col(df, temp_report, context)
        elif row_count == 2 and col_count == 0:
            self.handle_2_row_0_col(df, temp_report, context)
        elif row_count == 2 and col_count == 1:
            self.handle_2_row_1_col(df, temp_report, context)
        elif row_count == 3 and col_count == 0:
            self.handle_3_row_0_col(df, temp_report, context)
        else:
            context["error"] = (
                f"Configuration not supported: {row_count} rows, {col_count} columns"
            )

        # Chart data
        chart_data = self.generate_chart_data(df, temp_report)
        context["chart_data"] = chart_data
        context["total_count"] = len(data)
        context["total_amount"] = sum(
            [
                float(
                    df[agg["field"]].sum()
                    if agg["field"] in df.columns and agg.get("aggfunc") == "sum"
                    else 0
                )
                for agg in aggregate_columns_dict
            ]
        )

        columns = []
        for col in temp_report.selected_columns_list:
            field = model_class._meta.get_field(col)
            verbose_name = field.verbose_name.title()
            if field.choices:
                columns.append((verbose_name, f"get_{col}_display"))
            else:
                columns.append((verbose_name, col))

        list_view = HorillaListView(
            model=model_class,
            view_id="report-details-sec",
            search_url=reverse_lazy(
                "horilla_reports:report_detail", kwargs={"pk": report.pk}
            ),
            main_url=reverse_lazy(
                "horilla_reports:report_detail", kwargs={"pk": report.pk}
            ),
            table_width=False,
            columns=columns,
        )
        list_view.request = self.request
        list_view.table_width = False
        list_view.bulk_select_option = False
        list_view.clear_session_button_enabled = False
        list_view.list_column_visibility = False
        list_view.table_height = False
        list_view.table_height_as_class = "h-[200px]"
        if hasattr(report.model_class, "get_detail_url"):
            list_view.col_attrs = self.col_attrs()
        sort_field = self.request.GET.get("sort")
        sort_direction = self.request.GET.get("direction", "asc")

        if sort_field:
            queryset = list_view._apply_sorting(queryset, sort_field, sort_direction)
        else:
            queryset = queryset.order_by("-id")
        list_view.object_list = queryset
        context.update(list_view.get_context_data(object_list=queryset))
        session_referer_key = f"report_detail_referer_{report.pk}"
        current_referer = self.request.META.get("HTTP_REFERER")
        hx_current_url = self.request.headers.get("HX-Current-URL")
        stored_referer = self.request.session.get(session_referer_key)
        report_detail_base = f"/reports/report-detail/{report.pk}/"
        session_url_value = self.request.GET.get("session_url")

        if hx_current_url:
            hx_path = urlparse(hx_current_url).path
            is_from_report_detail = hx_path == report_detail_base
            if not is_from_report_detail and session_url_value != "False":
                self.request.session[session_referer_key] = hx_current_url
                previous_url = hx_current_url
            else:
                previous_url = (
                    stored_referer
                    if stored_referer
                    else reverse_lazy("horilla_reports:reports_list_view")
                )
        elif stored_referer:
            previous_url = stored_referer
        elif current_referer and self.request.get_host() in current_referer:
            referer_path = urlparse(current_referer).path
            if referer_path != report_detail_base:
                previous_url = current_referer
                self.request.session[session_referer_key] = current_referer
            else:
                previous_url = reverse_lazy("horilla_reports:reports_list_view")
        else:
            previous_url = reverse_lazy("horilla_reports:reports_list_view")
        context["previous_url"] = previous_url
        context["total_groups_count"] = len(temp_report.row_groups_list) + len(
            temp_report.column_groups_list
        )
        return context

    def create_temp_report(self, original_report, preview_data):
        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        if "chart_type" in preview_data:
            temp_report.chart_type = preview_data["chart_type"]
        if "chart_field" in preview_data:
            temp_report.chart_field = preview_data["chart_field"]
        if "chart_field_stacked" in preview_data:
            temp_report.chart_field_stacked = preview_data["chart_field_stacked"]
        return temp_report

    def get_configuration_type(self, report):
        row_count = len(report.row_groups_list)
        col_count = len(report.column_groups_list)
        return f"{row_count}_row_{col_count}_col"

    def get_verbose_name(self, field_name, model_class):
        """Get the verbose name of a field"""
        try:
            return model_class._meta.get_field(field_name).verbose_name.title()
        except:
            return field_name.title()

    def handle_0_row_0_col(self, df, report, context):
        try:
            aggregate_columns = []
            if report.aggregate_columns_dict:
                for agg in report.aggregate_columns_dict:
                    aggregate_field = agg.get("field")
                    aggfunc = agg.get("aggfunc", "sum")
                    aggregate_column_name = f"{aggfunc.title()} of {self.get_verbose_name(aggregate_field, report.model_class)}"
                    if aggregate_field and not df.empty:
                        if aggfunc == "sum":
                            total_value = df[aggregate_field].sum()
                        elif aggfunc == "avg":
                            total_value = df[aggregate_field].mean()
                        elif aggfunc == "min":
                            total_value = df[aggregate_field].min()
                        elif aggfunc == "max":
                            total_value = df[aggregate_field].max()
                        elif aggfunc == "count":
                            total_value = len(df)
                        else:
                            total_value = len(df)
                        aggregate_columns.append(
                            {
                                "name": aggregate_column_name,
                                "function": aggfunc,
                                "value": total_value,
                                "field": aggregate_field,
                            }
                        )
                context["simple_aggregate"] = {
                    "field": (
                        aggregate_columns[0]["field"]
                        if aggregate_columns
                        else "Records"
                    ),
                    "value": (
                        aggregate_columns[0]["value"] if aggregate_columns else len(df)
                    ),
                    "function": (
                        aggregate_columns[0]["function"]
                        if aggregate_columns
                        else "count"
                    ),
                }
            else:
                context["simple_aggregate"] = {
                    "field": "Records",
                    "value": len(df),
                    "function": "count",
                }
            context["aggregate_columns"] = aggregate_columns
        except Exception as e:
            context["error"] = f"Error in 0x0 configuration: {str(e)}"
            context["aggregate_columns"] = []

    def handle_1_row_0_col(self, df, report, context):
        try:
            if df.empty:
                context["pivot_index"] = []
                context["pivot_table"] = {}
                context["pivot_columns"] = ["Count"]
                context["aggregate_columns"] = []
                return

            model_class = report.model_class
            row_field = report.row_groups_list[0]

            # Always compute counts
            count_grouped = df.groupby(row_field).size().to_dict()
            display_grouped = {}
            display_rows = []
            pivot_columns = ["Count"]

            # Compute aggregate values for all aggregate columns
            aggregate_columns = []
            for agg in report.aggregate_columns_dict:
                aggregate_field = agg["field"]
                aggfunc = agg.get("aggfunc", "sum")
                aggregate_column_name = f"{aggfunc.title()} of {self.get_verbose_name(aggregate_field, model_class)}"
                pivot_columns.append(aggregate_column_name)

                if aggfunc == "sum":
                    aggregate_data = (
                        df.groupby(row_field)[aggregate_field].sum().to_dict()
                    )
                elif aggfunc == "avg":
                    aggregate_data = (
                        df.groupby(row_field)[aggregate_field].mean().to_dict()
                    )
                elif aggfunc == "min":
                    aggregate_data = (
                        df.groupby(row_field)[aggregate_field].min().to_dict()
                    )
                elif aggfunc == "max":
                    aggregate_data = (
                        df.groupby(row_field)[aggregate_field].max().to_dict()
                    )
                elif aggfunc == "count":
                    aggregate_data = df.groupby(row_field).size().to_dict()
                else:
                    aggregate_data = df.groupby(row_field).size().to_dict()

                aggregate_columns.append(
                    {
                        "name": aggregate_column_name,
                        "function": aggfunc,
                        "field": aggregate_field,
                        "data": aggregate_data,
                    }
                )

            # Convert to display values
            for row, count in count_grouped.items():
                display_info = self.get_display_value(row, row_field, model_class)
                composite_key = display_info["composite_key"]
                display_grouped[composite_key] = {
                    "Count": count,
                    "_display": display_info["display"],
                    "_id": display_info["id"],
                }
                for agg in aggregate_columns:
                    display_grouped[composite_key][agg["name"]] = agg["data"].get(
                        row, 0
                    )
                display_rows.append(composite_key)

            context["pivot_index"] = display_rows
            context["pivot_table"] = display_grouped
            context["pivot_columns"] = pivot_columns
            context["aggregate_columns"] = aggregate_columns

        except Exception as e:
            context["error"] = f"Error in 1x0 configuration: {str(e)}"
            context["aggregate_columns"] = []

    def handle_1_row_1_col(self, df, report, context):
        try:
            if df.empty:
                context["pivot_index"] = []
                context["pivot_table"] = {}
                context["pivot_columns"] = []
                context["aggregate_columns"] = []
                return

            model_class = report.model_class
            row_field = report.row_groups_list[0]
            col_field = report.column_groups_list[0]

            # Compute count-based pivot table
            pivot_table = pd.pivot_table(
                df, index=[row_field], columns=[col_field], aggfunc="size", fill_value=0
            )

            # Convert to display format
            pivot_dict = pivot_table.to_dict("index")
            transposed_dict = {}
            all_rows = pivot_table.index.tolist()
            all_columns = pivot_table.columns.tolist()

            # Convert row indices to display values
            display_rows = []
            display_columns = []
            for row in all_rows:
                display_info = self.get_display_value(row, row_field, model_class)
                composite_key = display_info["composite_key"]
                display_rows.append(composite_key)
                transposed_dict[composite_key] = {
                    "total": 0,
                    "_display": display_info["display"],
                    "_id": display_info["id"],
                }
                for col in all_columns:
                    col_info = self.get_display_value(col, col_field, model_class)
                    col_composite = col_info["composite_key"]
                    if col_composite not in display_columns:
                        display_columns.append(col_composite)
                    value = pivot_dict.get(row, {}).get(col, 0)
                    transposed_dict[composite_key][col_composite] = value
                    transposed_dict[composite_key]["total"] += value

            # Compute aggregate columns
            aggregate_columns = []
            for agg in report.aggregate_columns_dict:
                aggregate_field = agg["field"]
                aggfunc = agg.get("aggfunc", "sum")
                aggregate_column_name = f"{aggfunc.title()} of {self.get_verbose_name(aggregate_field, model_class)}"
                if aggfunc == "sum":
                    aggregate_data = (
                        df.groupby(row_field)[aggregate_field].sum().to_dict()
                    )
                elif aggfunc == "avg":
                    aggregate_data = (
                        df.groupby(row_field)[aggregate_field].mean().to_dict()
                    )
                elif aggfunc == "min":
                    aggregate_data = (
                        df.groupby(row_field)[aggregate_field].min().to_dict()
                    )
                elif aggfunc == "max":
                    aggregate_data = (
                        df.groupby(row_field)[aggregate_field].max().to_dict()
                    )
                elif aggfunc == "count":
                    aggregate_data = df.groupby(row_field).size().to_dict()
                else:
                    aggregate_data = df.groupby(row_field).size().to_dict()

                # Add aggregate values to transposed_dict
                for row in all_rows:
                    display_value = self.get_display_value(row, row_field, model_class)
                    display_info = self.get_display_value(row, row_field, model_class)
                    composite_key = display_info["composite_key"]
                    transposed_dict[composite_key][aggregate_column_name] = (
                        aggregate_data.get(row, 0)
                    )
                display_columns.append(aggregate_column_name)
                aggregate_columns.append(
                    {
                        "name": aggregate_column_name,
                        "function": aggfunc,
                        "field": aggregate_field,
                    }
                )

            context["pivot_table"] = transposed_dict
            context["pivot_index"] = display_rows
            context["pivot_columns"] = display_columns
            context["aggregate_columns"] = aggregate_columns

        except Exception as e:
            context["error"] = f"Error in 1x1 configuration: {str(e)}"
            context["aggregate_columns"] = []

    def handle_1_row_2_col(self, df, report, context):
        try:
            if df.empty:
                context["pivot_index"] = []
                context["pivot_table"] = {}
                context["pivot_columns"] = []
                context["column_hierarchy"] = []
                context["aggregate_columns"] = []
                return

            model_class = report.model_class
            row_field = report.row_groups_list[0]
            col_field1 = report.column_groups_list[0]
            col_field2 = report.column_groups_list[1]

            # Compute count-based pivot table
            pivot_table = pd.pivot_table(
                df,
                index=[row_field],
                columns=[col_field1, col_field2],
                aggfunc="size",
                fill_value=0,
            )

            # Handle multi-level columns
            pivot_dict = pivot_table.to_dict("index")
            transposed_dict = {}
            all_rows = pivot_table.index.tolist()
            column_hierarchy = []
            multi_level_columns = []

            for col_tuple in pivot_table.columns:
                col1_info = self.get_display_value(
                    col_tuple[0], col_field1, model_class
                )
                col2_info = self.get_display_value(
                    col_tuple[1], col_field2, model_class
                )
                col1_composite = col1_info["composite_key"]
                col2_composite = col2_info["composite_key"]
                column_key = f"{col1_composite}|{col2_composite}"
                multi_level_columns.append(column_key)
                column_hierarchy.append(
                    {
                        "level1": col1_composite,
                        "level1_display": col1_info["display"],
                        "level2": col2_composite,
                        "level2_display": col2_info["display"],
                        "key": column_key,
                    }
                )

            # Convert row data
            display_rows = []
            for row in all_rows:
                row_info = self.get_display_value(row, row_field, model_class)
                row_composite = row_info["composite_key"]
                display_rows.append(row_composite)
                transposed_dict[row_composite] = {
                    "total": 0,
                    "_display": row_info["display"],
                    "_id": row_info["id"],
                }
                for col_tuple in pivot_table.columns:
                    col1_info = self.get_display_value(
                        col_tuple[0], col_field1, model_class
                    )
                    col2_info = self.get_display_value(
                        col_tuple[1], col_field2, model_class
                    )
                    col1_composite = col1_info["composite_key"]
                    col2_composite = col2_info["composite_key"]
                    column_key = f"{col1_composite}|{col2_composite}"
                    value = pivot_dict.get(row, {}).get(col_tuple, 0)
                    transposed_dict[row_composite][column_key] = value
                    transposed_dict[row_composite]["total"] += value

            # Compute aggregate columns
            aggregate_columns = []
            for agg in report.aggregate_columns_dict:
                aggregate_field = agg["field"]
                aggfunc = agg.get("aggfunc", "sum")
                aggregate_column_name = f"{aggfunc.title()} of {self.get_verbose_name(aggregate_field, model_class)}"
                if aggfunc == "sum":
                    aggregate_data = (
                        df.groupby(row_field)[aggregate_field].sum().to_dict()
                    )
                elif aggfunc == "avg":
                    aggregate_data = (
                        df.groupby(row_field)[aggregate_field].mean().to_dict()
                    )
                elif aggfunc == "min":
                    aggregate_data = (
                        df.groupby(row_field)[aggregate_field].min().to_dict()
                    )
                elif aggfunc == "max":
                    aggregate_data = (
                        df.groupby(row_field)[aggregate_field].max().to_dict()
                    )
                elif aggfunc == "count":
                    aggregate_data = df.groupby(row_field).size().to_dict()
                else:
                    aggregate_data = df.groupby(row_field).size().to_dict()

                # Add aggregate values to transposed_dict
                for row in all_rows:
                    row_info = self.get_display_value(row, row_field, model_class)
                    row_composite = row_info["composite_key"]
                    transposed_dict[row_composite][aggregate_column_name] = (
                        aggregate_data.get(row, 0)
                    )
                multi_level_columns.append(aggregate_column_name)
                column_hierarchy.append(
                    {
                        "level1": aggregate_column_name,
                        "level1_display": aggregate_column_name,
                        "level2": "",
                        "level2_display": "",
                        "key": aggregate_column_name,
                    }
                )
                aggregate_columns.append(
                    {
                        "name": aggregate_column_name,
                        "function": aggfunc,
                        "field": aggregate_field,
                    }
                )

            context["pivot_table"] = transposed_dict
            context["pivot_index"] = display_rows
            context["pivot_columns"] = multi_level_columns
            context["column_hierarchy"] = column_hierarchy
            context["aggregate_columns"] = aggregate_columns

        except Exception as e:
            context["error"] = f"Error in 1x2 configuration: {str(e)}"
            context["aggregate_columns"] = []

    def handle_2_row_0_col(self, df, report, context):
        try:
            if df.empty:
                context["hierarchical_data"] = {"groups": [], "grand_total": 0}
                context["aggregate_columns"] = []
                return

            hierarchical_data = []
            primary_group = report.row_groups_list[0]
            secondary_group = report.row_groups_list[1]
            model_class = report.model_class
            pivot_columns = ["Count"]

            # Group by primary group
            primary_groups = df.groupby(primary_group)
            grand_total = 0

            # Compute aggregate columns
            aggregate_columns = []
            aggregate_data = {}
            for agg in report.aggregate_columns_dict:
                aggregate_field = agg["field"]
                aggfunc = agg.get("aggfunc", "sum")
                aggregate_column_name = f"{aggfunc.title()} of {self.get_verbose_name(aggregate_field, model_class)}"
                pivot_columns.append(aggregate_column_name)
                if aggfunc == "sum":
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([primary_group, secondary_group])[aggregate_field]
                        .sum()
                        .to_dict()
                    )
                elif aggfunc == "avg":
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([primary_group, secondary_group])[aggregate_field]
                        .mean()
                        .to_dict()
                    )
                elif aggfunc == "min":
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([primary_group, secondary_group])[aggregate_field]
                        .min()
                        .to_dict()
                    )
                elif aggfunc == "max":
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([primary_group, secondary_group])[aggregate_field]
                        .max()
                        .to_dict()
                    )
                elif aggfunc == "count":
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([primary_group, secondary_group]).size().to_dict()
                    )
                else:
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([primary_group, secondary_group]).size().to_dict()
                    )
                aggregate_columns.append(
                    {
                        "name": aggregate_column_name,
                        "function": aggfunc,
                        "field": aggregate_field,
                    }
                )

            for primary_value, primary_df in primary_groups:
                primary_info = self.get_display_value(
                    primary_value, primary_group, model_class
                )
                primary_composite = primary_info["composite_key"]
                group_data = {
                    "primary_group": primary_composite,
                    "primary_group_display": primary_info["display"],
                    "primary_group_id": primary_info["id"],
                    "items": [],
                    "subtotal": 0,
                }

                # Group by secondary group within primary group
                secondary_groups = primary_df.groupby(secondary_group)
                for secondary_value, secondary_df in secondary_groups:
                    secondary_info = self.get_display_value(
                        secondary_value, secondary_group, model_class
                    )
                    secondary_composite = secondary_info["composite_key"]
                    count_value = len(secondary_df)
                    item_data = {
                        "secondary_group": secondary_composite,
                        "secondary_group_display": secondary_info["display"],
                        "secondary_group_id": secondary_info["id"],
                        "values": {"Count": count_value},
                        "total": count_value,
                    }
                    for agg in aggregate_columns:
                        key = (primary_value, secondary_value)
                        item_data["values"][agg["name"]] = aggregate_data[
                            agg["name"]
                        ].get(key, 0)
                    group_data["items"].append(item_data)
                    group_data["subtotal"] += count_value

                hierarchical_data.append(group_data)
                grand_total += group_data["subtotal"]

            context["hierarchical_data"] = {
                "groups": hierarchical_data,
                "grand_total": grand_total,
            }
            context["pivot_columns"] = pivot_columns
            context["aggregate_columns"] = aggregate_columns

        except Exception as e:
            context["error"] = f"Error in 2x0 configuration: {str(e)}"
            context["aggregate_columns"] = []

    def handle_2_row_1_col(self, df, report, context):
        try:
            if df.empty:
                context["hierarchical_data"] = {"groups": [], "grand_total": 0}
                context["pivot_columns"] = []
                context["aggregate_columns"] = []
                return

            model_class = report.model_class
            primary_group = report.row_groups_list[0]
            secondary_group = report.row_groups_list[1]
            col_field = report.column_groups_list[0]

            # Get unique column values for headers
            unique_cols = df[col_field].unique().tolist()
            display_cols = []
            col_mapping = {}
            for col in unique_cols:
                col_info = self.get_display_value(col, col_field, model_class)
                col_composite = col_info["composite_key"]
                display_cols.append(col_composite)
                col_mapping[col] = col_composite

            # Compute aggregate columns
            aggregate_columns = []
            aggregate_data = {}
            for agg in report.aggregate_columns_dict:
                aggregate_field = agg["field"]
                aggfunc = agg.get("aggfunc", "sum")
                aggregate_column_name = f"{aggfunc.title()} of {self.get_verbose_name(aggregate_field, model_class)}"
                display_cols.append(aggregate_column_name)
                if aggfunc == "sum":
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([primary_group, secondary_group])[aggregate_field]
                        .sum()
                        .to_dict()
                    )
                elif aggfunc == "avg":
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([primary_group, secondary_group])[aggregate_field]
                        .mean()
                        .to_dict()
                    )
                elif aggfunc == "min":
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([primary_group, secondary_group])[aggregate_field]
                        .min()
                        .to_dict()
                    )
                elif aggfunc == "max":
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([primary_group, secondary_group])[aggregate_field]
                        .max()
                        .to_dict()
                    )
                elif aggfunc == "count":
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([primary_group, secondary_group]).size().to_dict()
                    )
                else:
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([primary_group, secondary_group]).size().to_dict()
                    )
                aggregate_columns.append(
                    {
                        "name": aggregate_column_name,
                        "function": aggfunc,
                        "field": aggregate_field,
                    }
                )

            hierarchical_data = []
            primary_groups = df.groupby(primary_group)
            grand_total = 0

            for primary_value, primary_df in primary_groups:
                primary_info = self.get_display_value(
                    primary_value, primary_group, model_class
                )
                primary_composite = primary_info["composite_key"]
                group_data = {
                    "primary_group": primary_composite,
                    "primary_group_display": primary_info["display"],
                    "primary_group_id": primary_info["id"],
                    "items": [],
                    "subtotal": 0,
                }

                secondary_groups = primary_df.groupby(secondary_group)
                for secondary_value, secondary_df in secondary_groups:
                    secondary_info = self.get_display_value(
                        secondary_value, secondary_group, model_class
                    )
                    secondary_composite = secondary_info["composite_key"]
                    item_data = {
                        "secondary_group": secondary_composite,
                        "secondary_group_display": secondary_info["display"],
                        "secondary_group_id": secondary_info["id"],
                        "values": {},
                        "total": 0,
                    }

                    # Compute counts for column groups
                    for col_value in unique_cols:
                        col_composite = col_mapping[col_value]
                        filtered_df = secondary_df[secondary_df[col_field] == col_value]
                        value = len(filtered_df)
                        item_data["values"][col_composite] = value
                        item_data["total"] += value

                    # Add aggregate values
                    for agg in aggregate_columns:
                        key = (primary_value, secondary_value)
                        item_data["values"][agg["name"]] = aggregate_data[
                            agg["name"]
                        ].get(key, 0)

                    group_data["items"].append(item_data)
                    group_data["subtotal"] += item_data["total"]

                hierarchical_data.append(group_data)
                grand_total += group_data["subtotal"]

            context["hierarchical_data"] = {
                "groups": hierarchical_data,
                "grand_total": grand_total,
            }
            context["pivot_columns"] = display_cols
            context["aggregate_columns"] = aggregate_columns

        except Exception as e:
            context["error"] = f"Error in 2x1 configuration: {str(e)}"
            context["aggregate_columns"] = []

    def handle_3_row_0_col(self, df, report, context):
        try:
            if df.empty:
                context["three_level_data"] = {"groups": [], "grand_total": 0}
                context["aggregate_columns"] = []
                return

            model_class = report.model_class
            level1_field = report.row_groups_list[0]
            level2_field = report.row_groups_list[1]
            level3_field = report.row_groups_list[2]

            three_level_data = []
            grand_total = 0

            # Compute aggregate columns
            aggregate_columns = []
            aggregate_data = {}
            for agg in report.aggregate_columns_dict:
                aggregate_field = agg["field"]
                aggfunc = agg.get("aggfunc", "sum")
                aggregate_column_name = f"{aggfunc.title()} of {self.get_verbose_name(aggregate_field, model_class)}"
                if aggfunc == "sum":
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([level1_field, level2_field, level3_field])[
                            aggregate_field
                        ]
                        .sum()
                        .to_dict()
                    )
                elif aggfunc == "avg":
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([level1_field, level2_field, level3_field])[
                            aggregate_field
                        ]
                        .mean()
                        .to_dict()
                    )
                elif aggfunc == "min":
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([level1_field, level2_field, level3_field])[
                            aggregate_field
                        ]
                        .min()
                        .to_dict()
                    )
                elif aggfunc == "max":
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([level1_field, level2_field, level3_field])[
                            aggregate_field
                        ]
                        .max()
                        .to_dict()
                    )
                elif aggfunc == "count":
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([level1_field, level2_field, level3_field])
                        .size()
                        .to_dict()
                    )
                else:
                    aggregate_data[aggregate_column_name] = (
                        df.groupby([level1_field, level2_field, level3_field])
                        .size()
                        .to_dict()
                    )
                aggregate_columns.append(
                    {
                        "name": aggregate_column_name,
                        "function": aggfunc,
                        "field": aggregate_field,
                    }
                )

            level1_groups = df.groupby(level1_field)
            for level1_value, level1_df in level1_groups:
                level1_info = self.get_display_value(
                    level1_value, level1_field, model_class
                )
                level1_composite = level1_info["composite_key"]
                level1_data = {
                    "level1_group": level1_composite,
                    "level1_group_display": level1_info["display"],
                    "level1_group_id": level1_info["id"],
                    "level2_groups": [],
                    "level1_total": 0,
                }

                level2_groups = level1_df.groupby(level2_field)
                for level2_value, level2_df in level2_groups:
                    level2_info = self.get_display_value(
                        level2_value, level2_field, model_class
                    )
                    level2_composite = level2_info["composite_key"]
                    level2_data = {
                        "level2_group": level2_composite,
                        "level2_group_display": level2_info["display"],
                        "level2_group_id": level2_info["id"],
                        "level3_items": [],
                        "level2_total": 0,
                    }

                    level3_groups = level2_df.groupby(level3_field)
                    for level3_value, level3_df in level3_groups:
                        level3_info = self.get_display_value(
                            level3_value, level3_field, model_class
                        )
                        level3_composite = level3_info["composite_key"]
                        count_value = len(level3_df)
                        aggregate_values = {
                            agg["name"]: aggregate_data[agg["name"]].get(
                                (level1_value, level2_value, level3_value), 0
                            )
                            for agg in aggregate_columns
                        }

                        level3_item = {
                            "level3_group": level3_composite,
                            "level3_group_display": level3_info["display"],
                            "level3_group_id": level3_info["id"],
                            "count": count_value,
                            "aggregate_values": aggregate_values,
                        }

                        level2_data["level3_items"].append(level3_item)
                        level2_data["level2_total"] += count_value

                    level1_data["level2_groups"].append(level2_data)
                    level1_data["level1_total"] += level2_data["level2_total"]

                three_level_data.append(level1_data)
                grand_total += level1_data["level1_total"]

            context["three_level_data"] = {
                "groups": three_level_data,
                "grand_total": grand_total,
            }
            context["aggregate_columns"] = aggregate_columns

        except Exception as e:
            context["error"] = f"Error in 3x0 configuration: {str(e)}"
            context["aggregate_columns"] = []

    def get_display_value(self, value, field_name, model_class):
        try:
            field = model_class._meta.get_field(field_name)
            if hasattr(field, "related_model") and field.related_model:
                try:
                    related_obj = field.related_model.objects.get(pk=value)
                    return {
                        "display": str(related_obj),
                        "id": related_obj.pk,
                        "composite_key": f"{str(related_obj)}||{related_obj.pk}",
                    }
                except field.related_model.DoesNotExist:
                    return {
                        "display": f"Unknown ({value})",
                        "id": value,
                        "composite_key": f"Unknown ({value})",
                    }
            if hasattr(field, "choices") and field.choices:
                choice_dict = dict(field.choices)
                display = choice_dict.get(value, value)
                return {"display": display, "id": value, "composite_key": str(display)}
            if value is None or value == "":
                return {
                    "display": "Unspecified (-)",
                    "id": None,
                    "composite_key": "Unspecified (-)",
                }
            return {"display": str(value), "id": value, "composite_key": str(value)}
        except:
            return {
                "display": str(value) if value is not None else "Unspecified (-)",
                "id": value,
                "composite_key": str(value) if value is not None else "Unspecified (-)",
            }

    def generate_chart_data(self, df, report):
        chart_data = {
            "labels": [],
            "data": [],
            "type": report.chart_type,
            "label_field": "Count",
            "stacked_data": {},
            "has_multiple_groups": False,
            "urls": [],
        }

        if df.empty:
            return chart_data

        config_type = self.get_configuration_type(report)
        model_class = report.model_class
        section_info = get_section_info_for_model(model_class)

        # Check if we have multiple grouping fields for stacked charts
        total_groups = len(report.row_groups_list) + len(report.column_groups_list)
        chart_data["has_multiple_groups"] = total_groups >= 2

        try:
            if config_type == "0_row_0_col":
                chart_data["labels"] = ["Records"]
                chart_data["data"] = [len(df)]
                chart_data["label_field"] = "Records"
                chart_data["urls"] = [section_info["url"]]

            elif (
                report.chart_type in ["stacked_vertical", "stacked_horizontal"]
                and chart_data["has_multiple_groups"]
            ):
                # Handle stacked charts with multiple grouping fields
                chart_data.update(
                    self._generate_stacked_chart_data(df, report, model_class)
                )

            else:
                # Handle single-dimension charts
                chart_field = None

                if (
                    hasattr(report, "chart_field")
                    and report.chart_field
                    and report.chart_field in df.columns
                ):
                    chart_field = report.chart_field
                elif report.row_groups_list and report.row_groups_list[0] in df.columns:
                    chart_field = report.row_groups_list[0]
                    if not report.chart_field:
                        report.chart_field = chart_field
                        report.save(update_fields=["chart_field"])
                elif (
                    report.column_groups_list
                    and report.column_groups_list[0] in df.columns
                ):
                    chart_field = report.column_groups_list[0]
                    if not report.chart_field:
                        report.chart_field = chart_field
                        report.save(update_fields=["chart_field"])

                if chart_field:
                    grouped = df.groupby(chart_field).size()

                    # Create unique labels with counter for duplicates
                    display_labels = []
                    display_count = {}

                    for k in grouped.index:
                        display_info = self.get_display_value(
                            k, chart_field, model_class
                        )
                        if isinstance(display_info, dict):
                            base_display = display_info["display"]
                        else:
                            base_display = str(display_info)

                        # Track duplicates and add counter
                        if base_display in display_count:
                            display_count[base_display] += 1
                            unique_label = (
                                f"{base_display} ({display_count[base_display]})"
                            )
                        else:
                            display_count[base_display] = 1
                            unique_label = base_display

                        display_labels.append(unique_label)

                    chart_data["labels"] = display_labels
                    chart_data["data"] = [float(v) for v in grouped.values]
                    chart_data["label_field"] = self.get_verbose_name(
                        chart_field, model_class
                    )
                    urls = []
                    for value in grouped.index:
                        query = urlencode(
                            {
                                "section": section_info["section"],
                                "apply_filter": "true",
                                "field": chart_field,
                                "operator": "exact",
                                "value": value if value is not None else "",
                            }
                        )
                        urls.append(f"{section_info['url']}?{query}")
                    chart_data["urls"] = urls
                else:
                    chart_data["labels"] = ["Records"]
                    chart_data["data"] = [len(df)]
                    chart_data["label_field"] = "Records"
                    chart_data["urls"] = [section_info["url"]]

        except Exception as e:
            chart_data["error"] = f"Error generating chart data: {str(e)}"

        return chart_data

    def _generate_stacked_chart_data(self, df, report, model_class):
        """Generate data for stacked charts when multiple grouping fields are available"""

        try:
            # Determine fields for stacking with priority to user-selected fields
            primary_field = None
            secondary_field = None

            # Priority 1: Use explicitly set chart fields for stacked charts
            if (
                hasattr(report, "chart_field")
                and report.chart_field
                and report.chart_field in df.columns
            ):
                primary_field = report.chart_field

                if (
                    hasattr(report, "chart_field_stacked")
                    and report.chart_field_stacked
                    and report.chart_field_stacked in df.columns
                    and report.chart_field_stacked != primary_field
                ):
                    secondary_field = report.chart_field_stacked

            # Priority 2: If chart_field_stacked is set but chart_field is not
            elif (
                hasattr(report, "chart_field_stacked")
                and report.chart_field_stacked
                and report.chart_field_stacked in df.columns
            ):
                secondary_field = report.chart_field_stacked
                all_fields = report.row_groups_list + report.column_groups_list
                primary_field = next(
                    (f for f in all_fields if f != secondary_field and f in df.columns),
                    None,
                )

            # Priority 3: Fallback to existing logic if no explicit fields are set
            if not primary_field or not secondary_field:
                if report.row_groups_list and report.column_groups_list:
                    if not primary_field:
                        primary_field = report.row_groups_list[0]
                    if not secondary_field:
                        secondary_field = report.column_groups_list[0]
                elif len(report.row_groups_list) >= 2:
                    if not primary_field:
                        primary_field = report.row_groups_list[0]
                    if not secondary_field:
                        secondary_field = report.row_groups_list[1]
                elif len(report.column_groups_list) >= 2:
                    if not primary_field:
                        primary_field = report.column_groups_list[0]
                    if not secondary_field:
                        secondary_field = report.column_groups_list[1]

            if not primary_field or not secondary_field:
                return self._fallback_chart_data(df, report, model_class)

            if primary_field not in df.columns or secondary_field not in df.columns:
                return self._fallback_chart_data(df, report, model_class)

            # Save chart fields if not already set
            fields_to_update = []
            if not report.chart_field:
                report.chart_field = primary_field
                fields_to_update.append("chart_field")
            if not report.chart_field_stacked:
                report.chart_field_stacked = secondary_field
                fields_to_update.append("chart_field_stacked")
            if fields_to_update:
                report.save(update_fields=fields_to_update)

            # Create pivot table for stacked data
            try:
                pivot_table = pd.pivot_table(
                    df,
                    index=[primary_field],
                    columns=[secondary_field],
                    aggfunc="size",
                    fill_value=0,
                )
            except Exception as pivot_error:
                return self._fallback_chart_data(df, report, model_class)

            if pivot_table.empty:
                return self._fallback_chart_data(df, report, model_class)

            # Prepare categories (x-axis labels) with unique names for duplicates
            categories = []
            category_count = {}

            for idx in pivot_table.index:
                display_info = self.get_display_value(idx, primary_field, model_class)
                if isinstance(display_info, dict):
                    base_display = display_info["display"]
                else:
                    base_display = str(display_info)

                # Track duplicates and add counter
                if base_display in category_count:
                    category_count[base_display] += 1
                    unique_label = f"{base_display} ({category_count[base_display]})"
                else:
                    category_count[base_display] = 1
                    unique_label = base_display

                categories.append(unique_label)

            # Prepare series data (stacked segments) with unique names for duplicates
            series = []
            series_name_count = {}

            for col in pivot_table.columns:
                col_display_info = self.get_display_value(
                    col, secondary_field, model_class
                )
                if isinstance(col_display_info, dict):
                    base_col_display = col_display_info["display"]
                else:
                    base_col_display = str(col_display_info)

                # Track duplicates and add counter
                if base_col_display in series_name_count:
                    series_name_count[base_col_display] += 1
                    col_display = (
                        f"{base_col_display} ({series_name_count[base_col_display]})"
                    )
                else:
                    series_name_count[base_col_display] = 1
                    col_display = base_col_display

                series_data = []

                for idx in pivot_table.index:
                    try:
                        value = pivot_table.loc[idx, col]
                        series_data.append(int(value) if pd.notna(value) else 0)
                    except Exception as val_error:
                        logger.error(
                            f"Value extraction error for {idx}, {col}: {str(val_error)}"
                        )
                        series_data.append(0)

                series.append({"name": col_display, "data": series_data})

            # Calculate totals for each category
            totals = []
            for i in range(len(categories)):
                total = sum(s["data"][i] for s in series if i < len(s["data"]))
                totals.append(total)

            section_info = get_section_info_for_model(model_class)
            urls = []
            for idx in pivot_table.index:
                query = urlencode(
                    {
                        "section": section_info["section"],
                        "apply_filter": "true",
                        "field": primary_field,
                        "operator": "exact",
                        "value": idx if idx is not None else "",
                    }
                )
                urls.append(f"{section_info['url']}?{query}")

            stacked_data = {"categories": categories, "series": series}

            primary_verbose = self.get_verbose_name(primary_field, model_class)
            secondary_verbose = self.get_verbose_name(secondary_field, model_class)

            return {
                "labels": categories,
                "data": totals,
                "urls": urls,
                "stacked_data": stacked_data,
                "label_field": f"{primary_verbose} by {secondary_verbose}",
                "has_stacked_data": True,
                "primary_field": primary_field,
                "secondary_field": secondary_field,
            }

        except Exception as e:
            logger.error(f"Error in stacked chart generation: {str(e)}")
            import traceback

            traceback.print_exc()
            return self._fallback_chart_data(df, report, model_class)

    def _fallback_chart_data(self, df, report, model_class):
        """Fallback to simple chart when stacking fails"""

        fallback_field = None
        if (
            hasattr(report, "chart_field")
            and report.chart_field
            and report.chart_field in df.columns
        ):
            fallback_field = report.chart_field
        elif report.row_groups_list and report.row_groups_list[0] in df.columns:
            fallback_field = report.row_groups_list[0]
        elif report.column_groups_list and report.column_groups_list[0] in df.columns:
            fallback_field = report.column_groups_list[0]

        section_info = get_section_info_for_model(model_class)

        if fallback_field:
            try:
                grouped = df.groupby(fallback_field).size()

                # Create unique labels with counter for duplicates
                display_labels = []
                display_count = {}

                for k in grouped.index:
                    display_info = self.get_display_value(
                        k, fallback_field, model_class
                    )
                    if isinstance(display_info, dict):
                        base_display = display_info["display"]
                    else:
                        base_display = str(display_info)

                    # Track duplicates and add counter
                    if base_display in display_count:
                        display_count[base_display] += 1
                        unique_label = f"{base_display} ({display_count[base_display]})"
                    else:
                        display_count[base_display] = 1
                        unique_label = base_display

                    display_labels.append(unique_label)

                urls = []
                for value in grouped.index:
                    query = urlencode(
                        {
                            "section": section_info["section"],
                            "apply_filter": "true",
                            "field": fallback_field,
                            "operator": "exact",
                            "value": value if value is not None else "",
                        }
                    )
                    urls.append(f"{section_info['url']}?{query}")

                return {
                    "labels": display_labels,
                    "data": [float(v) for v in grouped.values],
                    "urls": urls,
                    "stacked_data": {},
                    "label_field": self.get_verbose_name(fallback_field, model_class),
                    "has_stacked_data": False,
                }
            except Exception as e:
                logger.error(f"Fallback chart error: {str(e)}")

        # Ultimate fallback
        return {
            "labels": ["Records"],
            "data": [len(df)],
            "urls": [section_info["url"]],
            "stacked_data": {},
            "label_field": "Records",
            "has_stacked_data": False,
        }


@method_decorator(htmx_required, name="dispatch")
@method_decorator(
    permission_required_or_denied(
        ["horilla_reports.view_report", "horilla_reports.view_own_report"]
    ),
    name="dispatch",
)
class ReportDetailFilteredView(LoginRequiredMixin, View):

    def col_attrs(self):
        """Define column attributes for clickable rows in the report list view."""
        query_params = {}
        pk = self.kwargs.get("pk")
        report = Report.objects.get(pk=pk)

        model_class = report.model_class
        section = get_section_info_for_model(model_class)
        section_value = section["section"]
        query_params["section"] = section_value
        query_string = urlencode(query_params)
        attrs = {}

        if self.request.user.has_perm("horilla_reports.view_report"):
            attrs = {
                "hx-get": f"{{get_detail_url}}?{query_string}",
                "hx-target": "#mainContent",
                "hx-swap": "outerHTML",
                "hx-push-url": "true",
                "hx-on:click": "closeContentModal()",
                "hx-select": "#mainContent",
                "style": "cursor:pointer",
                "class": "hover:text-primary-600",
            }

        columns_with_attrs = []

        for col in report.selected_columns_list:
            columns_with_attrs.append({col: {**attrs}})

        return columns_with_attrs

    def get(self, request, pk, *args, **kwargs):
        # Get the report
        try:
            report = Report.objects.get(pk=pk)
        except Report.DoesNotExist:
            return render(request, "list_view.html")

        # Check if we have preview data in session
        session_key = f"report_preview_{report.pk}"
        preview_data = request.session.get(session_key, {})

        # Create a temporary report object with preview data
        temp_report = self.create_temp_report(report, preview_data)

        model_class = temp_report.model_class
        queryset = model_class.objects.all()

        # Apply original report filters using temp_report
        filters = temp_report.filters_dict
        if filters:
            try:
                # Use the same filter logic as ReportDetailView
                query = None
                for index, (field_name, filter_data) in enumerate(filters.items()):
                    if not filter_data.get("value"):
                        continue  # Skip empty filters
                    operator = filter_data.get("operator", "exact")
                    value = filter_data.get("value")
                    logic = (
                        filter_data.get("logic", "and") if index > 0 else "and"
                    )  # Default to AND for first filter

                    # Use original_field instead of field_name
                    actual_field = filter_data.get("original_field", field_name)

                    # Construct filter kwargs
                    filter_kwargs = {}
                    if operator == "exact":
                        filter_kwargs[actual_field] = value
                    elif operator == "icontains":
                        filter_kwargs[f"{actual_field}__icontains"] = value
                    elif operator == "gt":
                        filter_kwargs[f"{actual_field}__gt"] = value
                    elif operator == "lt":
                        filter_kwargs[f"{actual_field}__lt"] = value
                    elif operator == "gte":
                        filter_kwargs[f"{actual_field}__gte"] = value
                    elif operator == "lte":
                        filter_kwargs[f"{actual_field}__lte"] = value

                    # Combine filters with AND or OR
                    if not filter_kwargs:
                        continue
                    current_query = Q(**filter_kwargs)

                    if query is None:
                        query = current_query
                    elif logic == "or":
                        query |= current_query
                    else:  # logic == 'and'
                        query &= current_query

                if query:
                    queryset = queryset.filter(query)
            except Exception as e:
                logger.error(f"Filter Error in ReportDetailFilteredView: {e}")
                queryset = model_class.objects.none()

        row_group1 = request.GET.get("row_group1")
        row_group2 = request.GET.get("row_group2")
        row_group3 = request.GET.get("row_group3")
        col = request.GET.get("col")
        col1 = request.GET.get("col1")
        col2 = request.GET.get("col2")
        simple_aggregate = request.GET.get("simple_aggregate")

        # Use temp_report instead of report for row_groups and column_groups
        row_fields = temp_report.row_groups_list
        col_fields = temp_report.column_groups_list
        filter_kwargs = {}

        def get_dynamic_lookup_fields(related_model, original_field_name):
            """
            Dynamically determine lookup fields for a related model.
            Returns a list of field names to try for lookups.
            """
            lookup_fields = []

            # Get all fields from the related model
            for field in related_model._meta.get_fields():
                if hasattr(field, "name"):
                    field_name = field.name

                    # Prioritize common display fields
                    if field_name in ["name", "title", "display_name", "label"]:
                        lookup_fields.insert(0, field_name)  # Add to front
                    # Include string/char fields that might be used for display
                    elif hasattr(
                        field, "get_internal_type"
                    ) and field.get_internal_type() in ["CharField", "TextField"]:
                        lookup_fields.append(field_name)

            # Add the original field name as fallback
            if original_field_name not in lookup_fields:
                lookup_fields.append(original_field_name)

            # Add 'pk' and 'id' as final fallbacks
            if "pk" not in lookup_fields:
                lookup_fields.append("pk")
            if "id" not in lookup_fields:
                lookup_fields.append("id")

            return lookup_fields

        def get_filter_value(field_name, value, model):
            if not value or not field_name:
                return None
            try:
                field = model._meta.get_field(field_name)
                if isinstance(field, ForeignKey):
                    if isinstance(value, str) and "||" in value:
                        parts = value.split("||")
                        if len(parts) == 2:
                            try:
                                pk_value = int(parts[1])
                                related_model = field.related_model
                                try:
                                    return related_model.objects.get(pk=pk_value)
                                except related_model.DoesNotExist:
                                    logger.warning(
                                        f"Related object not found with pk={pk_value}"
                                    )
                                    return None
                            except (ValueError, TypeError):
                                value = parts[0]  # Use display part for lookup

                    related_model = field.related_model
                    lookup_fields = get_dynamic_lookup_fields(related_model, field_name)

                    for lookup_field in lookup_fields:
                        try:
                            related_obj = related_model.objects.get(
                                **{lookup_field: value}
                            )
                            return related_obj
                        except related_model.DoesNotExist:
                            continue
                        except AttributeError:
                            continue
                        except Exception as e:
                            logger.error(f"Error trying lookup {lookup_field}: {e}")
                            continue

                    # If exact match fails, try case-insensitive for string fields
                    for lookup_field in lookup_fields:
                        try:
                            field_obj = related_model._meta.get_field(lookup_field)
                            if hasattr(
                                field_obj, "get_internal_type"
                            ) and field_obj.get_internal_type() in [
                                "CharField",
                                "TextField",
                            ]:
                                related_obj = related_model.objects.get(
                                    **{f"{lookup_field}__iexact": value}
                                )
                                return related_obj
                        except (related_model.DoesNotExist, AttributeError, Exception):
                            continue
                    return None
                elif field.choices:
                    # Check if value is a composite key for choices
                    if isinstance(value, str) and "||" in value:
                        parts = value.split("||")
                        value = parts[0]  # Use display part

                    choice_map = {
                        display.lower(): value for value, display in field.choices
                    }
                    normalized_value = value.lower()
                    if normalized_value in choice_map:
                        return choice_map[normalized_value]
                    return value  # Fallback to original value if no match
                else:
                    # For non-FK fields, check if it's a composite key and extract the value
                    if isinstance(value, str) and "||" in value:
                        parts = value.split("||")
                        # For non-FK fields, try to convert to appropriate type
                        try:
                            if hasattr(field, "get_internal_type"):
                                field_type = field.get_internal_type()
                                if field_type in [
                                    "IntegerField",
                                    "BigIntegerField",
                                    "SmallIntegerField",
                                ]:
                                    return int(parts[1])
                                elif field_type in ["FloatField", "DecimalField"]:
                                    return float(parts[1])
                        except (ValueError, TypeError, IndexError):
                            pass
                        # If conversion fails, use the display part
                        value = parts[0]
                    return value
            except Exception as e:
                logger.error(
                    f"Error resolving filter value for {field_name}={value}: {e}"
                )
                return None

        if row_group1 and row_fields:
            filter_value = get_filter_value(row_fields[0], row_group1, model_class)
            if filter_value is not None:
                filter_kwargs[row_fields[0]] = filter_value
        if row_group2 and len(row_fields) > 1:
            filter_value = get_filter_value(row_fields[1], row_group2, model_class)
            if filter_value is not None:
                filter_kwargs[row_fields[1]] = filter_value
        if row_group3 and len(row_fields) > 2:
            filter_value = get_filter_value(row_fields[2], row_group3, model_class)
            if filter_value is not None:
                filter_kwargs[row_fields[2]] = filter_value
        if col and col_fields:
            filter_value = get_filter_value(col_fields[0], col, model_class)
            if filter_value is not None:
                filter_kwargs[col_fields[0]] = filter_value
        if col1 and col2 and len(col_fields) > 1:
            filter_value1 = get_filter_value(col_fields[0], col1, model_class)
            filter_value2 = get_filter_value(col_fields[1], col2, model_class)
            if filter_value1 is not None:
                filter_kwargs[col_fields[0]] = filter_value1
            if filter_value2 is not None:
                filter_kwargs[col_fields[1]] = filter_value2
        # if simple_aggregate and temp_report.aggregate_columns_dict.get('field'):
        #     filter_kwargs[temp_report.aggregate_columns_dict['field']] = simple_aggregate

        if filter_kwargs:
            try:
                queryset = queryset.filter(**filter_kwargs)
            except Exception as e:
                logger.error(f"Filter Error: {e}")
                queryset = model_class.objects.none()

        columns = []
        for col in temp_report.selected_columns_list:
            field = model_class._meta.get_field(col)
            verbose_name = field.verbose_name.title()
            if field.choices:
                columns.append((verbose_name, f"get_{col}_display"))
            else:
                columns.append((verbose_name, col))

        list_view = HorillaListView(
            model=model_class,
            view_id="report-details",
            search_url=reverse_lazy(
                "horilla_reports:report_detail_filtered", kwargs={"pk": report.pk}
            ),
            main_url=reverse_lazy(
                "horilla_reports:report_detail_filtered", kwargs={"pk": report.pk}
            ),
            table_width=False,
            columns=columns,
        )
        list_view.request = request
        list_view.table_width = False
        list_view.bulk_select_option = False
        list_view.clear_session_button_enabled = False
        list_view.list_column_visibility = False
        list_view.table_height = False
        list_view.table_height_as_class = "h-[200px]"
        if hasattr(report.model_class, "get_detail_url"):
            list_view.col_attrs = self.col_attrs()
        sort_field = self.request.GET.get("sort")
        sort_direction = self.request.GET.get("direction", "asc")
        if sort_field:
            queryset = list_view._apply_sorting(queryset, sort_field, sort_direction)
        else:
            queryset = queryset.order_by("-id")
        list_view.queryset = queryset  # Set object_list to avoid AttributeError
        context = list_view.get_context_data(object_list=queryset)

        # Add no_record_msg if queryset is empty
        if not queryset.exists():
            context["no_record_msg"] = "No records found"

        # Render only the list_view.html template
        return render(request, "list_view.html", context)

    def create_temp_report(self, original_report, preview_data):
        """Create a temporary report object with preview data (same as ReportDetailView)"""
        import copy

        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        if "chart_type" in preview_data:
            temp_report.chart_type = preview_data["chart_type"]
        if "chart_field" in preview_data:
            temp_report.chart_field = preview_data["chart_field"]
        if "chart_field_stacked" in preview_data:
            temp_report.chart_field_stacked = preview_data["chart_field_stacked"]

        return temp_report


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class ToggleAggregateView(LoginRequiredMixin, View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        field_name = request.POST.get("field_name")
        session_key = f"report_preview_{report.pk}"

        preview_data = request.session.get(session_key, {})
        current_aggregate = preview_data.get(
            "aggregate_columns", report.aggregate_columns
        )
        try:
            aggregate_list = json.loads(current_aggregate) if current_aggregate else []
            if not isinstance(aggregate_list, list):
                aggregate_list = [aggregate_list] if aggregate_list else []
        except (json.JSONDecodeError, TypeError):
            aggregate_list = []

        # Check how many times this field already appears
        field_count = sum(1 for agg in aggregate_list if agg.get("field") == field_name)

        # Define the aggregation functions in order
        aggfunc_order = ["sum", "avg", "count", "max", "min"]

        if field_count < len(aggfunc_order):
            # Add the next aggregation function for this field
            next_aggfunc = aggfunc_order[field_count]
            aggregate_list.append({"field": field_name, "aggfunc": next_aggfunc})

            preview_data["aggregate_columns"] = json.dumps(aggregate_list)
            request.session[session_key] = preview_data

        temp_report = self.create_temp_report(report, preview_data)
        detail_view = ReportDetailView()
        detail_view.request = self.request
        detail_view.object = temp_report
        detail_view.kwargs = {"pk": report.pk}
        context = detail_view.get_context_data()
        return render(request, "report_detail.html", context)

    def create_temp_report(self, original_report, preview_data):
        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        if "chart_type" in preview_data:
            temp_report.chart_type = preview_data["chart_type"]
        if "chart_field" in preview_data:
            temp_report.chart_field = preview_data["chart_field"]
        if "chart_field_stacked" in preview_data:
            temp_report.chart_field_stacked = preview_data["chart_field_stacked"]
        return temp_report


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class UpdateAggregateFunctionView(LoginRequiredMixin, View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        aggfunc = request.POST.get("aggfunc")
        field_name = request.POST.get("field_name")
        session_key = f"report_preview_{report.pk}"

        preview_data = request.session.get(session_key, {})
        current_aggregate = preview_data.get(
            "aggregate_columns", report.aggregate_columns
        )
        try:
            aggregate_list = json.loads(current_aggregate) if current_aggregate else []
            if not isinstance(aggregate_list, list):
                aggregate_list = [aggregate_list] if aggregate_list else []
        except (json.JSONDecodeError, TypeError):
            aggregate_list = []

        # Update the aggregation function for the specified field
        for agg in aggregate_list:
            if agg.get("field") == field_name:
                agg["aggfunc"] = aggfunc

        preview_data["aggregate_columns"] = json.dumps(aggregate_list)
        request.session[session_key] = preview_data

        temp_report = self.create_temp_report(report, preview_data)
        detail_view = ReportDetailView()
        detail_view.request = self.request
        detail_view.object = temp_report
        detail_view.kwargs = {"pk": report.pk}
        context = detail_view.get_context_data()
        return render(request, "report_detail.html", context)

    def create_temp_report(self, original_report, preview_data):
        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        if "chart_type" in preview_data:
            temp_report.chart_type = preview_data["chart_type"]
        if "chart_field" in preview_data:
            temp_report.chart_field = preview_data["chart_field"]
        if "chart_field_stacked" in preview_data:
            temp_report.chart_field_stacked = preview_data["chart_field_stacked"]
        return temp_report


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class SaveReportChangesView(LoginRequiredMixin, View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        session_key = f"report_preview_{report.pk}"
        preview_data = request.session.get(session_key, {})

        if preview_data:
            # Apply all changes to the actual model
            if "selected_columns" in preview_data:
                report.selected_columns = preview_data["selected_columns"]
            if "row_groups" in preview_data:
                report.row_groups = preview_data["row_groups"]
            if "column_groups" in preview_data:
                report.column_groups = preview_data["column_groups"]
            if "aggregate_columns" in preview_data:
                report.aggregate_columns = preview_data["aggregate_columns"]
            if "filters" in preview_data:
                report.filters = preview_data["filters"]
            if "chart_type" in preview_data:
                report.chart_type = preview_data["chart_type"]
            if "chart_field" in preview_data:
                report.chart_field = preview_data["chart_field"]
            if "chart_field_stacked" in preview_data:
                report.chart_field_stacked = preview_data["chart_field_stacked"]
            report.save()

            # Clear the session preview data
            if session_key in request.session:
                del request.session[session_key]

        # Use ReportDetailView to get the full context
        detail_view = ReportDetailView()
        detail_view.request = request
        detail_view.object = report
        context = detail_view.get_context_data()

        # Ensure panel is closed and no unsaved changes
        context["panel_open"] = False
        context["has_unsaved_changes"] = False

        # Render the report_detail.html template with the full context
        return render(request, "report_detail.html", context)


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class DiscardReportChangesView(LoginRequiredMixin, View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        session_key = f"report_preview_{pk}"

        # Clear the session preview data
        if session_key in request.session:
            del request.session[session_key]

        # Use ReportDetailView to get the full context
        detail_view = ReportDetailView()
        detail_view.request = request
        detail_view.object = report
        context = detail_view.get_context_data()

        # Ensure panel is closed and no unsaved changes
        context["panel_open"] = False
        context["has_unsaved_changes"] = False

        # Render the report_detail.html template with the full context
        return render(request, "report_detail.html", context)


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class ReportUpdateView(LoginRequiredMixin, DetailView):
    model = Report
    template_name = "partials/report_panel.html"
    context_object_name = "report"

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect_to_login(request.get_full_path())
        try:
            self.object = self.get_object()
        except Exception as e:
            if request.headers.get("HX-Request") == "true":
                messages.error(self.request, e)
                return HttpResponse(headers={"HX-Refresh": "true"})
            raise HorillaHttp404(e)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        report = self.object
        model_class = report.model_class

        # Get preview data for panel display
        session_key = f"report_preview_{report.pk}"
        preview_data = self.request.session.get(session_key, {})

        # Get the active tab from request (for maintaining state)
        active_tab = self.request.GET.get("active_tab", "columns")
        context["active_tab"] = active_tab

        temp_report = self.create_temp_report(report, preview_data)
        context["report"] = temp_report
        context["has_unsaved_changes"] = bool(preview_data)
        context["panel_open"] = True

        available_fields = []
        for field in model_class._meta.get_fields():
            if not field.many_to_many and not field.one_to_many:
                available_fields.append(
                    {
                        "name": field.name,
                        "verbose_name": field.verbose_name,
                        "field_type": field.__class__.__name__,
                    }
                )

        context["available_fields"] = available_fields
        return context

    def create_temp_report(self, original_report, preview_data):
        """Create a temporary report object with preview data"""
        import copy

        temp_report = copy.copy(original_report)

        # Apply preview changes
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]

        return temp_report


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class AddColumnView(LoginRequiredMixin, View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        field_name = request.POST.get("field_name")
        session_key = f"report_preview_{report.pk}"

        # Get current preview data or start from original
        preview_data = request.session.get(session_key, {})

        # Get current selected columns (from preview or original)
        current_columns = preview_data.get("selected_columns", report.selected_columns)
        selected_columns_list = (
            [col.strip() for col in current_columns.split(",") if col.strip()]
            if current_columns
            else []
        )

        if field_name and field_name not in selected_columns_list:
            selected_columns_list.append(field_name)
            preview_data["selected_columns"] = ",".join(selected_columns_list)
            request.session[session_key] = preview_data

        # Create a temporary report with updated preview data
        temp_report = self.create_temp_report(report, preview_data)

        # Instantiate ReportDetailView and set up required attributes
        detail_view = ReportDetailView()
        detail_view.request = self.request
        detail_view.object = temp_report
        detail_view.kwargs = {"pk": report.pk}

        # Get context from ReportDetailView
        context = detail_view.get_context_data()

        # Return the full report content
        return render(request, "report_detail.html", context)

    def create_temp_report(self, original_report, preview_data):
        import copy

        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        return temp_report


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class RemoveColumnView(LoginRequiredMixin, View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        field_name = request.POST.get("field_name")
        session_key = f"report_preview_{report.pk}"

        # Get current preview data or start from original
        preview_data = request.session.get(session_key, {})

        # Get current selected columns (from preview or original)
        current_columns = preview_data.get("selected_columns", report.selected_columns)
        selected_columns_list = (
            [col.strip() for col in current_columns.split(",") if col.strip()]
            if current_columns
            else []
        )

        if field_name and field_name in selected_columns_list:
            selected_columns_list.remove(field_name)
            preview_data["selected_columns"] = ",".join(selected_columns_list)
            request.session[session_key] = preview_data

        temp_report = self.create_temp_report(report, preview_data)

        detail_view = ReportDetailView()
        detail_view.request = self.request
        detail_view.object = temp_report
        detail_view.kwargs = {"pk": report.pk}

        context = detail_view.get_context_data()

        # Return the full report content
        return render(request, "report_detail.html", context)

    def create_temp_report(self, original_report, preview_data):
        import copy

        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        return temp_report


@method_decorator(
    permission_required_or_denied(
        ["horilla_reports.view_report", "horilla_reports.view_own_report"]
    ),
    name="dispatch",
)
class CloseReportPanelView(LoginRequiredMixin, View):
    def get(self, request, pk):
        """Close the report panel and redirect to detail view"""
        # Clear any session data if needed
        session_key = f"report_preview_{pk}"
        if session_key in request.session:
            pass

        return redirect("horilla_reports:report_detail", pk=pk)


@method_decorator(
    permission_required_or_denied(
        ["horilla_reports.view_report", "horilla_reports.view_own_report"]
    ),
    name="dispatch",
)
class AddFilterFieldView(LoginRequiredMixin, View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        field_name = request.POST.get("field_name")
        session_key = f"report_preview_{report.pk}"

        # Get current preview data or start from original
        preview_data = request.session.get(session_key, {})

        # Get current filters (from preview or original)
        current_filters = preview_data.get("filters", report.filters)
        try:
            filters_dict = json.loads(current_filters) if current_filters else {}
        except (json.JSONDecodeError, TypeError):
            filters_dict = {}

        # Generate a unique key for the filter
        base_field_name = field_name
        index = 1
        unique_field_name = field_name
        while unique_field_name in filters_dict:
            unique_field_name = f"{base_field_name}_{index}"
            index += 1

        # Add new filter with default logic 'and'
        filters_dict[unique_field_name] = {
            "value": "",
            "operator": "exact",
            "logic": "and",
            "original_field": base_field_name,
        }

        preview_data["filters"] = json.dumps(filters_dict)
        request.session[session_key] = preview_data

        # Determine if the field is a choice field or foreign key and get its options
        is_choice_or_fk = report.is_choice_or_foreign_key_field(field_name)
        field_choices = report.get_field_choices(field_name) if is_choice_or_fk else []

        # Create temp report for context
        temp_report = self.create_temp_report(report, preview_data)

        # Render the entire panel template with updated context
        return render(
            request,
            "partials/report_panel.html",
            {
                "report": temp_report,
                "available_fields": report.model_class._meta.get_fields(),
                "has_unsaved_changes": True,
                "is_choice_or_fk": is_choice_or_fk,
                "field_choices": field_choices,
            },
        )

    def create_temp_report(self, original_report, preview_data):
        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        return temp_report


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class UpdateFilterOperatorView(View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        field_name = request.POST.get("field_name")
        operator = request.POST.get("operator")
        session_key = f"report_preview_{report.pk}"

        # Get current preview data or start from original
        preview_data = request.session.get(session_key, {})

        # Get current filters (from preview or original)
        current_filters = preview_data.get("filters", report.filters)
        try:
            filters_dict = json.loads(current_filters) if current_filters else {}
        except (json.JSONDecodeError, TypeError):
            filters_dict = {}

        # Update operator and preserve or set default logic
        if field_name in filters_dict:
            if isinstance(filters_dict[field_name], dict):
                filters_dict[field_name]["operator"] = operator
                filters_dict[field_name].setdefault(
                    "logic", "and"
                )  # Preserve or set default logic
            else:
                filters_dict[field_name] = {
                    "value": str(filters_dict[field_name]),
                    "operator": operator,
                    "logic": "and",
                }
        else:
            filters_dict[field_name] = {
                "value": "",
                "operator": operator,
                "logic": "and",
            }

        preview_data["filters"] = json.dumps(filters_dict)
        request.session[session_key] = preview_data

        temp_report = self.create_temp_report(report, preview_data)
        detail_view = ReportDetailView()
        detail_view.request = self.request
        detail_view.object = temp_report
        detail_view.kwargs = {"pk": report.pk}
        context = detail_view.get_context_data()
        return render(request, "report_detail.html", context)

    def create_temp_report(self, original_report, preview_data):
        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        if "chart_type" in preview_data:
            temp_report.chart_type = preview_data["chart_type"]
        if "chart_field" in preview_data:
            temp_report.chart_field = preview_data["chart_field"]
        if "chart_field_stacked" in preview_data:
            temp_report.chart_field_stacked = preview_data["chart_field_stacked"]
        return temp_report

    def get_available_fields(self, model_class):
        available_fields = []
        for field in model_class._meta.get_fields():
            if not field.many_to_many and not field.one_to_many:
                available_fields.append(
                    {
                        "name": field.name,
                        "verbose_name": field.verbose_name,
                        "field_type": field.__class__.__name__,
                    }
                )
        return available_fields


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class UpdateFilterValueView(LoginRequiredMixin, View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        field_name = request.POST.get("field_name")
        value = request.POST.get("value")
        session_key = f"report_preview_{report.pk}"

        # Get current preview data or start from original
        preview_data = request.session.get(session_key, {})

        # Get current filters (from preview or original)
        current_filters = preview_data.get("filters", report.filters)
        try:
            filters_dict = json.loads(current_filters) if current_filters else {}
        except (json.JSONDecodeError, TypeError):
            filters_dict = {}

        # Update value and preserve or set default logic
        if field_name in filters_dict:
            if isinstance(filters_dict[field_name], dict):
                filters_dict[field_name]["value"] = value
                filters_dict[field_name].setdefault(
                    "logic", "and"
                )  # Preserve or set default logic
            else:
                filters_dict[field_name] = {
                    "value": value,
                    "operator": "exact",
                    "logic": "and",
                }
        else:
            filters_dict[field_name] = {
                "value": value,
                "operator": "exact",
                "logic": "and",
            }

        preview_data["filters"] = json.dumps(filters_dict)
        request.session[session_key] = preview_data

        temp_report = self.create_temp_report(report, preview_data)
        detail_view = ReportDetailView()
        detail_view.request = self.request
        detail_view.object = temp_report
        detail_view.kwargs = {"pk": report.pk}
        context = detail_view.get_context_data()
        return render(request, "report_detail.html", context)

    def create_temp_report(self, original_report, preview_data):
        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        if "chart_type" in preview_data:
            temp_report.chart_type = preview_data["chart_type"]
        if "chart_field" in preview_data:
            temp_report.chart_field = preview_data["chart_field"]
        if "chart_field_stacked" in preview_data:
            temp_report.chart_field_stacked = preview_data["chart_field_stacked"]
        return temp_report

    def get_available_fields(self, model_class):
        available_fields = []
        for field in model_class._meta.get_fields():
            if not field.many_to_many and not field.one_to_many:
                available_fields.append(
                    {
                        "name": field.name,
                        "verbose_name": field.verbose_name,
                        "field_type": field.__class__.__name__,
                    }
                )
        return available_fields


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class UpdateFilterLogicView(LoginRequiredMixin, View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        field_name = request.POST.get(
            "field_name"
        )  # This is the unique field name (e.g., field_name_1)
        logic = request.POST.get("logic")
        session_key = f"report_preview_{report.pk}"

        # Get current preview data or start from original
        preview_data = request.session.get(session_key, {})

        # Get current filters (from preview or original)
        current_filters = preview_data.get("filters", report.filters)
        try:
            filters_dict = json.loads(current_filters) if current_filters else {}
        except (json.JSONDecodeError, TypeError):
            filters_dict = {}

        # Update the logic for the specific filter
        if field_name in filters_dict:
            filters_dict[field_name]["logic"] = logic

        preview_data["filters"] = json.dumps(filters_dict)
        request.session[session_key] = preview_data

        temp_report = self.create_temp_report(report, preview_data)
        detail_view = ReportDetailView()
        detail_view.request = self.request
        detail_view.object = temp_report
        detail_view.kwargs = {"pk": report.pk}
        context = detail_view.get_context_data()
        return render(request, "report_detail.html", context)

    def create_temp_report(self, original_report, preview_data):
        import copy

        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        if "chart_type" in preview_data:
            temp_report.chart_type = preview_data["chart_type"]
        if "chart_field" in preview_data:
            temp_report.chart_field = preview_data["chart_field"]
        if "chart_field_stacked" in preview_data:
            temp_report.chart_field_stacked = preview_data["chart_field_stacked"]
        return temp_report


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class RemoveFilterView(LoginRequiredMixin, View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        field_name = request.POST.get("field_name")
        session_key = f"report_preview_{report.pk}"

        preview_data = request.session.get(session_key, {})
        current_filters = preview_data.get("filters", report.filters)
        try:
            filters_dict = json.loads(current_filters) if current_filters else {}
        except (json.JSONDecodeError, TypeError):
            filters_dict = {}

        if field_name in filters_dict:
            del filters_dict[field_name]

        preview_data["filters"] = json.dumps(filters_dict)
        request.session[session_key] = preview_data

        temp_report = self.create_temp_report(report, preview_data)
        detail_view = ReportDetailView()
        detail_view.request = self.request
        detail_view.object = temp_report
        detail_view.kwargs = {"pk": report.pk}
        context = detail_view.get_context_data()
        return render(request, "report_detail.html", context)

    def create_temp_report(self, original_report, preview_data):
        import copy

        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        if "chart_type" in preview_data:
            temp_report.chart_type = preview_data["chart_type"]
        if "chart_field" in preview_data:
            temp_report.chart_field = preview_data["chart_field"]
        if "chart_field_stacked" in preview_data:
            temp_report.chart_field_stacked = preview_data["chart_field_stacked"]
        return temp_report

    def get_available_fields(self, model_class):
        available_fields = []
        for field in model_class._meta.get_fields():
            if not field.many_to_many and not field.one_to_many:
                available_fields.append(
                    {
                        "name": field.name,
                        "verbose_name": field.verbose_name,
                        "field_type": field.__class__.__name__,
                    }
                )
        return available_fields


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class ToggleRowGroupView(LoginRequiredMixin, View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        field_name = request.POST.get("field_name")
        session_key = f"report_preview_{report.pk}"

        # Get current preview data or start from original
        preview_data = request.session.get(session_key, {})

        # Get current row groups (from preview or original)
        current_row_groups = preview_data.get("row_groups", report.row_groups)
        row_groups_list = (
            [group.strip() for group in current_row_groups.split(",") if group.strip()]
            if current_row_groups
            else []
        )

        if field_name in row_groups_list:
            row_groups_list.remove(field_name)
        else:
            row_groups_list.append(field_name)

        preview_data["row_groups"] = ",".join(row_groups_list)
        request.session[session_key] = preview_data

        # Create a temporary report with updated preview data
        temp_report = self.create_temp_report(report, preview_data)

        # Instantiate ReportDetailView and set up required attributes
        detail_view = ReportDetailView()
        detail_view.request = self.request
        detail_view.object = temp_report
        detail_view.kwargs = {"pk": report.pk}

        context = detail_view.get_context_data()

        return render(request, "report_detail.html", context)

    def create_temp_report(self, original_report, preview_data):
        import copy

        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        if "chart_type" in preview_data:
            temp_report.chart_type = preview_data["chart_type"]
        if "chart_field" in preview_data:
            temp_report.chart_field = preview_data["chart_field"]
        if "chart_field_stacked" in preview_data:
            temp_report.chart_field_stacked = preview_data["chart_field_stacked"]
        return temp_report


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class RemoveRowGroupView(LoginRequiredMixin, View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        field_name = request.POST.get("field_name")
        session_key = f"report_preview_{report.pk}"
        preview_data = request.session.get(session_key, {})
        current_row_groups = preview_data.get("row_groups", report.row_groups)
        row_groups_list = (
            [group.strip() for group in current_row_groups.split(",") if group.strip()]
            if current_row_groups
            else []
        )

        if field_name in row_groups_list:
            row_groups_list.remove(field_name)

        preview_data["row_groups"] = ",".join(row_groups_list)
        request.session[session_key] = preview_data

        # Create a temporary report with updated preview data
        temp_report = self.create_temp_report(report, preview_data)

        # Instantiate ReportDetailView and set up required attributes
        detail_view = ReportDetailView()
        detail_view.request = self.request
        detail_view.object = temp_report
        detail_view.kwargs = {"pk": report.pk}

        # Get context from ReportDetailView
        context = detail_view.get_context_data()

        # Return the full report content
        return render(request, "report_detail.html", context)

    def create_temp_report(self, original_report, preview_data):
        import copy

        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        if "chart_type" in preview_data:
            temp_report.chart_type = preview_data["chart_type"]
        if "chart_field" in preview_data:
            temp_report.chart_field = preview_data["chart_field"]
        if "chart_field_stacked" in preview_data:
            temp_report.chart_field_stacked = preview_data["chart_field_stacked"]

        return temp_report


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class ToggleColumnGroupView(LoginRequiredMixin, View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        field_name = request.POST.get("field_name")
        session_key = f"report_preview_{report.pk}"

        preview_data = request.session.get(session_key, {})

        current_column_groups = preview_data.get("column_groups", report.column_groups)
        column_groups_list = (
            [
                group.strip()
                for group in current_column_groups.split(",")
                if group.strip()
            ]
            if current_column_groups
            else []
        )

        if field_name in column_groups_list:
            column_groups_list.remove(field_name)
        else:
            column_groups_list.append(field_name)

        preview_data["column_groups"] = ",".join(column_groups_list)
        request.session[session_key] = preview_data

        temp_report = self.create_temp_report(report, preview_data)

        detail_view = ReportDetailView()
        detail_view.request = self.request
        detail_view.object = temp_report
        detail_view.kwargs = {"pk": report.pk}
        context = detail_view.get_context_data()
        return render(request, "report_detail.html", context)

    def create_temp_report(self, original_report, preview_data):
        import copy

        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        if "chart_type" in preview_data:
            temp_report.chart_type = preview_data["chart_type"]
        if "chart_field" in preview_data:
            temp_report.chart_field = preview_data["chart_field"]
        if "chart_field_stacked" in preview_data:
            temp_report.chart_field_stacked = preview_data["chart_field_stacked"]

        return temp_report


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class RemoveColumnGroupView(LoginRequiredMixin, View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        field_name = request.POST.get("field_name")
        session_key = f"report_preview_{report.pk}"

        preview_data = request.session.get(session_key, {})

        current_column_groups = preview_data.get("column_groups", report.column_groups)
        column_groups_list = (
            [
                group.strip()
                for group in current_column_groups.split(",")
                if group.strip()
            ]
            if current_column_groups
            else []
        )

        if field_name in column_groups_list:
            column_groups_list.remove(field_name)

        preview_data["column_groups"] = ",".join(column_groups_list)
        request.session[session_key] = preview_data

        temp_report = self.create_temp_report(report, preview_data)

        detail_view = ReportDetailView()
        detail_view.request = self.request
        detail_view.object = temp_report
        detail_view.kwargs = {"pk": report.pk}

        context = detail_view.get_context_data()

        return render(request, "report_detail.html", context)

    def create_temp_report(self, original_report, preview_data):
        import copy

        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        if "chart_type" in preview_data:
            temp_report.chart_type = preview_data["chart_type"]
        if "chart_field" in preview_data:
            temp_report.chart_field = preview_data["chart_field"]
        if "chart_field_stacked" in preview_data:
            temp_report.chart_field_stacked = preview_data["chart_field_stacked"]
        return temp_report


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class RemoveAggregateColumnView(LoginRequiredMixin, View):

    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        field_name = request.POST.get("field_name")
        session_key = f"report_preview_{report.pk}"
        preview_data = request.session.get(session_key, {})
        current_aggregate_columns = preview_data.get(
            "aggregate_columns", report.aggregate_columns
        )
        try:
            aggregate_list = (
                json.loads(current_aggregate_columns)
                if current_aggregate_columns
                else []
            )
            if not isinstance(aggregate_list, list):
                aggregate_list = [aggregate_list] if aggregate_list else []
        except (json.JSONDecodeError, TypeError):
            aggregate_list = []

        aggregate_list = [
            agg for agg in aggregate_list if agg.get("field") != field_name
        ]
        preview_data["aggregate_columns"] = json.dumps(aggregate_list)
        request.session[session_key] = preview_data
        request.session.modified = True
        temp_report = self.create_temp_report(report, preview_data)
        detail_view = ReportDetailView()
        detail_view.request = self.request
        detail_view.object = temp_report
        detail_view.kwargs = {"pk": report.pk}
        context = detail_view.get_context_data()
        return render(request, "report_detail.html", context)

    def create_temp_report(self, original_report, preview_data):
        import copy

        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        if "chart_type" in preview_data:
            temp_report.chart_type = preview_data["chart_type"]
        if "chart_field" in preview_data:
            temp_report.chart_field = preview_data["chart_field"]
        if "chart_field_stacked" in preview_data:
            temp_report.chart_field_stacked = preview_data["chart_field_stacked"]

        return temp_report


@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class SearchAvailableFieldsView(LoginRequiredMixin, DetailView):
    model = Report

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect_to_login(request.get_full_path())
        try:
            self.object = self.get_object()
        except Exception as e:
            if request.headers.get("HX-Request") == "true":
                messages.error(self.request, e)
                return HttpResponse(headers={"HX-Refresh": "true"})
            raise HorillaHttp404(e)
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        report = self.get_object()
        search_query = request.GET.get("search_columns", "").strip().lower()
        search_grouping = request.GET.get("search_grouping", "").strip().lower()
        search_filter = request.GET.get("search_filter", "").strip().lower()
        field_type = request.GET.get("field_type", "columns")

        # Get all available fields
        model_class = report.model_class
        available_fields = []
        for field in model_class._meta.get_fields():
            if not field.many_to_many and not field.one_to_many:
                available_fields.append(
                    {
                        "name": field.name,
                        "verbose_name": field.verbose_name,
                        "field_type": field.__class__.__name__,
                    }
                )

        # Get the appropriate search query based on field type
        search_term = ""
        if field_type == "columns":
            search_term = search_query
        elif field_type == "grouping":
            search_term = search_grouping
        elif field_type == "filter":
            search_term = search_filter

        # Filter fields based on search term
        if search_term:
            filtered_fields = [
                field
                for field in available_fields
                if search_term in field["verbose_name"].lower()
                or search_term in field["name"].lower()
            ]
        else:
            filtered_fields = available_fields

        # Get preview data for temp report
        session_key = f"report_preview_{report.pk}"
        preview_data = self.request.session.get(session_key, {})
        temp_report = self.create_temp_report(report, preview_data)

        # Render the appropriate template based on field type
        if field_type == "columns":
            html = render_to_string(
                "partials/available_columns_list.html",
                {"available_fields": filtered_fields, "report": temp_report},
            )
        elif field_type == "grouping":
            html = render_to_string(
                "partials/available_grouping_list.html",
                {"available_fields": filtered_fields, "report": temp_report},
            )
        elif field_type == "filter":
            html = render_to_string(
                "partials/available_filter_list.html",
                {"available_fields": filtered_fields, "report": temp_report},
            )
        else:
            html = "<div>Invalid field type</div>"

        return HttpResponse(html)

    def create_temp_report(self, original_report, preview_data):
        """Create a temporary report object with preview data"""
        import copy

        temp_report = copy.copy(original_report)

        # Apply preview changes
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]

        return temp_report


@method_decorator(htmx_required, name="dispatch")
@method_decorator(
    permission_required_or_denied(["reports.view_report", "reports.view_own_report"]),
    name="dispatch",
)
class ChangeChartTypeView(LoginRequiredMixin, HorillaSingleFormView):

    model = Report
    fields = ["chart_type"]
    modal_height = False
    full_width_fields = ["chart_type"]
    form_class = ChangeChartReportForm

    @cached_property
    def form_url(self):
        pk = self.kwargs.get("pk") or self.request.GET.get("id")
        if pk:
            return reverse_lazy("horilla_reports:change_chart_type", kwargs={"pk": pk})


@method_decorator(htmx_required, name="dispatch")
@method_decorator(
    permission_required_or_denied(
        ["horilla_reports.view_report", "horilla_reports.view_own_report"]
    ),
    name="dispatch",
)
class ChangeChartFieldView(LoginRequiredMixin, HorillaSingleFormView):
    model = Report
    fields = ["chart_field", "chart_field_stacked"]
    modal_height = False
    full_width_fields = ["chart_field", "chart_field_stacked"]

    def get_form_class(self):
        report = get_object_or_404(Report, pk=self.kwargs["pk"])

        # Check if we have preview data in session
        session_key = f"report_preview_{report.pk}"
        preview_data = self.request.session.get(session_key, {})

        if preview_data:
            temp_report = self.create_temp_report(report, preview_data)
        else:
            temp_report = report

        field_choices = []

        # Add row groups to choices
        for field_name in temp_report.row_groups_list:
            try:
                field = temp_report.model_class._meta.get_field(field_name)
                verbose_name = field.verbose_name.title()
                field_choices.append((field_name, f"{verbose_name} (Row Group)"))
            except:
                field_choices.append((field_name, f"{field_name.title()} (Row Group)"))

        # Add column groups to choices
        for field_name in temp_report.column_groups_list:
            try:
                field = temp_report.model_class._meta.get_field(field_name)
                verbose_name = field.verbose_name.title()
                field_choices.append((field_name, f"{verbose_name} (Column Group)"))
            except:
                field_choices.append(
                    (field_name, f"{field_name.title()} (Column Group)")
                )

        # Add empty choice for clearing the field
        field_choices.insert(0, ("", "-- Select Chart Field --"))

        class ChartFieldForm(HorillaModelForm):
            chart_field = forms.ChoiceField(
                choices=field_choices,
                label="Primary Chart Field",
                required=False,  # Allow empty selection
                widget=forms.Select(attrs={"class": "w-full p-2 border rounded"}),
            )

            chart_field_stacked = forms.ChoiceField(
                choices=field_choices,
                label="Secondary Field (For Stacked Charts)",
                required=False,  # Allow empty selection
                widget=forms.Select(attrs={"class": "w-full p-2 border rounded"}),
            )

            class Meta:
                model = Report
                fields = ["chart_field", "chart_field_stacked"]

        return ChartFieldForm

    def get_initial(self):
        """Get initial form data from preview or database"""
        report = get_object_or_404(Report, pk=self.kwargs["pk"])

        session_key = f"report_preview_{report.pk}"
        preview_data = self.request.session.get(session_key, {})

        initial = super().get_initial()

        if preview_data:
            initial["chart_field"] = preview_data.get("chart_field", "")
            initial["chart_field_stacked"] = preview_data.get("chart_field_stacked", "")
        else:
            initial["chart_field"] = report.chart_field or ""
            initial["chart_field_stacked"] = report.chart_field_stacked or ""

        return initial

    def create_temp_report(self, original_report, preview_data):
        """Create a temporary report object with preview data applied"""
        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        if "chart_type" in preview_data:
            temp_report.chart_type = preview_data["chart_type"]
        if "chart_field" in preview_data:
            temp_report.chart_field = preview_data["chart_field"]
        if "chart_field_stacked" in preview_data:
            temp_report.chart_field_stacked = preview_data["chart_field_stacked"]
        return temp_report

    def form_valid(self, form):
        report = get_object_or_404(Report, pk=self.kwargs["pk"])
        chart_field_value = form.cleaned_data.get("chart_field")
        chart_field_stacked_value = form.cleaned_data.get("chart_field_stacked")

        # Check if we have preview data in session (preview mode)
        session_key = f"report_preview_{report.pk}"
        preview_data = self.request.session.get(session_key, {})

        if preview_data:
            preview_data["chart_field"] = chart_field_value
            preview_data["chart_field_stacked"] = chart_field_stacked_value
            self.request.session[session_key] = preview_data
            self.request.session.modified = True
        else:
            report.chart_field = chart_field_value
            report.chart_field_stacked = chart_field_stacked_value
            report.save(update_fields=["chart_field", "chart_field_stacked"])

        return HttpResponse("<script>$('#reloadButton').click();closeModal();</script>")

    @cached_property
    def form_url(self):
        pk = self.kwargs.get("pk") or self.request.GET.get("id")
        if pk:
            return reverse_lazy("horilla_reports:change_chart_field", kwargs={"pk": pk})


@method_decorator(htmx_required, name="dispatch")
class CreateReportView(LoginRequiredMixin, HorillaSingleFormView):
    model = Report
    fields = ["name", "module", "folder", "selected_columns", "report_owner"]
    modal_height = False
    form_class = ReportForm
    hidden_fields = ["report_owner"]
    full_width_fields = ["name", "module", "folder", "selected_columns"]

    @cached_property
    def form_url(self):
        return reverse_lazy("horilla_reports:create_report")

    def get_initial(self):
        initial = super().get_initial()
        pk = self.request.GET.get("pk")
        initial["folder"] = pk if pk else None
        initial["report_owner"] = self.request.user
        return initial

    def form_valid(self, form):
        response = super().form_valid(form)
        report_pk = self.object.pk
        report_detail_url = reverse_lazy(
            "horilla_reports:report_detail", kwargs={"pk": report_pk}
        )
        return HttpResponse(
            f'<div id="htmx-trigger" '
            f'hx-get="{report_detail_url}" '
            'hx-target="#mainContent" '
            'hx-swap="outerHTML" '
            'hx-push-url="true" '
            'hx-select="#mainContent" '
            'hx-trigger="load" '
            'hx-on::after-request="closeModal();">'
            "</div>"
        )

    def form_invalid(self, form):
        module_id = self.request.POST.get("module") or (
            form.instance.module.id if form.instance.module else None
        )
        selected_values = self.request.POST.getlist("selected_columns") or (
            form.instance.selected_columns.split(",")
            if form.instance.selected_columns
            else []
        )
        choices = []
        if module_id:
            try:
                content_type = ContentType.objects.get(id=module_id)
                temp_report = Report(module=content_type)
                fields = temp_report.get_available_fields()
                choices = [
                    (field["name"], f"{field['verbose_name']}") for field in fields
                ]
            except ContentType.DoesNotExist:
                choices = []

        form.fields["selected_columns"].choices = choices
        form.fields["selected_columns"].widget.choices = choices
        if selected_values:
            form.fields["selected_columns"].widget.value = selected_values
        return super().form_invalid(form)

    def get(self, request, *args, **kwargs):
        report_id = self.kwargs.get("pk")
        if request.user.has_perm("reports.change_report") or request.user.has_perm(
            "reports.add_report"
        ):
            return super().get(request, *args, **kwargs)

        if report_id:
            report = get_object_or_404(Report, pk=report_id)
            if report.report_owner == request.user:
                return super().get(request, *args, **kwargs)

        return render(request, "error/403.html")


@method_decorator(htmx_required, name="dispatch")
class UpdateReportView(LoginRequiredMixin, HorillaSingleFormView):
    model = Report
    fields = ["name"]
    modal_height = False
    full_width_fields = ["name"]

    @cached_property
    def form_url(self):
        pk = self.kwargs.get("pk") or self.request.GET.get("id")
        if pk:
            return reverse_lazy("horilla_reports:update_report", kwargs={"pk": pk})

    def form_valid(self, form):
        response = super().form_valid(form)
        report_pk = self.object.pk
        report_detail_url = reverse_lazy(
            "horilla_reports:report_detail", kwargs={"pk": report_pk}
        )
        return HttpResponse(
            f'<div id="htmx-trigger" '
            f'hx-get="{report_detail_url}" '
            'hx-target="#mainContent" '
            'hx-swap="outerHTML" '
            'hx-push-url="true" '
            'hx-select="#mainContent" '
            'hx-trigger="load" '
            'hx-on::after-request="closeModal();">'
            "</div>"
        )

    def get(self, request, *args, **kwargs):
        report_id = self.kwargs.get("pk")
        if request.user.has_perm(
            "horilla_reports.change_report"
        ) or request.user.has_perm("horilla_reports.add_report"):
            return super().get(request, *args, **kwargs)

        if report_id:
            try:
                report = get_object_or_404(Report, pk=report_id)
            except Http404:
                messages.error(
                    request,
                    f"{self.model._meta.verbose_name.title()} not found or no longer exists.",
                )
                return HttpResponse(
                    "<script>$('#reloadButton').click();closeModal();</script>"
                )
            if report.report_owner == request.user:
                return super().get(request, *args, **kwargs)

        return render(request, "error/403.html")


@method_decorator(htmx_required, name="dispatch")
class MoveReportView(LoginRequiredMixin, HorillaSingleFormView):
    model = Report
    fields = ["folder"]
    modal_height = False
    full_width_fields = ["folder"]

    @cached_property
    def form_url(self):
        pk = self.kwargs.get("pk") or self.request.GET.get("id")
        if pk:
            return reverse_lazy(
                "horilla_reports:move_report_to_folder", kwargs={"pk": pk}
            )

    def get(self, request, *args, **kwargs):
        report_id = self.kwargs.get("pk")
        if request.user.has_perm(
            "horilla_reports.change_report"
        ) or request.user.has_perm("horilla_reports.add_report"):
            return super().get(request, *args, **kwargs)

        if report_id:
            try:
                report = get_object_or_404(Report, pk=report_id)
            except Http404:
                messages.error(
                    request,
                    f"{self.model._meta.verbose_name.title()} not found or no longer exists.",
                )
                return HttpResponse(
                    "<script>$('#reloadButton').click();closeModal();</script>"
                )
            if report.report_owner == request.user:
                return super().get(request, *args, **kwargs)

        return render(request, "error/403.html")

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        user = getattr(self.request, "user", None)
        if user:
            form.fields["folder"].widget.attrs.update(
                {
                    "class": "js-example-basic-single",
                }
            )
            if not user.is_superuser:
                form.fields["folder"].queryset = ReportFolder.objects.filter(
                    report_folder_owner=user
                )
        return form


@method_decorator(htmx_required, name="dispatch")
class MoveFolderView(LoginRequiredMixin, HorillaSingleFormView):
    model = ReportFolder
    fields = ["parent"]
    modal_height = False
    full_width_fields = ["parent"]

    @cached_property
    def form_url(self):
        pk = self.kwargs.get("pk") or self.request.GET.get("id")
        if pk:
            return reverse_lazy(
                "horilla_reports:move_folder_to_folder", kwargs={"pk": pk}
            )

    def get(self, request, *args, **kwargs):
        folder_id = self.kwargs.get("pk")
        if request.user.has_perm(
            "horilla_reports.change_report"
        ) or request.user.has_perm("horilla_reports.add_report"):
            return super().get(request, *args, **kwargs)

        if folder_id:
            try:
                folder = get_object_or_404(ReportFolder, pk=folder_id)
            except Http404:
                messages.error(
                    request,
                    f"{self.model._meta.verbose_name.title()} not found or no longer exists.",
                )
                return HttpResponse(
                    "<script>$('#reloadButton').click();closeModal();</script>"
                )
            if folder.report_folder_owner == request.user:
                return super().get(request, *args, **kwargs)

        return render(request, "error/403.html")

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        user = getattr(self.request, "user", None)
        if user:
            form.fields["parent"].widget.attrs.update(
                {
                    "class": "js-example-basic-single",
                }
            )
            if not user.is_superuser:
                form.fields["parent"].queryset = ReportFolder.objects.filter(
                    report_folder_owner=user
                )
        return form


@method_decorator(htmx_required, name="dispatch")
class GetModuleColumnsHTMXView(LoginRequiredMixin, View):
    """HTMX view to return updated selected_columns field based on module selection"""

    def get(self, request, *args, **kwargs):
        module_id = request.GET.get("module")

        widget_html = self.get_columns_widget_html(module_id)

        return HttpResponse(widget_html)

    def get_columns_widget_html(self, module_id):
        """Generate HTML for the select widget with choices based on module"""
        choices = []

        if module_id:
            try:
                content_type = ContentType.objects.get(id=module_id)
                temp_report = Report(module=content_type)
                fields = temp_report.get_available_fields()

                choices = [
                    (field["name"], f"{field['verbose_name']}") for field in fields
                ]
            except ContentType.DoesNotExist:
                choices = []

        widget = forms.SelectMultiple(
            attrs={
                "class": "js-example-basic-multiple headselect w-full",
                "id": "id_columns",
                "name": "selected_columns",
                "tabindex": "-1",
                "aria-hidden": "true",
                "multiple": True,
            }
        )

        field = forms.MultipleChoiceField(
            choices=choices, widget=widget, required=False
        )
        return field.widget.render("selected_columns", None, attrs=widget.attrs)


@method_decorator(htmx_required, name="dispatch")
class CreateFolderView(LoginRequiredMixin, HorillaSingleFormView):
    model = ReportFolder
    fields = ["name", "parent", "report_folder_owner"]
    modal_height = False
    full_width_fields = ["name", "parent", "report_folder_owner"]
    hidden_fields = ["parent", "report_folder_owner"]

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if self.kwargs.get("pk"):
            form.fields = {k: v for k, v in form.fields.items() if k in ["name"]}
        return form

    def get_initial(self):
        initial = super().get_initial()
        pk = self.request.GET.get("pk")
        initial["parent"] = pk if pk else None
        initial["report_folder_owner"] = self.request.user
        return initial

    @cached_property
    def form_url(self):
        pk = self.kwargs.get("pk")
        if pk:
            return reverse_lazy("horilla_reports:update_folder", kwargs={"pk": pk})
        return reverse_lazy("horilla_reports:create_folder")

    def get(self, request, *args, **kwargs):
        folder_id = self.kwargs.get("pk")
        if request.user.has_perm(
            "horilla_reports.change_report"
        ) or request.user.has_perm("horilla_reports.add_report"):
            return super().get(request, *args, **kwargs)

        if folder_id:
            try:
                folder = get_object_or_404(ReportFolder, pk=folder_id)
            except Http404:
                messages.error(
                    request,
                    f"{self.model._meta.verbose_name.title()} not found or no longer exists.",
                )
                return HttpResponse(
                    "<script>$('#reloadButton').click();closeModal();</script>"
                )

            if folder.report_folder_owner == request.user:
                return super().get(request, *args, **kwargs)

        return render(request, "error/403.html")


@method_decorator(
    permission_required_or_denied(
        ["horilla_reports.view_reportfolder", "horilla_reports.view_own_reportfolder"]
    ),
    name="dispatch",
)
class ReportFolderListView(LoginRequiredMixin, HorillaListView):
    template_name = "report_folder_detail.html"
    model = ReportFolder
    view_id = "folder-list-view"
    table_width = False
    sorting_target = f"#tableview-{view_id}"

    columns = ["name"]

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.filter(parent=None)
        return queryset

    @cached_property
    def col_attrs(self):
        query_params = {}
        if "section" in self.request.GET:
            query_params["section"] = self.request.GET.get("section")
        query_string = urlencode(query_params)
        attrs = {}
        if self.request.user.has_perm(
            "horilla_reports.view_reportfolder"
        ) or self.request.user.has_perm("horilla_reports.view_own_reportfolder"):
            attrs = {
                "hx-get": f"{{get_detail_view_url}}?{query_string}",
                "hx-target": "#mainContent",
                "hx-swap": "outerHTML",
                "hx-select": "#mainContent",
                "hx-push-url": "true",
                "style": "cursor:pointer",
                "class": "hover:text-primary-600",
            }
        return [
            {
                "name": {
                    **attrs,
                }
            }
        ]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Folders"
        return context

    @cached_property
    def action_method(self):
        action_method = ""
        if self.request.user.has_perm(
            "horilla_reports.change_report"
        ) or self.request.user.has_perm("horilla_reports.delete_report"):
            action_method = "actions"
        return action_method


@method_decorator(
    permission_required_or_denied(
        ["horilla_reports.view_report", "horilla_reports.view_own_report"]
    ),
    name="dispatch",
)
class FavouriteReportFolderListView(LoginRequiredMixin, HorillaListView):
    template_name = "favourite_folder_list.html"
    model = ReportFolder
    table_width = False
    view_id = "favourite-folder-list-view"
    sorting_target = f"#tableview-{view_id}"

    def action_method(self):
        action_method = ""
        if self.request.user.has_perm(
            "horilla_reports.change_report"
        ) or self.request.user.has_perm("horilla_reports.delete_report"):
            action_method = "actions"
        return action_method

    columns = ["name"]

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.filter(parent=None, is_favourite=True)
        return queryset

    @cached_property
    def col_attrs(self):
        query_params = {}
        if "section" in self.request.GET:
            query_params["section"] = self.request.GET.get("section")
        query_string = urlencode(query_params)
        attrs = {}
        if self.request.user.has_perm(
            "horilla_reports.view_reportfolder"
        ) or self.request.user.has_perm("horilla_reports.view_own_reportfolder"):
            attrs = {
                "hx-get": f"{{get_detail_view_url}}?{query_string}&source=favourites",
                "hx-target": "#mainContent",
                "hx-swap": "outerHTML",
                "hx-select": "#mainContent",
                "hx-push-url": "true",
                "style": "cursor:pointer",
                "class": "hover:text-primary-600",
            }
        return [
            {
                "name": {
                    **attrs,
                }
            }
        ]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Favourite Folders"
        return context


@method_decorator(
    permission_required_or_denied(
        ["horilla_reports.view_report", "horilla_reports.view_own_report"]
    ),
    name="dispatch",
)
class ReportFolderDetailView(LoginRequiredMixin, HorillaListView):
    template_name = "report_folder_detail.html"
    model = ReportFolder
    table_width = False
    view_id = "report-folder-detail-view"
    bulk_select_option = False
    sorting_target = f"#tableview-{view_id}"

    columns = [
        (_("Name"), "name"),
        (_("Type"), "get_item_type"),
    ]

    def action_method(self):
        action_method = ""
        if self.request.user.has_perm(
            "horilla_reports.change_reportfolder"
        ) or self.request.user.has_perm("horilla_reports.delete_report"):
            action_method = "actions_detail"
        return action_method

    def get_queryset(self):
        folder_id = self.kwargs.get("pk")
        from django.db import models

        folders = ReportFolder.objects.filter(parent__id=folder_id).annotate(
            content_type=models.Value("folder", output_field=models.CharField())
        )
        reports = Report.objects.filter(folder__id=folder_id).annotate(
            content_type=models.Value("report", output_field=models.CharField())
        )
        return folders

    def get(self, request, *args, **kwargs):
        folder_id = self.kwargs.get("pk")
        if not self.model.objects.filter(
            report_folder_owner_id=self.request.user, pk=self.kwargs["pk"]
        ).first() and not self.request.user.has_perm("horilla_reports.view_report"):
            return render(self.request, "error/403.html")
        try:
            ReportFolder.objects.get(pk=folder_id)
        except Exception as e:
            if request.headers.get("HX-Request") == "true":
                messages.error(self.request, e)
                return HttpResponse(headers={"HX-Refresh": "true"})
            raise HorillaHttp404(e)

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        folder_id = self.kwargs.get("pk")
        reports = Report.objects.filter(folder__id=folder_id)
        folders = list(context["object_list"])
        reports_list = list(reports)
        title = ReportFolder.objects.filter(id=folder_id).first()
        context["title"] = title.name if title else "All Folders"
        context["pk"] = folder_id

        for folder in folders:
            folder.item_type = "Folder"
            folder.hx_target = "#mainContent"
            folder.hx_swap = "outerHTML"
            folder.hx_select = "#mainContent"
        for report in reports_list:
            report.item_type = "Report"
            report.hx_target = "#mainContent"
            report.hx_swap = "outerHTML"
            report.hx_select = "#mainContent"

        combined = folders + reports_list
        combined.sort(key=lambda x: x.name.lower())
        context["object_list"] = combined
        context["queryset"] = combined

        query_params = QueryDict(mutable=True)
        if "section" in self.request.GET:
            query_params["section"] = self.request.GET.get("section")
        if "source" in self.request.GET:
            query_params["source"] = self.request.GET.get("source")
        query_string = urlencode(query_params)

        context["col_attrs"] = {
            "name": {
                "hx-get": f"{{get_detail_view_url}}?{query_string}",
                "hx-target": "{hx_target}",
                "hx-swap": "{hx_swap}",
                "hx-push-url": "true",
                "hx-select": "{hx_select}",
                "style": "cursor:pointer",
                "class": "hover:text-primary-600",
            },
            "get_item_type": {},
        }

        # Add breadcrumbs
        breadcrumbs = []
        source = self.request.GET.get("source")
        if source == "favourites":
            breadcrumbs.append(
                {
                    "name": "Favourites",
                    "url": f"{reverse('horilla_reports:favourite_folder_list_view')}?{query_string}",
                    "active": False,
                }
            )
        else:
            breadcrumbs.append(
                {
                    "name": "All Folders",
                    "url": f"{reverse('horilla_reports:report_folder_list')}?{query_string}",
                    "active": False,
                }
            )

        # Build dynamic breadcrumbs for parent folders
        current_folder = ReportFolder.objects.filter(id=folder_id).first()
        folder_chain = []
        while current_folder:
            folder_chain.append(
                {
                    "name": current_folder.name,
                    "url": f"{reverse('horilla_reports:report_folder_detail', kwargs={'pk': current_folder.id})}?{query_string}",
                    "active": current_folder.id == folder_id,
                }
            )
            current_folder = current_folder.parent

        # Reverse parent breadcrumbs to correct order
        folder_chain.reverse()

        # Combine base + folder chain
        breadcrumbs.extend(folder_chain)

        context["breadcrumbs"] = breadcrumbs
        # current_folder = ReportFolder.objects.filter(id=folder_id).first()
        # while current_folder:
        #     breadcrumbs.append(
        #         {
        #             "name": current_folder.name,
        #             "url": f"{reverse('horilla_reports:report_folder_detail', kwargs={'pk': current_folder.id})}?{query_string}",
        #             "active": current_folder.id == folder_id,
        #         }
        #     )
        #     current_folder = current_folder.parent

        # breadcrumbs.reverse()
        # context["breadcrumbs"] = breadcrumbs
        return context


@method_decorator(htmx_required, name="dispatch")
class MarkFolderAsFavouriteView(LoginRequiredMixin, View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        folder = get_object_or_404(ReportFolder, pk=pk)
        user = request.user
        if (
            user.is_superuser
            or user.has_perm("horilla_reports.change_report")
            or folder.report_folder_owner == user
        ):
            folder.is_favourite = not folder.is_favourite
            folder.save(update_fields=["is_favourite"])

        return HttpResponse("<script>$('#reloadButton').click();</script>")

    def get(self, request, *args, **kwargs):
        return render(request, "error/403.html")


@method_decorator(htmx_required, name="dispatch")
class MarkReportAsFavouriteView(LoginRequiredMixin, View):
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        report = get_object_or_404(Report, pk=pk)
        user = request.user
        if (
            user.is_superuser
            or user.has_perm("horilla_reports.change_report")
            or report.report_owner == user
        ):
            report.is_favourite = not report.is_favourite
            report.save(update_fields=["is_favourite"])

        return HttpResponse("<script>$('#reloadButton').click();</script>")

    def get(self, request, *args, **kwargs):
        return render(request, "error/403.html")


@method_decorator(htmx_required, name="dispatch")
@method_decorator(
    permission_required_or_denied("horilla_reports.delete_report"), name="dispatch"
)
class ReportDeleteView(LoginRequiredMixin, HorillaSingleDeleteView):
    model = Report

    def get_post_delete_response(self):
        return HttpResponse("<script>htmx.trigger('#reloadButton','click');</script>")


@method_decorator(htmx_required, name="dispatch")
@method_decorator(
    permission_required_or_denied("horilla_reports.delete_reportfolder"),
    name="dispatch",
)
class FolderDeleteView(LoginRequiredMixin, HorillaSingleDeleteView):
    model = ReportFolder

    def get_post_delete_response(self):
        return HttpResponse("<script>htmx.trigger('#reloadButton','click');</script>")


@method_decorator(
    permission_required_or_denied("horilla_reports.change_report"), name="dispatch"
)
class ReportExportView(LoginRequiredMixin, View):
    """
    Export pivot table data in various formats: Excel, CSV, PDF
    """

    def get(self, request, pk):
        try:
            report = get_object_or_404(Report, pk=pk)
        except Exception as e:
            if request.headers.get("HX-Request") == "true":
                messages.error(self.request, e)
                return HttpResponse(headers={"HX-Refresh": "true"})
            raise HorillaHttp404(e)
        export_format = request.GET.get("format", "excel")

        session_key = f"report_preview_{report.pk}"
        preview_data = request.session.get(session_key, {})
        temp_report = self.create_temp_report(report, preview_data)

        df, context = self.get_report_data(temp_report, request)

        detail_view = ReportDetailView()
        detail_view.request = request
        detail_view.object = report
        detail_context = detail_view.get_context_data()

        if export_format == "excel":
            return self.export_excel(report, df, detail_context, temp_report)
        elif export_format == "csv":
            return self.export_csv(report, df, detail_context, temp_report)
        elif export_format == "pdf":
            return self.export_pdf(report, df, detail_context, temp_report)
        else:
            return self.export_excel(report, df, detail_context, temp_report)

    def create_temp_report(self, original_report, preview_data):
        """Create temporary report with preview data - same as ReportDetailView"""
        temp_report = copy.copy(original_report)
        if "selected_columns" in preview_data:
            temp_report.selected_columns = preview_data["selected_columns"]
        if "row_groups" in preview_data:
            temp_report.row_groups = preview_data["row_groups"]
        if "column_groups" in preview_data:
            temp_report.column_groups = preview_data["column_groups"]
        if "aggregate_columns" in preview_data:
            temp_report.aggregate_columns = preview_data["aggregate_columns"]
        if "filters" in preview_data:
            temp_report.filters = preview_data["filters"]
        return temp_report

    def get_report_data(self, temp_report, request):
        """Get processed report data - simplified version of ReportDetailView logic"""
        model_class = temp_report.model_class
        queryset = model_class.objects.all()

        # Apply filters
        filters = temp_report.filters_dict
        if filters:
            query = None
            for index, (field_name, filter_data) in enumerate(filters.items()):
                if not filter_data.get("value"):
                    continue

                operator = filter_data.get("operator", "exact")
                value = filter_data.get("value")
                logic = filter_data.get("logic", "and") if index > 0 else "and"
                actual_field = filter_data.get("original_field", field_name)

                filter_kwargs = {}
                if operator == "exact":
                    filter_kwargs[f"{actual_field}"] = value
                elif operator == "icontains":
                    filter_kwargs[f"{actual_field}__icontains"] = value
                elif operator == "gt":
                    filter_kwargs[f"{actual_field}__gt"] = value
                elif operator == "lt":
                    filter_kwargs[f"{actual_field}__lt"] = value
                elif operator == "gte":
                    filter_kwargs[f"{actual_field}__gte"] = value
                elif operator == "lte":
                    filter_kwargs[f"{actual_field}__lte"] = value

                if filter_kwargs:
                    current_query = Q(**filter_kwargs)
                    if query is None:
                        query = current_query
                    elif logic == "or":
                        query |= current_query
                    else:
                        query &= current_query

            if query:
                queryset = queryset.filter(query)

        # Get fields and convert to DataFrame
        fields = []
        aggregate_columns_dict = temp_report.aggregate_columns_dict
        if not isinstance(aggregate_columns_dict, list):
            aggregate_columns_dict = (
                [aggregate_columns_dict] if aggregate_columns_dict else []
            )

        if temp_report.selected_columns_list:
            fields.extend(temp_report.selected_columns_list)
        if temp_report.row_groups_list:
            fields.extend(temp_report.row_groups_list)
        if temp_report.column_groups_list:
            fields.extend(temp_report.column_groups_list)

        for agg in aggregate_columns_dict:
            if agg.get("field"):
                fields.append(agg["field"])

        fields = list(dict.fromkeys(fields))
        data = list(queryset.values(*fields)) if fields else list(queryset.values())
        df = pd.DataFrame(data)

        # Create context for export
        context = {
            "total_count": len(data),
            "configuration_type": self.get_configuration_type(temp_report),
            "aggregate_columns_dict": aggregate_columns_dict,
        }

        return df, context

    def get_configuration_type(self, report):
        row_count = len(report.row_groups_list)
        col_count = len(report.column_groups_list)
        return f"{row_count}_row_{col_count}_col"

    def get_display_value(self, value, field_name, model_class):
        """Get display value for field - same logic as ReportDetailView"""
        try:
            field = model_class._meta.get_field(field_name)
            if hasattr(field, "related_model") and field.related_model:
                try:
                    related_obj = field.related_model.objects.get(pk=value)
                    return str(related_obj)
                except field.related_model.DoesNotExist:
                    return f"Unknown ({value})"

            if hasattr(field, "choices") and field.choices:
                choice_dict = dict(field.choices)
                return choice_dict.get(value, value)

            if value is None or value == "":
                return "Unspecified (-)"

            return str(value)
        except:
            return str(value) if value is not None else "Unspecified (-)"

    def get_verbose_name(self, field_name, model_class):
        """Get the verbose name of a field"""
        try:
            return model_class._meta.get_field(field_name).verbose_name.title()
        except:
            return field_name.title()

    def _create_pivot_sheet(self, ws, df, detail_context, temp_report):
        """Create pivot table sheet that matches the web detail view exactly"""
        pivot_table = detail_context.get("pivot_table", {})
        pivot_index = detail_context.get("pivot_index", [])
        pivot_columns = detail_context.get("pivot_columns", [])

        if not pivot_table or not pivot_index:
            ws["A1"] = "No pivot table data available"
            return

        # Create two-row header structure like your web interface
        # Row 1: Group headers (Campaign, Email, etc.)
        # Row 2: Campaign type headers (Education, Finance, etc.)

        ws.cell(row=1, column=1, value="Lead Status")
        ws.cell(row=2, column=1, value="Lead Status")

        # Merge the Lead Status cells
        ws.merge_cells("A1:A2")
        ws["A1"].font = Font(bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Parse column structure to create grouped headers
        current_col = 2
        group_starts = {}  # Track where each group starts for merging

        # First pass: identify groups and their ranges
        current_group = None
        group_start_col = 2

        for col_name in pivot_columns:
            if "|" in col_name:
                group_name, campaign_type = col_name.split("|", 1)
            else:
                group_name = "Other"
                campaign_type = col_name

            # If this is a new group, record the previous group's range
            if current_group != group_name:
                if current_group is not None:
                    group_starts[current_group] = (group_start_col, current_col - 1)
                current_group = group_name
                group_start_col = current_col

            current_col += 1

        # Record the last group
        if current_group is not None:
            group_starts[current_group] = (group_start_col, current_col - 1)

        # Create group headers (row 1) and campaign headers (row 2)
        current_col = 2
        for col_name in pivot_columns:
            if "|" in col_name:
                group_name, campaign_type = col_name.split("|", 1)
            else:
                group_name = "Other"
                campaign_type = col_name

            # Set group header (row 1)
            cell1 = ws.cell(row=1, column=current_col, value=group_name)
            cell1.font = Font(bold=True)
            cell1.alignment = Alignment(horizontal="center")

            # Set campaign type header (row 2)
            cell2 = ws.cell(row=2, column=current_col, value=campaign_type)
            cell2.font = Font(bold=True)
            cell2.alignment = Alignment(horizontal="center")

            current_col += 1

        # Merge cells for group headers
        for group_name, (start_col, end_col) in group_starts.items():
            if start_col < end_col:  # Only merge if there are multiple columns
                start_cell = openpyxl.utils.get_column_letter(start_col)
                end_cell = openpyxl.utils.get_column_letter(end_col)
                ws.merge_cells(f"{start_cell}1:{end_cell}1")

        # Data rows starting from row 3
        lead_statuses = ["New", "Contacted", "Qualified", "Proposal", "Lost"]

        for row_idx, status in enumerate(lead_statuses, 3):
            # Lead Status column
            ws.cell(row=row_idx, column=1, value=status)

            # Data values for each campaign
            for col_idx, col_name in enumerate(pivot_columns, 2):
                if status in pivot_table and col_name in pivot_table[status]:
                    value = pivot_table[status][col_name]
                else:
                    value = 0
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Total row
        total_row = len(lead_statuses) + 3
        total_cell = ws.cell(row=total_row, column=1, value="Total")
        total_cell.font = Font(bold=True)

        # Calculate totals for each column
        for col_idx, col_name in enumerate(pivot_columns, 2):
            total_value = 0
            for status in lead_statuses:
                if status in pivot_table and col_name in pivot_table[status]:
                    total_value += pivot_table[status][col_name]

            cell = ws.cell(row=total_row, column=col_idx, value=total_value)
            cell.font = Font(bold=True)

        # Apply styling and borders
        max_row = total_row
        max_col = len(pivot_columns) + 1

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Apply borders to all cells
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

                # Center align numeric data
                if row > 2 and col > 1:
                    cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for col_idx in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_length = 0

            for row_idx in range(1, max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            # Set appropriate column width
            if col_idx == 1:  # Lead Status column
                adjusted_width = max(max_length + 2, 12)
            else:  # Campaign columns
                adjusted_width = max(max_length + 2, 10)

            ws.column_dimensions[col_letter].width = min(adjusted_width, 20)

        # Add header row background colors
        for col_idx in range(1, max_col + 1):
            # Group header row (row 1)
            cell1 = ws.cell(row=1, column=col_idx)
            cell1.fill = PatternFill(
                start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"
            )

            # Campaign header row (row 2)
            cell2 = ws.cell(row=2, column=col_idx)
            cell2.fill = PatternFill(
                start_color="E8F4FD", end_color="E8F4FD", fill_type="solid"
            )

    def export_excel(self, report, df, detail_context, temp_report):
        """Export pivot table as Excel file"""
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{report.name}_pivot.xlsx"'
        )

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pivot Table"

        # Write pivot table data
        self._create_pivot_sheet(ws, df, detail_context, temp_report)

        # Add metadata sheet
        meta_ws = wb.create_sheet("Report Info")
        meta_ws["A1"] = "Report Name"
        meta_ws["B1"] = report.name
        meta_ws["A2"] = "Export Date"
        meta_ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        meta_ws["A3"] = "Total Records"
        meta_ws["B3"] = detail_context.get("total_count", 0)

        # Save to response
        wb.save(response)
        return response

    def export_csv(self, report, df, detail_context, temp_report):
        """Export pivot table as CSV"""
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="{report.name}_pivot.csv"'
        )

        writer = csv.writer(response)
        pivot_table = detail_context.get("pivot_table", {})
        pivot_index = detail_context.get("pivot_index", [])
        pivot_columns = detail_context.get("pivot_columns", [])

        if not pivot_table or not pivot_index:
            writer.writerow(["No pivot table data available"])
            return response

        # Create two-row header structure for CSV
        # First row: Group names
        group_header = ["Lead Status"]
        campaign_header = ["Lead Status"]

        current_group = None
        group_size = 0

        for col_name in pivot_columns:
            if "|" in col_name:
                group_name, campaign_type = col_name.split("|", 1)
            else:
                group_name = "Other"
                campaign_type = col_name

            if current_group != group_name:
                current_group = group_name
                group_size = 1
            else:
                group_size += 1

            group_header.append(group_name)
            campaign_header.append(campaign_type)

        # Write both header rows
        writer.writerow(group_header)
        writer.writerow(campaign_header)

        # Write data rows
        lead_statuses = ["New", "Contacted", "Qualified", "Proposal", "Lost"]

        for status in lead_statuses:
            row = [status]
            for col_name in pivot_columns:
                value = 0
                if status in pivot_table and col_name in pivot_table[status]:
                    value = pivot_table[status][col_name]
                row.append(value)
            writer.writerow(row)

        # Add totals row
        total_row = ["Total"]
        for col_name in pivot_columns:
            total_value = 0
            for status in lead_statuses:
                if status in pivot_table and col_name in pivot_table[status]:
                    total_value += pivot_table[status][col_name]
            total_row.append(total_value)
        writer.writerow(total_row)

        return response

    def export_pdf(self, report, df, detail_context, temp_report):
        """Export pivot table as PDF"""
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="{report.name}_pivot.pdf"'
        )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1 * inch)
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=16,
            spaceAfter=30,
            alignment=1,  # Center
        )

        # Title
        elements.append(Paragraph(f"Lead Report: {report.name}", title_style))
        elements.append(Spacer(1, 12))

        pivot_table = detail_context.get("pivot_table", {})
        pivot_index = detail_context.get("pivot_index", [])
        pivot_columns = detail_context.get("pivot_columns", [])

        if not pivot_table or not pivot_index:
            elements.append(
                Paragraph("No pivot table data available", styles["Normal"])
            )
        else:
            # Clean column names
            clean_columns = []
            for col_name in pivot_columns:
                if "|" in col_name:
                    clean_name = col_name.split("|")[-1]
                else:
                    clean_name = col_name
                clean_columns.append(clean_name)

            # Prepare data with headers
            data_rows = [["Lead Status"] + clean_columns]

            # Add data rows
            lead_statuses = ["New", "Contacted", "Qualified", "Proposal", "Lost"]

            for status in lead_statuses:
                row = [status]
                for col_name in pivot_columns:
                    value = 0
                    if status in pivot_table and col_name in pivot_table[status]:
                        value = pivot_table[status][col_name]
                    row.append(str(value))
                data_rows.append(row)

            # Add totals row
            total_row = ["Total"]
            for col_name in pivot_columns:
                total_value = 0
                for status in lead_statuses:
                    if status in pivot_table and col_name in pivot_table[status]:
                        total_value += pivot_table[status][col_name]
                total_row.append(str(total_value))
            data_rows.append(total_row)

            # Create table with proper styling
            data_table = Table(data_rows)
            data_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        # Bold total row
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                        ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
                    ]
                )
            )

            elements.append(data_table)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        response.write(buffer.getvalue())
        buffer.close()
        return response
